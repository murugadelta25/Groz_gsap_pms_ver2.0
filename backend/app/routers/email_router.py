from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from sqlalchemy.orm import Session
from datetime import datetime
import pytz as _pytz
_IST = _pytz.timezone('Asia/Kolkata')
def _now_ist(): return datetime.now(_IST).replace(tzinfo=None)
from typing import Optional, List
from pydantic import BaseModel
import smtplib, io, csv
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from ..models import (EmailGroup, EmailRecipient, EmailSchedule,
                      EmailSmtpConfig, OEEEntry, ProductionPlan,
                      BreakdownTicket, EmailLog, get_db)
from ..auth import get_current_user, require_role

router = APIRouter(prefix="/api/email", tags=["email"])

# ── Pydantic schemas ──────────────────────────────────────────────────────────

class SmtpConfigIn(BaseModel):
    smtp_server: str = "smtp.gmail.com"
    smtp_port: int = 587
    email_address: str
    email_password: str

class RecipientIn(BaseModel):
    group_id: int
    name: str
    email: str
    active: int = 1

class ScheduleIn(BaseModel):
    name: str
    group_ids: str          # "1,2,3"
    report_type: str = "daily"
    send_hour: int = 18
    send_minute: int = 0
    attach_report: int = 1
    active: int = 1

REPORT_DEFAULTS = {
    "management": "oee,planning,breakdown,maintenance,tools,deviation_alerts",
    "production":  "oee,planning,breakdown,tools,deviation_alerts",
    "maintenance": "breakdown,maintenance,tools,deviation_alerts",
}
ALL_REPORTS = ["oee", "planning", "breakdown", "maintenance", "data_entry", "loss_tracker", "tools", "deviation_alerts"]

class GroupIn(BaseModel):
    name: str
    description: str = ""
    report_types: str = "oee,planning,breakdown"

class ManualSendIn(BaseModel):
    group_ids: List[int]
    subject: str
    body: str
    attach_report: bool = False
    report_type: str = "daily"
    report_date: Optional[str] = None  # YYYY-MM-DD, defaults to yesterday

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_smtp(db: Session):
    cfg = db.query(EmailSmtpConfig).first()
    if not cfg or not cfg.email_address or not cfg.email_password:
        raise HTTPException(400, "SMTP not configured. Go to Alerts > Email Settings.")
    return cfg

def get_group_emails(db: Session, group_ids: List[int]) -> List[str]:
    recipients = db.query(EmailRecipient).filter(
        EmailRecipient.group_id.in_(group_ids),
        EmailRecipient.active == 1
    ).all()
    return [r.email for r in recipients]

def build_oee_xlsx(db: Session, report_date=None) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from ..models import OEEDefectLog, User
    from datetime import date as date_type, timedelta
    from .oee import collect_merged_oee_report_rows

    if report_date is None:
        report_date = date_type.today() - timedelta(days=1)
    elif isinstance(report_date, str):
        report_date = date_type.fromisoformat(report_date)

    entries = collect_merged_oee_report_rows(
        db,
        entry_date=report_date,
        prefer_live=True,
    )
    user_map = {u.id: u.username for u in db.query(User).all()}

    def fmt_ist(dt_val):
        if not dt_val: return ''
        return dt_val.strftime('%d-%m-%Y %H:%M:%S IST')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OEE Report"

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    red_font  = Font(bold=True, color="DC2626")
    grn_font  = Font(bold=True, color="059669")
    amb_font  = Font(bold=True, color="D97706")

    headers = ["Date","Station","Machine","Shift","Work Order","Model / Variant",
               "Current Operation","Next Operation","CT (sec)",
               "Avail (min)","Op Time (min)","Plan Qty","Possible Qty","Actual Qty",
               "Prod Loss","Accepted Qty","Defect Qty",
               "AR%","PR%","QR%","OEE%",
               "AR% (original)","PR% (original)","QR% (original)","OEE% (original)",
               "Source"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    if not entries:
        ws.append([f"No live or data-entry OEE records found for {report_date}"])

    for e in entries:
        ct = (e.process_time or 0) + (e.loading_unloading or 0)
        prod_loss = max(0, (e.possible_qty or 0) - (e.actual_qty or 0))
        ar_val = float(e.ar or 0)
        pr_val = float(e.pr or 0)
        qr_val = float(e.qr or 0)
        oee_val = float(e.oee or 0)
        ar_raw = float(e.ar_raw or 0) if e.ar_raw is not None else None
        pr_raw = float(e.pr_raw or 0) if e.pr_raw is not None else None
        qr_raw = float(e.qr_raw or 0) if e.qr_raw is not None else None
        oee_raw = float(e.oee_raw or 0) if e.oee_raw is not None else None
        plan_qty = e.planned_qty if e.planned_qty is not None else ""
        ws.append([
            str(e.entry_date), e.station_name or str(e.station_no or ""),
            e.machine_name or "",
            e.shift, e.work_order_no or "", e.model_variant or "",
            e.current_operation, e.next_operation,
            ct, e.available_shift_time, e.operating_time,
            plan_qty, e.possible_qty, e.actual_qty, prod_loss,
            e.accp_qty, e.defect_qty,
            ar_val, pr_val, qr_val, oee_val,
            ar_raw if ar_raw is not None else "—",
            pr_raw if pr_raw is not None else "—",
            qr_raw if qr_raw is not None else "—",
            oee_raw if oee_raw is not None else "—",
            "Live" if e.source == "realtime" else "Data Entry",
        ])
        row_idx = ws.max_row
        ws.cell(row_idx, 21).font = grn_font if oee_val >= 85 else (amb_font if oee_val >= 65 else red_font)
        for col, raw in [(18, ar_raw), (19, pr_raw), (20, qr_raw), (21, oee_raw)]:
            if raw is not None:
                ws.cell(row_idx, col).font = amb_font

    col_widths = [12,14,12,8,18,16,14,14,10,12,14,10,12,12,10,12,12,8,8,8,8,14,14,14,14,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # ── QC Logs sheet ────────────────────────────────────────────────────────
    entry_ids = [e.id for e in entries if e.id]
    qc_logs = []
    if entry_ids:
        qc_logs = db.query(OEEDefectLog)\
                    .filter(OEEDefectLog.oee_entry_id.in_(entry_ids))\
                    .order_by(OEEDefectLog.updated_at.desc()).all()
    entry_map = {e.id: e for e in entries if e.id}

    ws_qc = wb.create_sheet("QC Logs")
    qc_headers = [
        "Current Operation", "Next Operation", "Entry Date", "Shift", "Station",
        "Updated At (IST)", "Updated By",
        "Before Defect", "Before Accp", "Before QR%", "Before OEE%",
        "After Defect",  "After Accp",  "After QR%",  "After OEE%",
        "Note"
    ]
    ws_qc.append(qc_headers)
    for col in range(1, len(qc_headers) + 1):
        c = ws_qc.cell(1, col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    for l in qc_logs:
        e = entry_map.get(l.oee_entry_id)
        before_oee = float(l.before_oee or 0)
        after_oee  = float(l.after_oee  or 0)
        ws_qc.append([
            e.current_operation if e else "",
            e.next_operation if e else "",
            str(e.entry_date) if e else "",
            e.shift if e else "",
            (e.station_name if e else "") or "",
            fmt_ist(l.updated_at),
            user_map.get(l.updated_by, str(l.updated_by) if l.updated_by else ""),
            l.before_defect_qty, l.before_accp_qty,
            f"{float(l.before_qr or 0):.2f}%", f"{before_oee:.2f}%",
            l.after_defect_qty,  l.after_accp_qty,
            f"{float(l.after_qr  or 0):.2f}%", f"{after_oee:.2f}%",
            l.note or "",
        ])
        row_i = ws_qc.max_row
        bf = ws_qc.cell(row_i, 11)
        bf.font = grn_font if before_oee >= 85 else (amb_font if before_oee >= 65 else red_font)
        af = ws_qc.cell(row_i, 15)
        af.font = grn_font if after_oee  >= 85 else (amb_font if after_oee  >= 65 else red_font)

    qc_widths = [18,18,12,8,14, 22,14, 14,12,12,12, 12,12,12,12, 35]
    for i, w in enumerate(qc_widths, 1):
        ws_qc.column_dimensions[ws_qc.cell(1, i).column_letter].width = w
    ws_qc.freeze_panes = "A2"

    # ── OEE Daywise sheet ──
    from collections import defaultdict
    day_groups = defaultdict(list)
    for e in entries:
        day_groups[(str(e.entry_date), e.station_no)].append(e)

    ws_day = wb.create_sheet("OEE Daywise")
    day_hdrs = ["Date","Station","AR%","PR%","QR%","OEE%","Total Produced","Accepted Qty","Defects"]
    ws_day.append(day_hdrs)
    for col in range(1, len(day_hdrs)+1):
        c = ws_day.cell(1, col); c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    ws_day.freeze_panes = "A2"
    for (dt, pno), grp in sorted(day_groups.items()):
        n = len(grp)
        avg_ar  = round(sum(float(e.ar  or 0) for e in grp) / n, 2)
        avg_pr  = round(sum(float(e.pr  or 0) for e in grp) / n, 2)
        avg_qr  = round(sum(float(e.qr  or 0) for e in grp) / n, 2)
        avg_oee = round(sum(float(e.oee or 0) for e in grp) / n, 2)
        station_label = grp[0].station_name if grp else str(pno)
        ws_day.append([dt, station_label,
                       avg_ar, avg_pr, avg_qr, avg_oee,
                       sum(e.actual_qty or 0 for e in grp),
                       sum(e.accp_qty   or 0 for e in grp),
                       sum(e.defect_qty or 0 for e in grp)])
        ws_day.cell(ws_day.max_row, 6).font = grn_font if avg_oee>=85 else (amb_font if avg_oee>=65 else red_font)
    for i, w in enumerate([12,14,8,8,8,8,14,14,10], 1):
        ws_day.column_dimensions[ws_day.cell(1,i).column_letter].width = w

    # ── OEE Shiftwise sheet ──
    shift_groups = defaultdict(list)
    for e in entries:
        shift_groups[(str(e.entry_date), e.station_no, e.shift)].append(e)

    ws_shift = wb.create_sheet("OEE Shiftwise")
    shift_hdrs = ["Date","Station","Shift","Actual","Prod Loss","Accepted","Defects","AR%","PR%","QR%","OEE%"]
    ws_shift.append(shift_hdrs)
    for col in range(1, len(shift_hdrs)+1):
        c = ws_shift.cell(1, col); c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    ws_shift.freeze_panes = "A2"
    for (dt, pno, sh), grp in sorted(shift_groups.items()):
        n = len(grp)
        avg_ar  = round(sum(float(e.ar  or 0) for e in grp) / n, 2)
        avg_pr  = round(sum(float(e.pr  or 0) for e in grp) / n, 2)
        avg_qr  = round(sum(float(e.qr  or 0) for e in grp) / n, 2)
        avg_oee = round(sum(float(e.oee or 0) for e in grp) / n, 2)
        station_label = grp[0].station_name if grp else str(pno)
        ws_shift.append([dt, station_label, sh,
                         sum(e.actual_qty or 0 for e in grp),
                         sum(max(0,(e.possible_qty or 0)-(e.actual_qty or 0)) for e in grp),
                         sum(e.accp_qty   or 0 for e in grp),
                         sum(e.defect_qty or 0 for e in grp),
                         avg_ar, avg_pr, avg_qr, avg_oee])
        ws_shift.cell(ws_shift.max_row, 11).font = grn_font if avg_oee>=85 else (amb_font if avg_oee>=65 else red_font)
    for i, w in enumerate([12,14,8,10,10,12,10,8,8,8,8], 1):
        ws_shift.column_dimensions[ws_shift.cell(1,i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_plan_xlsx(db: Session) -> bytes:
    from .plans import _build_excel
    from ..models import Station
    plans = db.query(ProductionPlan).order_by(ProductionPlan.plan_date.desc()).all()
    # patch station_no to display_name for the excel builder
    station_map = {p.id: (p.display_name or p.name) for p in db.query(Station).all()}
    for p in plans:
        p._station_label = station_map.get(p.station_no, str(p.station_no))
    return _build_excel(plans).read()

def _build_breakdown_xlsx(tickets, machine_map: dict, user_map: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Breakdown Tickets"

    headers = ["Ticket #", "Raised By", "Acknowledged By", "Machine", "Status",
               "Description", "Resolution Notes",
               "Raised At", "Acknowledged At", "Work Started At", "Resolved At",
               "Duration (mins)"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    def fmt(dt):
        return dt.strftime("%d-%m-%Y %H:%M:%S") if dt else ""

    for t in tickets:
        duration = ""
        if t.created_at and t.resolved_time:
            mins = round((t.resolved_time - t.created_at).total_seconds() / 60, 1)
            duration = mins
        ws.append([
            t.id,
            user_map.get(t.raised_by, str(t.raised_by) if t.raised_by else ""),
            user_map.get(t.acknowledged_by, str(t.acknowledged_by) if t.acknowledged_by else ""),
            machine_map.get(t.machine_id, str(t.machine_id)),
            t.status, t.description or "", t.resolution_notes or "",
            fmt(t.created_at), fmt(t.ack_time),
            fmt(t.start_troubleshoot), fmt(t.resolved_time),
            duration,
        ])

    col_widths = [10, 14, 16, 18, 12, 22, 28, 18, 18, 18, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _build_maintenance_xlsx(tickets, machine_map: dict, user_map: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Maintenance Tickets"

    headers = ["Ticket #", "Machine", "Status", "Raised By", "Acknowledged By",
               "Description", "Serviced By", "Resolution Notes",
               "Raised At", "Acknowledged At", "Work Started At", "Resolved At",
               "Service Duration (ack→resolved)", "Issue Duration (raised→resolved)"]
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    def fmt(dt):
        return dt.strftime("%d-%m-%Y %H:%M:%S") if dt else ""

    def fmt_duration(from_dt, to_dt):
        if not from_dt or not to_dt:
            return ""
        mins = int((to_dt - from_dt).total_seconds() / 60)
        if mins < 1:
            return "<1m"
        if mins < 60:
            return f"{mins}m"
        return f"{mins // 60}h {mins % 60}m"

    import re
    status_labels = {
        "raised": "Raised", "acknowledged": "Acknowledged",
        "in_progress": "In Progress", "resolved": "Resolved"
    }

    for t in tickets:
        notes = t.resolution_notes or ""
        serviced_by_match = re.search(r'\[Serviced by: ([^\]]+)\]', notes)
        serviced_by  = serviced_by_match.group(1) if serviced_by_match else ""
        clean_notes  = re.sub(r'\[Serviced by: [^\]]+\]\s*', '', notes).strip()
        ws.append([
            t.id,
            machine_map.get(t.machine_id, str(t.machine_id)),
            status_labels.get(t.status, t.status),
            user_map.get(t.raised_by, str(t.raised_by) if t.raised_by else ""),
            user_map.get(t.acknowledged_by, str(t.acknowledged_by) if t.acknowledged_by else ""),
            t.description or "",
            serviced_by,
            clean_notes,
            fmt(t.created_at), fmt(t.ack_time),
            fmt(t.start_troubleshoot), fmt(t.resolved_time),
            fmt_duration(t.ack_time, t.resolved_time),
            fmt_duration(t.created_at, t.resolved_time),
        ])

    col_widths = [10, 18, 12, 14, 16, 20, 14, 22, 20, 20, 20, 20, 26, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _get_breakdown_lookups(db: Session):
    from ..models import Machine, User
    machine_map = {m.id: m.name for m in db.query(Machine).all()}
    user_map    = {u.id: u.username for u in db.query(User).all()}
    return machine_map, user_map

def build_breakdown_xlsx(db: Session) -> bytes:
    tickets = db.query(BreakdownTicket).order_by(BreakdownTicket.created_at.desc()).all()
    machine_map, user_map = _get_breakdown_lookups(db)
    return _build_breakdown_xlsx(tickets, machine_map, user_map)

def build_maintenance_xlsx(db: Session) -> bytes:
    tickets = db.query(BreakdownTicket).filter(
        BreakdownTicket.status.in_(["acknowledged", "in_progress", "resolved"])
    ).order_by(BreakdownTicket.created_at.desc()).all()
    machine_map, user_map = _get_breakdown_lookups(db)
    return _build_maintenance_xlsx(tickets, machine_map, user_map)

def build_data_entry_xlsx(db: Session, report_date=None) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import date as date_type, timedelta
    from .oee import collect_merged_oee_report_rows

    # Default: previous day's completed data
    if report_date is None:
        report_date = date_type.today() - timedelta(days=1)
    elif isinstance(report_date, str):
        report_date = date_type.fromisoformat(report_date)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    alt_fill = PatternFill("solid", fgColor="F1F5F9")   # light alternating row
    oee_good  = Font(bold=True, color="059669")
    oee_warn  = Font(bold=True, color="D97706")
    oee_bad   = Font(bold=True, color="DC2626")

    HEADERS = [
        "Date", "Station", "Machine", "Shift", "Work Order", "Model / Variant",
        "Current Operation", "Next Operation",
        "Process Time (s)", "Loading & Unloading (s)", "CT (s)",
        "Start Time", "Stop Time", "Total Minutes",
        "Lunch Break", "Tea Break", "TPM Cleaning", "Other Cleaning", "Mgmt Meeting",
        "Total Breaks", "Shift Working Min",
        "No Load", "New Model Trial", "Power Cut", "Planned Maintenance", "No Manpower",
        "Mgmt Loss Total", "Available Shift Time (min)",
        "Setting Time", "Tool Change", "Dim Correction", "Scrap Removal", "Break Down",
        "Total Down Time", "Operating Time (min)", "Plan Qty", "Possible Qty",
        "Actual Qty", "Production Loss", "Accepted Qty", "Defect Qty",
        "AR%", "PR%", "QR%", "OEE%", "Source",
    ]
    COL_WIDTHS = [
        12,14,12,8,18,16, 14,14, 16,20,10, 10,10,14,
        12,10,14,14,14, 12,18,
        10,16,12,20,14, 14,22,
        12,12,14,14,12, 14,18,10,12,
        12,14,12,12, 8,8,8,8, 12,
    ]
    OEE_COL = len(HEADERS) - 1  # OEE% column (1-based), Source is last

    def entry_row(e):
        ct = (e.process_time or 0) + (e.loading_unloading or 0)
        prod_loss = max(0, (e.possible_qty or 0) - (e.actual_qty or 0))
        plan_qty = e.planned_qty if e.planned_qty is not None else ""
        return [
            str(e.entry_date), e.station_name or str(e.station_no or ""),
            e.machine_name or "", e.shift, e.work_order_no or "", e.model_variant or "",
            e.current_operation or "", e.next_operation or "",
            e.process_time or 0, e.loading_unloading or 0, ct,
            e.start_time or "", e.stop_time or "", e.total_minutes or 0,
            e.lunch_break or 0, e.tea_break or 0, e.tpm_cleaning or 0,
            e.other_cleaning or 0, e.management_meeting or 0,
            e.total_breaks or 0, e.shift_working_minutes or 0,
            e.no_load or 0, e.new_model_trial or 0, e.power_cut or 0,
            e.planned_maintenance or 0, e.no_manpower_planned or 0,
            e.management_loss_total or 0, e.available_shift_time or 0,
            e.setting_time or 0, e.tool_change or 0,
            e.dimension_correction or 0, e.scrap_removal or 0, e.break_down or 0,
            e.total_down_time or 0, e.operating_time or 0, plan_qty, e.possible_qty or 0,
            e.actual_qty or 0, prod_loss, e.accp_qty or 0, e.defect_qty or 0,
            float(e.ar or 0), float(e.pr or 0),
            float(e.qr or 0), float(e.oee or 0),
            "Live" if e.source == "realtime" else "Data Entry",
        ]

    def style_sheet(ws, entries_for_sheet):
        """Write header + rows with alternating fill and OEE colour."""
        ws.append(HEADERS)
        for col in range(1, len(HEADERS) + 1):
            c = ws.cell(1, col)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
        for idx, e in enumerate(entries_for_sheet, start=2):
            ws.append(entry_row(e))
            if idx % 2 == 0:
                for col in range(1, len(HEADERS) + 1):
                    ws.cell(idx, col).fill = alt_fill
            oee_val = float(e.oee or 0)
            oee_cell = ws.cell(idx, OEE_COL)
            oee_cell.font = oee_good if oee_val >= 85 else (oee_warn if oee_val >= 65 else oee_bad)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # ── Per-shift sheets for report_date (live first, else data entry) ────────
    day_entries = collect_merged_oee_report_rows(
        db, entry_date=report_date, prefer_live=True,
    )

    shifts_present = sorted(set(e.shift for e in day_entries if e.shift)) or ["A", "B"]
    for shift in shifts_present:
        ws = wb.create_sheet(title=f"Shift {shift} — {report_date}")
        style_sheet(ws, [e for e in day_entries if e.shift == shift])

    if not day_entries:
        ws = wb.create_sheet(f"No Data — {report_date}")
        ws.append([f"No live or data-entry records found for {report_date}"])

    # ── Monthly Consolidated sheet ────────────────────────────────────────────
    month_start = report_date.replace(day=1)
    monthly_entries = collect_merged_oee_report_rows(
        db,
        date_from=month_start,
        date_to=report_date,
        prefer_live=True,
    )

    month_label = report_date.strftime("%b %Y")
    ws_monthly = wb.create_sheet(title=f"Monthly — {month_label}")
    style_sheet(ws_monthly, monthly_entries)

    # Add a summary row at the bottom of monthly sheet
    if monthly_entries:
        ws_monthly.append([])  # blank separator
        totals = [
            "TOTALS / AVG", "", "", "", "", "",
            "", "", "", "", "",
            "", "", "",
            "", "", "", "", "",
            "", "",
            "", "", "", "", "",
            "", "",
            "", "", "", "", "",
            "", "", "",
            sum(e.possible_qty or 0 for e in monthly_entries),
            sum(e.actual_qty or 0 for e in monthly_entries),
            sum(max(0,(e.possible_qty or 0)-(e.actual_qty or 0)) for e in monthly_entries),
            sum(e.accp_qty or 0 for e in monthly_entries),
            sum(e.defect_qty or 0 for e in monthly_entries),
            round(sum(float(e.ar or 0) for e in monthly_entries)/len(monthly_entries), 2),
            round(sum(float(e.pr or 0) for e in monthly_entries)/len(monthly_entries), 2),
            round(sum(float(e.qr or 0) for e in monthly_entries)/len(monthly_entries), 2),
            round(sum(float(e.oee or 0) for e in monthly_entries)/len(monthly_entries), 2),
            "",
        ]
        ws_monthly.append(totals)
        tr = ws_monthly.max_row
        summary_fill = PatternFill("solid", fgColor="1E3A5F")
        summary_font = Font(bold=True, color="FFFFFF")
        for col in range(1, len(HEADERS) + 1):
            c = ws_monthly.cell(tr, col)
            c.fill = summary_fill; c.font = summary_font

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def build_loss_tracker_xlsx(db: Session, report_date=None) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from ..models import MachineStatusLog, Machine, Station
    from datetime import date as date_type, timedelta, datetime as dt

    # Default: yesterday
    if report_date is None:
        target = date_type.today() - timedelta(days=1)
    elif isinstance(report_date, str):
        target = date_type.fromisoformat(report_date)
    else:
        target = report_date

    # changed_at is stored as IST naive — filter directly against IST day boundaries
    from datetime import timezone, timedelta as td
    day_start = dt.combine(target, dt.min.time())  # IST 00:00
    day_end   = dt.combine(target + timedelta(days=1), dt.min.time())  # IST 00:00 next day

    logs = (db.query(MachineStatusLog)
              .filter(MachineStatusLog.changed_at >= day_start,
                      MachineStatusLog.changed_at < day_end)
              .order_by(MachineStatusLog.machine_id, MachineStatusLog.changed_at.desc())
              .all())

    machine_map = {m.id: m for m in db.query(Machine).all()}
    station_map    = {p.id: (p.display_name or p.name) for p in db.query(Station).all()}

    from ..deviation_alert_service import get_limits_min
    limits_min = get_limits_min(db)
    LIMITS = {k: int(v) * 60 for k, v in limits_min.items()}
    TRACK = set(LIMITS.keys())

    def fmt_ist(dt_val):
        if not dt_val: return ''
        return dt_val.strftime('%d-%m-%Y %H:%M:%S IST')

    def dur_label(sec):
        if sec < 0: sec = 0
        h, rem = divmod(int(sec), 3600)
        m, s   = divmod(rem, 60)
        if h: return f"{h}h {m}m"
        if m: return f"{m}m {s}s"
        return f"{s}s"

    # Build rows with computed duration
    all_rows = []
    # group logs by machine_id to compute durations
    from itertools import groupby
    sorted_logs = sorted(logs, key=lambda l: (l.machine_id, l.changed_at))
    # We need per-machine sorted ascending to compute duration
    machines_logs = {}
    for l in sorted_logs:
        machines_logs.setdefault(l.machine_id, []).append(l)

    now_ist_naive = dt.now(_pytz.timezone('Asia/Kolkata')).replace(tzinfo=None)
    for mid, mlogs in machines_logs.items():
        mach = machine_map.get(mid)
        station_name = station_map.get(mach.station_id, '') if mach else ''
        # mlogs ascending; duration = next_log.changed_at - this.changed_at
        for i, log in enumerate(mlogs):
            if i < len(mlogs) - 1:
                next_log = mlogs[i + 1]
                duration_sec = (next_log.changed_at - log.changed_at).total_seconds()
                end_time = next_log.changed_at
                is_ongoing = False
            else:
                duration_sec = (now_ist_naive - log.changed_at).total_seconds()
                end_time = None
                is_ongoing = True
            limit_sec = LIMITS.get(log.status)
            breached  = limit_sec is not None and duration_sec > limit_sec
            all_rows.append({
                'id': log.id,
                'machine_id': mid,
                'machine_name': mach.name if mach else str(mid),
                'station_name': station_name,
                'status': log.status,
                'changed_at': log.changed_at,
                'end_time': end_time,
                'duration_sec': duration_sec,
                'is_ongoing': is_ongoing,
                'limit_sec': limit_sec,
                'breached': breached,
                'source': log.source or '',
                'reason': log.deviation_reason or '',
            })

    hdr_fill  = PatternFill('solid', fgColor='1E3A5F')
    hdr_font  = Font(bold=True, color='FFFFFF')
    red_fill  = PatternFill('solid', fgColor='FEE2E2')
    red_font  = Font(bold=True, color='DC2626')
    green_font = Font(color='059669')

    STATUS_SHEETS = [
        ('idle',           'Idle Deviations'),
        ('breakdown',      'Breakdown'),
        ('alarm',          'Alarm'),
        ('offline',        'Offline'),
        ('setting_change', 'Setting Change'),
    ]

    HEADERS = ['Machine', 'Station', 'Status', 'Start Time (IST)', 'End Time (IST)',
               'Duration', 'Ongoing', 'Limit', 'EXCEEDED', 'Deviation Reason', 'Source']

    def write_rows(ws, rows):
        ws.append(HEADERS)
        for col in range(1, len(HEADERS)+1):
            c = ws.cell(1, col)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal='center')
        for r in rows:
            eff_status = 'Ld/UnLd' if r['status'] == 'idle' and r['duration_sec'] < 60 else r['status'].replace('_', ' ').title()
            row = [
                r['machine_name'], r['station_name'], eff_status,
                fmt_ist(r['changed_at']),
                fmt_ist(r['end_time']) if r.get('end_time') else '(ongoing)',
                dur_label(r['duration_sec']),
                'Yes' if r['is_ongoing'] else 'No',
                dur_label(r['limit_sec']) if r['limit_sec'] else '—',
                'YES' if r['breached'] else '',
                r['reason'], r['source'],
            ]
            ws.append(row)
            row_idx = ws.max_row
            if r['breached']:
                for col in range(1, len(HEADERS)+1):
                    ws.cell(row_idx, col).fill = red_fill
                ws.cell(row_idx, 9).font = red_font
            if r['reason']:
                ws.cell(row_idx, 10).font = green_font
        col_widths = [22, 14, 16, 24, 24, 12, 10, 12, 10, 35, 12]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(1,i).column_letter].width = w

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Shift Utilization (created first) ────────────────────────────
    for status_key, sheet_name in STATUS_SHEETS:
        deviated = sorted(
            [r for r in all_rows if r['status'] == status_key and r['breached']],
            key=lambda r: r['duration_sec'], reverse=True
        )
        ws = wb.create_sheet(sheet_name)
        write_rows(ws, deviated)
        if not deviated:
            ws.append([f'No {sheet_name} deviations found for {target}'])

    # Ld/UnLd sheet — idle records < 60s
    ldunld_rows = sorted(
        [r for r in all_rows if r['status'] == 'idle' and r['duration_sec'] < 60],
        key=lambda r: r['duration_sec'], reverse=True
    )
    ws_ld = wb.create_sheet('Ld-UnLd (Load-Unload)')
    write_rows(ws_ld, ldunld_rows)
    if not ldunld_rows:
        ws_ld.append([f'No Ld-UnLd records found for {target}'])

    # Top-20 sheet — all statuses, highest duration first
    top20 = sorted(
        [r for r in all_rows if r['breached']],
        key=lambda r: r['duration_sec'], reverse=True
    )[:20]
    ws_top = wb.create_sheet('Top 20 Deviations')
    write_rows(ws_top, top20)
    if not top20:
        ws_top.append([f'No deviations found for {target}'])

    if not wb.sheetnames:
        wb.create_sheet('No Data').append([f'No LOSS TRACKER data for {target}'])

    # ── Load shift config from DB ─────────────────────────────────────────────
    import json
    from datetime import timezone, timedelta as td
    from ..models import SiteConfig
    IST_TZ = timezone(td(hours=5, minutes=30))
    _cfg_row = db.query(SiteConfig).first()
    _cfg = json.loads(_cfg_row.config_json) if _cfg_row else {}
    _shifts = [s for s in _cfg.get('shifts', [
        {'id': 'A', 'start': '08:00', 'end': '20:00'},
        {'id': 'B', 'start': '20:00', 'end': '08:00'},
    ]) if s.get('enabled', True)]

    def shift_utc_bounds(shift, date_ist):
        """Return (start_utc, end_utc, [hour_slots_utc]) for a shift on a given IST date."""
        sh, sm = map(int, shift['start'].split(':'))
        eh, em = map(int, shift['end'].split(':'))
        start_ist = dt.combine(date_ist, dt.min.time().replace(hour=sh, minute=sm))
        if eh <= sh:  # crosses midnight
            end_ist = dt.combine(date_ist + timedelta(days=1), dt.min.time().replace(hour=eh, minute=em))
        else:
            end_ist = dt.combine(date_ist, dt.min.time().replace(hour=eh, minute=em))
        start_utc = start_ist - td(hours=5, minutes=30)
        end_utc   = end_ist   - td(hours=5, minutes=30)
        # build hourly slots (IST datetimes)
        slots = []
        cur = start_ist
        while cur < end_ist:
            nxt = cur + timedelta(hours=1)
            slots.append((cur, min(nxt, end_ist)))
            cur = nxt
        return start_utc, end_utc, slots

    def to_ist(utc_dt):
        """changed_at is now IST naive — return as-is."""
        return utc_dt

    def fmt_dur(sec):
        if sec <= 0: return '0m'
        h, r = divmod(int(sec), 3600)
        m    = r // 60
        return f'{h}h {m}m' if h else f'{m}m'

    STATUSES = ['running', 'idle', 'breakdown', 'alarm', 'offline', 'setting_change']
    LD_UNLD_MAX_SEC = 60  # idle < 60s treated as ld/unld

    def classify(row):
        if row['status'] == 'idle' and row['duration_sec'] < LD_UNLD_MAX_SEC:
            return 'ld_unld'
        return row['status']

    # Build a lookup: machine_id -> rows sorted by changed_at asc
    mach_rows = {}
    for r in all_rows:
        mach_rows.setdefault(r['machine_id'], []).append(r)
    for v in mach_rows.values():
        v.sort(key=lambda r: r['changed_at'])

    # ── Sheet 1: Shift Utilization ────────────────────────────────────────────
    ws_util = wb.create_sheet('Shift Utilization')
    util_hdrs = [
        'Date', 'Station', 'Machine', 'Shift',
        'Running Duration', 'Running Count',
        'Ld/UnLd Duration', 'Ld/UnLd Count',
        'Idle Duration', 'Idle Count',
        'Breakdown Duration', 'Breakdown Count',
        'Alarm Duration', 'Alarm Count',
        'Offline Duration', 'Offline Count',
        'Setting Change Duration', 'Setting Change Count',
        'Unaccounted Duration',
    ]
    ws_util.append(util_hdrs)
    for col in range(1, len(util_hdrs)+1):
        c = ws_util.cell(1, col); c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal='center')
    ws_util.freeze_panes = 'A2'

    for mid, mrows in sorted(mach_rows.items()):
        mach = machine_map.get(mid)
        station_name  = station_map.get(mach.station_id, '') if mach else ''
        mach_name  = mach.name if mach else str(mid)
        for sh in _shifts:
            s_utc, e_utc, _ = shift_utc_bounds(sh, target)
            shift_dur_sec = (e_utc - s_utc).total_seconds()
            # filter rows whose changed_at falls inside this shift window
            in_shift = [r for r in mrows if s_utc <= r['changed_at'] < e_utc]
            if not in_shift:
                continue
            buckets = {'running': [0, 0], 'ld_unld': [0, 0], 'idle': [0, 0],
                       'breakdown': [0, 0], 'alarm': [0, 0],
                       'offline': [0, 0], 'setting_change': [0, 0]}
            for r in in_shift:
                key = classify(r)
                if key in buckets:
                    buckets[key][0] += r['duration_sec']
                    buckets[key][1] += 1
            tracked_sec = sum(v[0] for v in buckets.values())
            unaccounted_sec = max(0, shift_dur_sec - tracked_sec)
            ws_util.append([
                str(target), station_name, mach_name, sh['id'],
                fmt_dur(buckets['running'][0]),      buckets['running'][1],
                fmt_dur(buckets['ld_unld'][0]),      buckets['ld_unld'][1],
                fmt_dur(buckets['idle'][0]),         buckets['idle'][1],
                fmt_dur(buckets['breakdown'][0]),    buckets['breakdown'][1],
                fmt_dur(buckets['alarm'][0]),        buckets['alarm'][1],
                fmt_dur(buckets['offline'][0]),      buckets['offline'][1],
                fmt_dur(buckets['setting_change'][0]), buckets['setting_change'][1],
                fmt_dur(unaccounted_sec),
            ])
    for i, w in enumerate([12,14,20,8, 18,14, 16,14, 14,12, 20,18, 16,14, 16,14, 22,20, 20], 1):
        ws_util.column_dimensions[ws_util.cell(1,i).column_letter].width = w

    # ── Sheet 2: CT Distribution ──────────────────────────────────────────────
    ws_ct = wb.create_sheet('CT Distribution')

    SHIFT_COLOURS = ['0070C0', '00B050', 'FF0000', 'FF9900']
    TOTAL_COL_COLOUR = '00B0F0'   # light blue for shift totals
    DAY_TOTAL_COLOUR = '00B0F0'

    # Build per-shift slot lists (IST datetime pairs)
    shift_slots = {}  # shift_id -> [(s_ist, e_ist), ...]
    for sh in _shifts:
        _, _, slots = shift_utc_bounds(sh, target)
        shift_slots[sh['id']] = slots

    # Pre-compute column layout:
    # Cols: DATE(1) Pair(2) Machine(3) Day Total(4) State(5)
    #       [ShiftA slots... ShiftA Total] [ShiftB slots... ShiftB Total] ...
    # shift_layout: list of {sh, col_start, col_end, total_col}
    shift_layout = []
    col_cursor = 6
    for sh in _shifts:
        slots = shift_slots.get(sh['id'], [])
        if not slots:
            continue
        col_start  = col_cursor
        col_end    = col_cursor + len(slots) - 1
        total_col  = col_end + 1
        shift_layout.append({
            'sh': sh, 'slots': slots,
            'col_start': col_start, 'col_end': col_end, 'total_col': total_col
        })
        col_cursor = total_col + 1

    total_cols = col_cursor - 1  # last used column

    from openpyxl.utils import get_column_letter

    # ─ Row 1 & 2: headers ─
    ws_ct.cell(1, 1).value = 'DATE'
    ws_ct.cell(1, 2).value = 'Station'
    ws_ct.cell(1, 3).value = 'Machine'
    ws_ct.cell(1, 4).value = 'Day Total'
    ws_ct.cell(1, 5).value = 'State'
    for col in range(1, 6):
        ws_ct.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        c = ws_ct.cell(1, col)
        c.font = hdr_font
        c.fill = hdr_fill if col != 4 else PatternFill('solid', fgColor=DAY_TOTAL_COLOUR)
        c.alignment = Alignment(horizontal='center', vertical='center')

    for li in shift_layout:
        sh     = li['sh']
        slots  = li['slots']
        cs     = li['col_start']
        ce     = li['col_end']
        tc     = li['total_col']
        colour = SHIFT_COLOURS[_shifts.index(sh) % len(SHIFT_COLOURS)]
        sh_fill   = PatternFill('solid', fgColor=colour)
        tot_fill  = PatternFill('solid', fgColor=TOTAL_COL_COLOUR)

        # Row 1: merge shift name across hour slots (NOT including total col)
        ws_ct.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=ce)
        sh_cell = ws_ct.cell(1, cs)
        sh_cell.value = sh.get('name', f"Shift {sh['id']}")
        sh_cell.font  = Font(bold=True, color='FFFFFF')
        sh_cell.fill  = sh_fill
        sh_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Row 1: shift total header (separate, not merged into shift name)
        tc_cell = ws_ct.cell(1, tc)
        tc_cell.value = f"{sh.get('name', sh['id'])} Total"
        tc_cell.font  = Font(bold=True, color='FFFFFF')
        tc_cell.fill  = tot_fill
        tc_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_ct.merge_cells(start_row=1, start_column=tc, end_row=2, end_column=tc)
        ws_ct.cell(1, tc).alignment = Alignment(horizontal='center', vertical='center')

        # Row 2: hourly slot labels
        for i, (s_ist, e_ist) in enumerate(slots):
            label = f"{s_ist.strftime('%H:%M')}-{e_ist.strftime('%H:%M')}"
            sc = ws_ct.cell(2, cs + i)
            sc.value = label
            sc.font  = Font(bold=True, color='FFFFFF')
            sc.fill  = sh_fill
            sc.alignment = Alignment(horizontal='center')

    ws_ct.freeze_panes = 'F3'
    ws_ct.row_dimensions[1].height = 22
    ws_ct.row_dimensions[2].height = 18

    # ─ Data rows ─
    STATES = ['Idle', 'Running', 'Ld/UnLd']
    state_classify = {'idle': 'Idle', 'running': 'Running', 'ld_unld': 'Ld/UnLd'}
    state_colours  = {'Idle': 'FFF2CC', 'Running': 'E2EFDA', 'Ld/UnLd': 'EDE7F6'}

    data_row = 3
    for mid, mrows in sorted(mach_rows.items()):
        mach      = machine_map.get(mid)
        station_name = station_map.get(mach.station_id, '') if mach else ''
        mach_name = mach.name if mach else str(mid)
        group_start = data_row

        for si, state_label in enumerate(STATES):
            r          = data_row + si
            state_fill = PatternFill('solid', fgColor=state_colours[state_label])
            ws_ct.cell(r, 5).value     = state_label
            ws_ct.cell(r, 5).alignment = Alignment(horizontal='center')
            ws_ct.cell(r, 5).fill      = state_fill

            shift_total_cols = []  # collect total_col letters for Day Total formula

            for li in shift_layout:
                sh      = li['sh']
                slots   = li['slots']
                cs      = li['col_start']
                tc      = li['total_col']
                s_utc_sh, e_utc_sh, _ = shift_utc_bounds(sh, target)
                in_shift = [row for row in mrows if s_utc_sh <= row['changed_at'] < e_utc_sh]

                for i, (s_ist, e_ist) in enumerate(slots):
                    s_utc_slot = s_ist - td(hours=5, minutes=30)
                    e_utc_slot = e_ist - td(hours=5, minutes=30)
                    count = sum(
                        1 for row in in_shift
                        if s_utc_slot <= row['changed_at'] < e_utc_slot
                        and state_classify.get(classify(row)) == state_label
                    )
                    c = ws_ct.cell(r, cs + i)
                    c.value     = count
                    c.fill      = state_fill
                    c.alignment = Alignment(horizontal='center')

                # Shift Total = SUM(slots for this shift)
                cs_letter = get_column_letter(cs)
                ce_letter = get_column_letter(li['col_end'])
                tc_cell   = ws_ct.cell(r, tc)
                tc_cell.value     = f'=SUM({cs_letter}{r}:{ce_letter}{r})'
                tc_cell.font      = Font(bold=True)
                tc_cell.fill      = PatternFill('solid', fgColor=TOTAL_COL_COLOUR)
                tc_cell.alignment = Alignment(horizontal='center')
                shift_total_cols.append(get_column_letter(tc))

            # Day Total = sum of all shift total columns
            day_formula = '+'.join(f'{col}{r}' for col in shift_total_cols)
            dt_cell = ws_ct.cell(r, 4)
            dt_cell.value     = f'={day_formula}'
            dt_cell.font      = Font(bold=True)
            dt_cell.fill      = PatternFill('solid', fgColor=DAY_TOTAL_COLOUR)
            dt_cell.alignment = Alignment(horizontal='center')

        # Merge A, B, C, D across 3 state rows
        for col in [1, 2, 3]:
            ws_ct.merge_cells(start_row=group_start, start_column=col,
                              end_row=group_start + 2, end_column=col)
            ws_ct.cell(group_start, col).alignment = Alignment(horizontal='center', vertical='center')

        ws_ct.cell(group_start, 1).value = str(target)
        ws_ct.cell(group_start, 2).value = station_name
        ws_ct.cell(group_start, 3).value = mach_name

        data_row += 3

    # Column widths
    for col, w in zip(range(1, 6), [13, 14, 22, 12, 12]):
        ws_ct.column_dimensions[get_column_letter(col)].width = w
    for col in range(6, total_cols + 1):
        # Shift total columns slightly wider
        is_total = any(col == li['total_col'] for li in shift_layout)
        ws_ct.column_dimensions[get_column_letter(col)].width = 14 if is_total else 11

    # ── Reorder sheets: Shift Utilization first, CT Distribution second, rest after ──
    desired_order = ['Shift Utilization', 'CT Distribution'] + [
        s for s in wb.sheetnames if s not in ('Shift Utilization', 'CT Distribution')
    ]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_tools_xlsx(db: Session, report_date=None) -> bytes:
    """Tool inventory, life status, open alerts, and recent consumption history."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from ..models import ToolStock, ToolAlert, ToolEvent, WorkOrder, Machine
    from ..tool_service import serialize_tool, refresh_tool_status

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    warn_fill = PatternFill("solid", fgColor="FEF3C7")
    alert_fill = PatternFill("solid", fgColor="FEE2E2")

    # ── Inventory ──
    ws = wb.active
    ws.title = "Tool Inventory"
    inv_headers = [
        "Tool Code", "Tool Name", "Stock", "Min Stock", "Below Min", "Unit",
        "Life Used", "Set Life", "Life %", "Status", "Cycles/Part",
        "Source", "SAP Material", "QR Code", "Last Synced", "Notes", "Active",
    ]
    ws.append(inv_headers)
    for col, _ in enumerate(inv_headers, 1):
        cell = ws.cell(1, col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    tools = db.query(ToolStock).order_by(ToolStock.tool_code).all()
    for tool in tools:
        refresh_tool_status(tool)
        row = serialize_tool(tool)
        values = [
            row["tool_code"], row["tool_name"], row["stock_qty"], row["min_stock"],
            "YES" if row["below_min"] else "NO", row["unit"],
            row["cycles_used"], row["life_cycles_limit"] or "",
            row["life_used_pct"] if row["life_used_pct"] is not None else "",
            row["tool_status"], row["cycles_per_part"],
            row["stock_source"], row["sap_material_no"] or "",
            row["qr_code"] or row["tool_code"],
            row["last_synced_at"] or "", row["notes"] or "",
            "Yes" if row["active"] else "No",
        ]
        ws.append(values)
        r_idx = ws.max_row
        if row["below_min"] or row["tool_status"] in ("eol", "blocked"):
            for c in range(1, len(inv_headers) + 1):
                ws.cell(r_idx, c).fill = alert_fill
        elif row["tool_status"] in ("near_eol", "correction_ack"):
            for c in range(1, len(inv_headers) + 1):
                ws.cell(r_idx, c).fill = warn_fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 14

    # ── Open Alerts ──
    ws_a = wb.create_sheet("Open Alerts")
    alert_headers = ["Alert ID", "Tool Code", "Tool Name", "Type", "Severity", "Message", "Suppressed", "Acknowledged", "Created At"]
    ws_a.append(alert_headers)
    for col, _ in enumerate(alert_headers, 1):
        cell = ws_a.cell(1, col)
        cell.fill = hdr_fill
        cell.font = hdr_font
    tool_map = {t.id: t for t in tools}
    alerts = (
        db.query(ToolAlert)
        .filter(ToolAlert.acknowledged == 0)
        .order_by(ToolAlert.created_at.desc())
        .limit(500)
        .all()
    )
    for a in alerts:
        t = tool_map.get(a.tool_id)
        ws_a.append([
            a.id,
            t.tool_code if t else "",
            t.tool_name if t else "",
            a.alert_type,
            a.severity,
            a.message,
            "Yes" if a.suppressed else "No",
            "Yes" if a.acknowledged else "No",
            str(a.created_at or ""),
        ])
    for col in ws_a.columns:
        ws_a.column_dimensions[col[0].column_letter].width = 16

    # ── Recent Events ──
    ws_e = wb.create_sheet("Life & Consumption")
    ev_headers = [
        "When", "Tool Code", "Event", "Qty Δ", "Cycles Before", "Cycles After", "Cycles Δ",
        "Work Order", "Plan ID", "Machine", "Location", "Notes", "QR Scanned", "QR Suppressed",
    ]
    ws_e.append(ev_headers)
    for col, _ in enumerate(ev_headers, 1):
        cell = ws_e.cell(1, col)
        cell.fill = hdr_fill
        cell.font = hdr_font
    events = db.query(ToolEvent).order_by(ToolEvent.created_at.desc()).limit(1000).all()
    machines = {m.id: m.name for m in db.query(Machine).all()}
    wo_map = {w.id: w.work_order_no for w in db.query(WorkOrder).all()}
    for e in events:
        t = tool_map.get(e.tool_id)
        ws_e.append([
            str(e.created_at or ""),
            t.tool_code if t else "",
            e.event_type,
            float(e.qty_delta) if e.qty_delta is not None else "",
            float(e.cycles_before) if e.cycles_before is not None else "",
            float(e.cycles_after) if e.cycles_after is not None else "",
            float(e.cycles_delta) if e.cycles_delta is not None else "",
            wo_map.get(e.work_order_id, ""),
            e.plan_id or "",
            machines.get(e.machine_id, ""),
            e.location or "",
            e.notes or "",
            "Yes" if e.qr_scanned else "No",
            "Yes" if e.qr_suppressed else "No",
        ])
    for col in ws_e.columns:
        ws_e.column_dimensions[col[0].column_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_attachments_for_report_types(report_types: str, db: Session, report_date=None) -> dict:
    from datetime import date as date_type, timedelta
    rts = [r.strip() for r in report_types.split(",")]
    attachments = {}
    if "oee"         in rts: attachments["oee_report.xlsx"]         = build_oee_xlsx(db, report_date)
    if "planning"    in rts: attachments["production_plan.xlsx"]    = build_plan_xlsx(db)
    if "breakdown"   in rts: attachments["breakdown_report.xlsx"]   = build_breakdown_xlsx(db)
    if "maintenance" in rts: attachments["maintenance_report.xlsx"] = build_maintenance_xlsx(db)
    if "data_entry"    in rts: attachments["data_entry_report.xlsx"]    = build_data_entry_xlsx(db, report_date)
    if "loss_tracker" in rts:
        # Pair Loss Tracker with TPM Loss Logger (same date) for scheduled + manual sends
        if report_date is None:
            date_str = (date_type.today() - timedelta(days=1)).isoformat()
        elif isinstance(report_date, str):
            date_str = report_date[:10]
        else:
            date_str = report_date.isoformat()
        attachments[f"loss_tracker_{date_str}.xlsx"] = build_loss_tracker_xlsx(db, report_date)
        attachments[f"TPM_Loss_logger_{date_str}.xlsx"] = build_tpm_loss_logger_xlsx(db, report_date)
    if "tools"         in rts: attachments["tool_management_report.xlsx"] = build_tools_xlsx(db, report_date)
    return attachments

def do_send(cfg, to_list: List[str], subject: str, body: str,
            attachments: dict = None):
    """attachments = {filename: bytes}"""
    if not to_list:
        return
    msg = MIMEMultipart()
    msg['From']    = cfg.email_address
    msg['To']      = ", ".join(to_list)
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'plain'))

    for fname, data in (attachments or {}).items():
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(data)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename="{fname}"')
        msg.attach(part)

    server = smtplib.SMTP(cfg.smtp_server, cfg.smtp_port, timeout=15)
    server.starttls()
    server.login(cfg.email_address, cfg.email_password)
    server.sendmail(cfg.email_address, to_list, msg.as_string())
    server.quit()

# ── SMTP Config ───────────────────────────────────────────────────────────────

@router.get("/smtp")
def get_smtp_config(db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    cfg = db.query(EmailSmtpConfig).first()
    if not cfg:
        return {}
    return {"smtp_server": cfg.smtp_server, "smtp_port": cfg.smtp_port,
            "email_address": cfg.email_address, "password_set": bool(cfg.email_password)}

@router.post("/smtp")
def save_smtp_config(data: SmtpConfigIn, db: Session = Depends(get_db),
                     _=Depends(require_role("admin"))):
    cfg = db.query(EmailSmtpConfig).first()
    if cfg:
        cfg.smtp_server    = data.smtp_server
        cfg.smtp_port      = data.smtp_port
        cfg.email_address  = data.email_address
        cfg.email_password = data.email_password
    else:
        cfg = EmailSmtpConfig(**data.model_dump())
        db.add(cfg)
    db.commit()
    return {"ok": True}

@router.post("/smtp/test")
def test_smtp(db: Session = Depends(get_db), _=Depends(require_role("admin"))):
    cfg = get_smtp(db)
    try:
        server = smtplib.SMTP(cfg.smtp_server, cfg.smtp_port, timeout=10)
        server.starttls()
        server.login(cfg.email_address, cfg.email_password)
        server.quit()
        return {"ok": True, "message": "SMTP connection successful"}
    except Exception as e:
        raise HTTPException(400, f"SMTP test failed: {str(e)}")

# ── Groups ────────────────────────────────────────────────────────────────────

@router.get("/groups")
def get_groups(db: Session = Depends(get_db), _=Depends(get_current_user)):
    groups = db.query(EmailGroup).all()
    result = []
    for g in groups:
        members = db.query(EmailRecipient).filter(EmailRecipient.group_id == g.id).all()
        result.append({"id": g.id, "name": g.name, "description": g.description,
                       "report_types": g.report_types or "oee,planning,breakdown",
                       "count": len([m for m in members if m.active]),
                       "members": [{"id": m.id, "name": m.name, "email": m.email,
                                    "active": m.active} for m in members]})
    return result

@router.post("/groups")
def create_group(data: GroupIn, db: Session = Depends(get_db),
                 _=Depends(require_role("admin"))):
    if db.query(EmailGroup).filter(EmailGroup.name == data.name).first():
        raise HTTPException(400, "Group name already exists")
    # auto-apply defaults by name if report_types not explicitly set
    rts = data.report_types or REPORT_DEFAULTS.get(data.name.lower(), "oee,planning,breakdown")
    g = EmailGroup(name=data.name.strip(), description=data.description.strip(), report_types=rts)
    db.add(g)
    db.commit()
    db.refresh(g)
    return g

@router.patch("/groups/{gid}")
def update_group(gid: int, data: GroupIn, db: Session = Depends(get_db),
                 _=Depends(require_role("admin"))):
    g = db.query(EmailGroup).filter(EmailGroup.id == gid).first()
    if not g: raise HTTPException(404, "Not found")
    g.name = data.name.strip()
    g.description = data.description.strip()
    g.report_types = data.report_types
    db.commit()
    return g

@router.delete("/groups/{gid}")
def delete_group(gid: int, db: Session = Depends(get_db),
                 _=Depends(require_role("admin"))):
    g = db.query(EmailGroup).filter(EmailGroup.id == gid).first()
    if not g: raise HTTPException(404, "Not found")
    db.query(EmailRecipient).filter(EmailRecipient.group_id == gid).delete(synchronize_session=False)
    db.delete(g)
    db.commit()
    return {"ok": True}

# ── Recipients ────────────────────────────────────────────────────────────────

@router.post("/recipients")
def add_recipient(data: RecipientIn, db: Session = Depends(get_db),
                  _=Depends(require_role("admin", "supervisor"))):
    r = EmailRecipient(**data.model_dump())
    db.add(r)
    db.commit()
    db.refresh(r)
    return r

@router.patch("/recipients/{rid}")
def update_recipient(rid: int, data: RecipientIn, db: Session = Depends(get_db),
                     _=Depends(require_role("admin", "supervisor"))):
    r = db.query(EmailRecipient).filter(EmailRecipient.id == rid).first()
    if not r: raise HTTPException(404, "Not found")
    for k, v in data.model_dump().items():
        setattr(r, k, v)
    db.commit()
    return r

@router.delete("/recipients/{rid}")
def delete_recipient(rid: int, db: Session = Depends(get_db),
                     _=Depends(require_role("admin", "supervisor"))):
    r = db.query(EmailRecipient).filter(EmailRecipient.id == rid).first()
    if not r: raise HTTPException(404, "Not found")
    db.delete(r)
    db.commit()
    return {"ok": True}

# ── Schedules ─────────────────────────────────────────────────────────────────

@router.get("/schedules")
def get_schedules(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return db.query(EmailSchedule).all()

@router.post("/schedules")
def create_schedule(data: ScheduleIn, db: Session = Depends(get_db),
                    _=Depends(require_role("admin", "supervisor"))):
    s = EmailSchedule(**data.model_dump())
    db.add(s)
    db.commit()
    db.refresh(s)
    from ..scheduler_service import reload_schedules
    reload_schedules(db)
    return s

@router.patch("/schedules/{sid}")
def update_schedule(sid: int, data: ScheduleIn, db: Session = Depends(get_db),
                    _=Depends(require_role("admin", "supervisor"))):
    s = db.query(EmailSchedule).filter(EmailSchedule.id == sid).first()
    if not s: raise HTTPException(404, "Not found")
    for k, v in data.model_dump().items():
        setattr(s, k, v)
    db.commit()
    from ..scheduler_service import reload_schedules
    reload_schedules(db)
    return s

@router.delete("/schedules/{sid}")
def delete_schedule(sid: int, db: Session = Depends(get_db),
                    _=Depends(require_role("admin", "supervisor"))):
    s = db.query(EmailSchedule).filter(EmailSchedule.id == sid).first()
    if not s: raise HTTPException(404, "Not found")
    db.delete(s)
    db.commit()
    from ..scheduler_service import reload_schedules
    reload_schedules(db)
    return {"ok": True}

# ── Manual Send ───────────────────────────────────────────────────────────────

@router.post("/send")
def manual_send(data: ManualSendIn, bg: BackgroundTasks,
                db: Session = Depends(get_db),
                user=Depends(require_role("admin", "supervisor"))):
    cfg = get_smtp(db)
    to_list = get_group_emails(db, data.group_ids)
    if not to_list:
        raise HTTPException(400, "No active recipients in selected groups")

    # collect report types from selected groups
    grps = db.query(EmailGroup).filter(EmailGroup.id.in_(data.group_ids)).all()
    combined = set()
    for grp in grps:
        for r in (grp.report_types or "oee,planning,breakdown").split(","):
            combined.add(r.strip())
    report_type_str = ",".join(combined)

    # use provided report_date or default to yesterday
    from datetime import date as date_type, timedelta
    if data.report_date:
        report_date = data.report_date
    else:
        report_date = (date_type.today() - timedelta(days=1)).isoformat()

    # capture primitive values only — never pass ORM objects or open sessions to bg task
    smtp_server   = cfg.smtp_server
    smtp_port     = cfg.smtp_port
    smtp_email    = cfg.email_address
    smtp_password = cfg.email_password
    attach        = data.attach_report
    subject       = data.subject
    body          = data.body
    sent_by       = user.id if user else None
    rpt_date      = report_date

    log = EmailLog(sent_at=_now_ist(), recipients=", ".join(to_list),
                   subject=subject, report_type="manual",
                   status="pending", sent_by=sent_by)
    db.add(log)
    db.commit()
    log_id = log.id

    def _send_bg(log_id, server, port, email, password, rtype_str, do_attach, rpt_date):
        from ..models import SessionLocal
        _db = SessionLocal()
        try:
            attachments = build_attachments_for_report_types(rtype_str, _db, rpt_date) if do_attach else {}
            class _Cfg:
                smtp_server    = server
                smtp_port      = port
                email_address  = email
                email_password = password
            do_send(_Cfg(), to_list, subject, body, attachments)
            _log = _db.query(EmailLog).filter(EmailLog.id == log_id).first()
            if _log:
                _log.status = "sent"
                _db.commit()
        except Exception as e:
            _log = _db.query(EmailLog).filter(EmailLog.id == log_id).first()
            if _log:
                _log.status = "failed"
                _log.error_msg = str(e)
                _db.commit()
            print(f"[ManualSend] Error: {e}")
        finally:
            _db.close()

    bg.add_task(_send_bg, log_id, smtp_server, smtp_port, smtp_email, smtp_password,
                report_type_str, attach, rpt_date)
    return {"ok": True, "recipients": len(to_list), "to": to_list,
            "sent_at": _now_ist().isoformat(),
            "reports": [report_type_str] if attach else []}


@router.get("/logs")
def get_email_logs(db: Session = Depends(get_db), _=Depends(get_current_user)):
    logs = db.query(EmailLog).order_by(EmailLog.sent_at.desc()).limit(200).all()
    return [{"id": l.id, "sent_at": str(l.sent_at), "recipients": l.recipients,
             "subject": l.subject, "report_type": l.report_type,
             "status": l.status, "error_msg": l.error_msg} for l in logs]


@router.get("/download/loss-tracker")
def download_loss_tracker(
    report_date: str = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    from fastapi.responses import Response
    from datetime import date as date_type, timedelta
    if not report_date:
        report_date = (date_type.today() - timedelta(days=1)).isoformat()
    data = build_loss_tracker_xlsx(db, report_date)
    filename = f"loss_tracker_{report_date}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


def build_tpm_loss_logger_xlsx(db: Session, report_date=None) -> bytes:
    """
    TPM Loss Logger workbook from tablet OperatorLossLog rows.
    Sheet 1: All TPM losses for the date.
    Sheet 2+: one sheet per Loss Assigner type (LOSS-1 … LOSS-16).
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import date as date_type, timedelta, datetime as dt
    from ..models import OperatorLossLog, Machine, Station, Operator
    from ..loss_mapping import TPM_LOSS_ASSIGNER_TYPES

    if report_date is None:
        target = date_type.today() - timedelta(days=1)
    elif isinstance(report_date, str):
        target = date_type.fromisoformat(report_date)
    else:
        target = report_date

    rows = (
        db.query(OperatorLossLog)
        .filter(OperatorLossLog.entry_date == target)
        .order_by(OperatorLossLog.started_at.asc(), OperatorLossLog.id.asc())
        .all()
    )

    machine_map = {m.id: m for m in db.query(Machine).all()}
    station_map = {p.id: (p.display_name or p.name) for p in db.query(Station).all()}
    op_map = {o.id: o for o in db.query(Operator).all()}

    def fmt_ist(val):
        if not val:
            return ""
        if hasattr(val, "strftime"):
            return val.strftime("%d-%m-%Y %H:%M:%S")
        return str(val)

    def operator_label(row: OperatorLossLog) -> str:
        if row.operator_id and row.operator_id in op_map:
            op = op_map[row.operator_id]
            name = (op.name or "").strip()
            code = (op.employee_code or "").strip()
            if name and code and name.lower() != code.lower():
                return f"{name} ({code})"
            return name or code
        return row.username or ""

    def minutes_of(row: OperatorLossLog) -> float:
        if row.status == "open" and row.started_at:
            now = dt.now(_IST).replace(tzinfo=None)
            return round(max(0.0, (now - row.started_at).total_seconds() / 60.0), 2)
        if row.minutes is not None:
            return float(row.minutes)
        if row.started_at and row.ended_at:
            return round(max(0.0, (row.ended_at - row.started_at).total_seconds() / 60.0), 2)
        return 0.0

    HEADERS = [
        "ID", "Machine", "Station", "Operator", "Tab ID",
        "Loss Code", "Loss Description", "Sub-division / Detail",
        "Minutes", "Status", "Shift", "Entry Date",
        "Started At", "Ended At", "OEE Field", "Notes",
    ]

    hdr_fill = PatternFill("solid", fgColor="0F3D68")
    hdr_font = Font(bold=True, color="FFFFFF")
    open_fill = PatternFill("solid", fgColor="FEF3C7")

    def sheet_title(code: str, label: str) -> str:
        # Excel sheet name max 31 chars
        raw = f"{code} {label}".replace("/", "-")
        return raw[:31]

    def write_sheet(ws, loss_rows, empty_msg: str):
        ws.append(HEADERS)
        for col in range(1, len(HEADERS) + 1):
            c = ws.cell(1, col)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        if not loss_rows:
            ws.append([empty_msg])
            return
        for row in loss_rows:
            mach = machine_map.get(row.machine_id)
            station = station_map.get(mach.station_id, "") if mach else ""
            ws.append([
                row.id,
                mach.name if mach else row.machine_id,
                station,
                operator_label(row),
                row.tab_id or "",
                row.loss_code or "",
                row.loss_description or "",
                row.sub_division or "",
                minutes_of(row),
                row.status or "",
                row.shift or "",
                row.entry_date.isoformat() if row.entry_date else "",
                fmt_ist(row.started_at),
                fmt_ist(row.ended_at) if row.ended_at else ("(open)" if row.status == "open" else ""),
                row.oee_field or "",
                (row.notes or "")[:500],
            ])
            if row.status == "open":
                for col in range(1, len(HEADERS) + 1):
                    ws.cell(ws.max_row, col).fill = open_fill
        widths = [8, 16, 14, 22, 16, 12, 28, 22, 10, 10, 8, 12, 20, 20, 16, 36]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = openpyxl.Workbook()
    # Sheet 1 — all tablet TPM losses
    ws_all = wb.active
    ws_all.title = "All TPM Losses"
    write_sheet(
        ws_all,
        rows,
        f"No TPM losses recorded on tablets for {target.isoformat()}",
    )

    # Sheet 2+ — one per Loss Assigner type
    by_code = {}
    for r in rows:
        code = (r.loss_code or "").strip().upper()
        by_code.setdefault(code, []).append(r)

    for loss_type in TPM_LOSS_ASSIGNER_TYPES:
        code = loss_type["code"]
        title = sheet_title(code, loss_type["label"])
        ws = wb.create_sheet(title)
        write_sheet(
            ws,
            by_code.get(code, []),
            f"No {code} · {loss_type['description']} records for {target.isoformat()}",
        )

    # Any codes not in the standard 16 (legacy / manual)
    known = {t["code"] for t in TPM_LOSS_ASSIGNER_TYPES}
    extras = sorted(c for c in by_code.keys() if c and c not in known)
    for code in extras:
        title = sheet_title(code, "Other")
        ws = wb.create_sheet(title)
        write_sheet(ws, by_code[code], f"No records for {code}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/download/tpm-loss-logger")
def download_tpm_loss_logger(
    report_date: str = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    from fastapi.responses import Response
    from datetime import date as date_type, timedelta
    if not report_date:
        report_date = (date_type.today() - timedelta(days=1)).isoformat()
    data = build_tpm_loss_logger_xlsx(db, report_date)
    filename = f"TPM_Loss_logger_{report_date}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
