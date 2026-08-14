from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract, or_
from typing import Optional, List
from datetime import date, datetime, timedelta
from types import SimpleNamespace
from calendar import monthrange
import csv, io, json, logging
from ..models import (
    OEEEntry, OEEDefectLog, Machine, Station, ProductionPlan,
    WorkOrder, MachineStatusLog, SiteConfig, get_db, now_ist,
)
from ..auth import get_current_user
from pydantic import BaseModel

log = logging.getLogger(__name__)

router = APIRouter(prefix="/api/oee", tags=["oee"])

class OEECreate(BaseModel):
    entry_date: date
    station_no: int
    machine_id: Optional[int] = None
    shift: str
    current_operation: Optional[str] = ""
    next_operation: Optional[str] = ""
    model_variant: Optional[str] = None
    process_time: float
    loading_unloading: float
    start_time: str
    stop_time: str
    total_minutes: int
    lunch_break: int = 0
    tea_break: int = 0
    tpm_cleaning: int = 0
    other_cleaning: int = 0
    management_meeting: int = 0
    no_load: int = 0
    new_model_trial: int = 0
    power_cut: int = 0
    planned_maintenance: int = 0
    no_manpower_planned: int = 0
    setting_time: int = 0
    tool_change: int = 0
    dimension_correction: int = 0
    scrap_removal: int = 0
    break_down: int = 0
    actual_qty: int
    defect_qty: int = 0

def _cap(value: float) -> float:
    """Cap a percentage at 100. Returns the capped value."""
    return min(float(value), 100.0)


def _calculate_display_rates(ar_raw: float, pr_raw: float, qr_raw: float) -> dict:
    ar = _cap(ar_raw)
    pr = _cap(pr_raw)
    qr = _cap(qr_raw)
    oee = round(ar * pr * qr / 10000, 2)
    return {
        "ar": ar,
        "pr": pr,
        "qr": qr,
        "oee": oee,
    }


def calculate_oee(data: OEECreate) -> dict:
    # CT = Process Time + Loading & Unloading (seconds)
    ct = data.process_time + data.loading_unloading

    # Shift Working Min = Total Minutes - Total Breaks
    total_breaks = data.lunch_break + data.tea_break + data.tpm_cleaning + data.other_cleaning + data.management_meeting
    shift_working = data.total_minutes - total_breaks

    # Available Shift Time = Shift Working Min - Mgmt Loss Total
    mgmt_loss = data.no_load + data.new_model_trial + data.power_cut + data.planned_maintenance + data.no_manpower_planned
    available = shift_working - mgmt_loss

    # Operating Time = Available Shift Time - Total Down Time
    total_down = data.setting_time + data.tool_change + data.dimension_correction + data.scrap_removal + data.break_down
    operating = available - total_down

    # Possible Qty = (Operating Time * 60) / CT  [operating in min, CT in sec]
    possible_qty = int((operating * 60) / ct) if ct > 0 else 0

    # Production Loss = Possible Qty - Actual Qty
    production_loss = max(0, possible_qty - data.actual_qty)

    # Accepted Qty = Actual Qty - Defect Qty
    accp_qty = max(0, data.actual_qty - data.defect_qty)

    # Raw (uncapped) rates
    ar_raw = round((operating / available * 100), 2) if available > 0 else 0
    pr_raw = round((data.actual_qty / possible_qty * 100), 2) if possible_qty > 0 else 0
    qr_raw = round((accp_qty / data.actual_qty * 100), 2) if data.actual_qty > 0 else 0
    oee_raw = round(ar_raw * pr_raw * qr_raw / 10000, 2)

    # Capped rates (display / stored values — never exceed 100%)
    rates = _calculate_display_rates(ar_raw, pr_raw, qr_raw)
    ar = rates["ar"]
    pr = rates["pr"]
    qr = rates["qr"]
    oee = rates["oee"]

    return {
        "available_shift_time": available,
        "operating_time": operating,
        "possible_qty": possible_qty,
        "accp_qty": accp_qty,
        "ar": ar, "pr": pr, "qr": qr, "oee": oee,
        # Original uncapped values tracked for audit
        "ar_raw": ar_raw, "pr_raw": pr_raw, "qr_raw": qr_raw, "oee_raw": oee_raw,
        # Flags indicating which values were capped
        "ar_capped":  ar_raw  > 100,
        "pr_capped":  pr_raw  > 100,
        "qr_capped":  qr_raw  > 100,
        "oee_capped": oee_raw > oee,
    }

@router.post("/")
def create_entry(data: OEECreate, db: Session = Depends(get_db), user=Depends(get_current_user)):
    from ..models import ProductionPlan
    calc = calculate_oee(data)
    # Only store raw values when capping actually occurred
    entry_fields = {
        k: v for k, v in calc.items()
        if k not in ("ar_capped", "pr_capped", "qr_capped", "oee_capped")
    }
    if not calc["ar_capped"]:  entry_fields.pop("ar_raw",  None)
    if not calc["pr_capped"]:  entry_fields.pop("pr_raw",  None)
    if not calc["qr_capped"]:  entry_fields.pop("qr_raw",  None)
    if not calc["oee_capped"]: entry_fields.pop("oee_raw", None)
    entry = OEEEntry(**data.model_dump(), **entry_fields, created_by=user.id)
    db.add(entry)
    db.flush()

    # Sync actual_qty to matching production plan
    plan = db.query(ProductionPlan).filter(
        ProductionPlan.plan_date == data.entry_date,
        ProductionPlan.shift == data.shift,
        ProductionPlan.station_no == data.station_no,
        ProductionPlan.current_operation == data.current_operation,
        ProductionPlan.next_operation == data.next_operation,
    ).first()
    if plan:
        plan.actual_qty = data.actual_qty
        if data.actual_qty >= plan.planned_qty and plan.status in ("pending", "running"):
            from .plans import _validate_may_complete_plan
            _validate_may_complete_plan(plan)
            plan.status = "completed"

    db.commit()
    db.refresh(entry)
    return entry

@router.get("/")
def get_entries(
    shift: Optional[str] = None,
    entry_date: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    station_no: Optional[int] = None,
    machine_id: Optional[int] = None,
    current_operation: Optional[str] = None,
    model: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    from ..history_archive import query_oee_entries_federated

    term = (search or model or current_operation or "").strip()
    result, meta = query_oee_entries_federated(
        db,
        shift=shift,
        entry_date=entry_date,
        date_from=date_from,
        date_to=date_to,
        month=month,
        year=year,
        station_no=station_no,
        machine_id=machine_id,
        term=term,
    )
    # Attach federation meta on empty list via custom response is awkward;
    # clients read rows; meta available on /summary. Keep rows as list for compat.
    for row in result:
        row.setdefault("hot_cutoff_date", meta.get("hot_cutoff_date"))
    return result

@router.get("/summary")
def get_summary(
    shift: Optional[str] = None,
    entry_date: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    station_no: Optional[int] = None,
    machine_id: Optional[int] = None,
    current_operation: Optional[str] = None,
    model: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    from ..history_archive import query_oee_entries_federated, summarize_oee_dicts

    term = (search or model or current_operation or "").strip()
    entries, meta = query_oee_entries_federated(
        db,
        shift=shift,
        entry_date=entry_date,
        date_from=date_from,
        date_to=date_to,
        month=month,
        year=year,
        station_no=station_no,
        machine_id=machine_id,
        term=term,
    )
    out = summarize_oee_dicts(entries)
    out["data_sources"] = meta.get("sources") or ["live"]
    out["hot_cutoff_date"] = meta.get("hot_cutoff_date")
    out["archive_enabled"] = meta.get("archive_enabled")
    return out

def _compute_realtime_oee_for_date(
    db: Session,
    target_date: date,
    cfg: dict,
    *,
    shift: Optional[str] = None,
    station_no: Optional[int] = None,
    machine_id: Optional[int] = None,
    sync_actuals: bool = False,
) -> list:
    """Compute per-machine realtime OEE rows for a single plan date."""
    from .hourly_output import (
        _break_windows, _build_status_segments,
        _countable_running_segments, _running_part_threshold_ratio,
        auto_transition_shift_plans, _shift_window,
        sync_plan_actuals_from_status_logs,
    )

    enabled_shifts = [s for s in cfg.get("shifts", []) if s.get("enabled")]
    if shift:
        enabled_shifts = [s for s in enabled_shifts if s["id"] == shift]
    if not enabled_shifts:
        return []

    _now = datetime.now()
    if sync_actuals:
        for sh_def in enabled_shifts:
            s_start, s_end = _shift_window(target_date, sh_def)
            if s_start <= _now < s_end:
                auto_transition_shift_plans(db, target_date, sh_def['id'], cfg)

        sync_plan_actuals_from_status_logs(
            db,
            entry_date=target_date,
            machine_id=machine_id,
            shift=shift,
            commit=True,
        )

    plan_q = db.query(ProductionPlan).filter(ProductionPlan.plan_date == target_date)
    if station_no:
        plan_q = plan_q.filter(ProductionPlan.station_no == station_no)
    if machine_id:
        plan_q = plan_q.filter(ProductionPlan.machine_id == machine_id)
    all_plans = plan_q.all()
    if not all_plans:
        return []

    wo_ids = {p.work_order_id for p in all_plans if p.work_order_id}
    wo_map = {}
    if wo_ids:
        wos = db.query(WorkOrder).filter(WorkOrder.id.in_(wo_ids)).all()
        wo_map = {w.id: w.work_order_no for w in wos}

    machine_ids = {p.machine_id for p in all_plans if p.machine_id}
    machines = db.query(Machine).filter(Machine.id.in_(machine_ids)).all() if machine_ids else []
    machine_map = {m.id: m for m in machines}

    station_ids = {m.station_id for m in machines}
    stations = db.query(Station).filter(Station.id.in_(station_ids)).all() if station_ids else []
    station_map = {st.id: st for st in stations}

    threshold_ratio = _running_part_threshold_ratio(cfg)

    results = []
    for sh in enabled_shifts:
        sh_id = sh["id"]
        break_cfg = cfg.get("breaks", {}).get(sh_id, {})
        breaks = _break_windows(break_cfg)

        sh_start_hm = sh["start"]
        sh_end_hm = sh["end"]
        start_h, start_min = map(int, sh_start_hm.split(":"))
        end_h, end_min = map(int, sh_end_hm.split(":"))

        shift_start = datetime.combine(target_date, datetime.min.time()).replace(
            hour=start_h, minute=start_min
        )
        overnight = (end_h * 60 + end_min) <= (start_h * 60 + start_min)
        if overnight:
            shift_end = (datetime.combine(target_date, datetime.min.time()) + timedelta(days=1)).replace(
                hour=end_h, minute=end_min
            )
        else:
            shift_end = datetime.combine(target_date, datetime.min.time()).replace(
                hour=end_h, minute=end_min
            )

        shift_total_mins = (shift_end - shift_start).total_seconds() / 60.0
        total_break_mins = sum(b["minutes"] for b in breaks)
        available_mins = shift_total_mins - total_break_mins

        shift_plans = [
            p for p in all_plans
            if p.shift == sh_id and p.machine_id is not None
        ]

        plans_by_machine = {}
        for p in shift_plans:
            plans_by_machine.setdefault(p.machine_id, []).append(p)

        now = datetime.now()
        effective_end = min(shift_end, now)

        for mid, mplans in plans_by_machine.items():
            m = machine_map.get(mid)
            if not m:
                continue
            st = station_map.get(m.station_id)

            segments = _build_status_segments(db, mid, shift_start, effective_end)

            op_mins = 0.0
            for seg in segments:
                if seg["state"] in ("running", "ld_unld"):
                    lo = max(seg["start"], shift_start)
                    hi = min(seg["end"], effective_end)
                    if hi > lo:
                        op_mins += (hi - lo).total_seconds() / 60.0

            running_segs = [s for s in segments if s["state"] == "running"
                            and s["start"] >= shift_start and s["start"] < effective_end]

            primary_ct = 0.0
            for p in mplans:
                ct = float(p.process_time or 0) + float(p.loading_unloading or 0)
                if ct > 0:
                    primary_ct = ct
                    break

            actual_qty = _countable_running_segments(
                running_segs, primary_ct if primary_ct > 0 else None, threshold_ratio,
            )

            total_planned = sum(p.planned_qty or 0 for p in mplans)
            if primary_ct > 0 and available_mins > 0:
                possible_qty = int(available_mins * (60.0 / primary_ct))
            else:
                possible_qty = total_planned

            expected_qty = min(possible_qty, total_planned) if total_planned > 0 else possible_qty

            ar = round(min(op_mins / available_mins * 100, 100.0), 2) if available_mins > 0 else 0.0
            pr = round(actual_qty / expected_qty * 100, 2) if expected_qty > 0 else 0.0
            qr = 100.0
            oee_val = round(ar * pr * qr / 10000, 2)

            prod_loss = max(0, possible_qty - actual_qty)
            accp_qty = actual_qty

            wo_no_list = list({wo_map.get(p.work_order_id, "") for p in mplans if p.work_order_id})
            wo_display = ", ".join(w for w in wo_no_list if w) or "—"

            model_variants = list({(p.model_variant or "").strip() for p in mplans})
            model_display = ", ".join(v for v in model_variants if v) or "—"

            cur_op = mplans[0].current_operation if mplans else ""
            next_op = mplans[0].next_operation if mplans else ""

            results.append({
                "id": f"rt_{mid}_{sh_id}_{target_date.isoformat()}",
                "source": "realtime",
                "entry_date": target_date.isoformat(),
                "station_no": m.station_id,
                "station_name": st.display_name if st else f"Station {m.station_id}",
                "machine_id": mid,
                "machine_name": m.name,
                "shift": sh_id,
                "work_order_no": wo_display,
                "model_variant": model_display,
                "current_operation": cur_op,
                "next_operation": next_op,
                "process_time": float(primary_ct) if primary_ct else 0,
                "loading_unloading": 0,
                "cycle_time": float(primary_ct) if primary_ct else 0,
                "available_shift_time": round(available_mins),
                "operating_time": round(op_mins),
                "possible_qty": possible_qty,
                "actual_qty": actual_qty,
                "production_loss": prod_loss,
                "accp_qty": accp_qty,
                "defect_qty": 0,
                "ar": ar,
                "pr": pr,
                "qr": qr,
                "oee": oee_val,
                "planned_qty": total_planned,
            })

    return results


@router.get("/realtime")
def realtime_oee(
    entry_date: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    shift: Optional[str] = None,
    station_no: Optional[int] = None,
    machine_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Compute per-machine OEE from status logs + production plans (real-time).

    Returns entries compatible with the manual OEE data but flagged as
    ``source: 'realtime'`` so the Dashboard can display both sources.
    """
    from .config import DEFAULT_CONFIG, merge_config, parse_stored_config

    row = db.query(SiteConfig).first()
    if row:
        cfg = merge_config(parse_stored_config(row.config_json))
    else:
        cfg = dict(DEFAULT_CONFIG)

    today = date.today()
    target_dates: list[date] = []

    if date_from and date_to:
        if date_from > date_to:
            date_from, date_to = date_to, date_from
        cursor = date_from
        while cursor <= date_to:
            target_dates.append(cursor)
            cursor += timedelta(days=1)
    else:
        target_dates = [entry_date or today]

    results = []
    for target_date in target_dates:
        results.extend(_compute_realtime_oee_for_date(
            db,
            target_date,
            cfg,
            shift=shift,
            station_no=station_no,
            machine_id=machine_id,
            sync_actuals=(target_date == today),
        ))

    return results


class DefectUpdate(BaseModel):
    defect_qty: int
    note: Optional[str] = ""


class DefectUpdateByMachine(BaseModel):
    machine_id: int
    entry_date: date
    shift: str
    defect_qty: int
    note: Optional[str] = ""


def _load_site_config(db: Session) -> dict:
    from .config import DEFAULT_CONFIG, merge_config, parse_stored_config

    try:
        row = db.query(SiteConfig).first()
        if row:
            return merge_config(parse_stored_config(row.config_json))
    except Exception:
        pass
    return dict(DEFAULT_CONFIG)


def _as_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        return date.fromisoformat(val[:10])
    return val


def _resolve_report_dates(
    db: Session,
    *,
    entry_date: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
) -> List[date]:
    """Resolve which calendar dates a report should cover."""
    if entry_date:
        return [entry_date]
    if date_from or date_to:
        start = date_from or date_to
        end = date_to or date_from
        if start > end:
            start, end = end, start
        out = []
        cur = start
        while cur <= end:
            out.append(cur)
            cur += timedelta(days=1)
        return out
    if month and year:
        _, last = monthrange(year, month)
        return [date(year, month, d) for d in range(1, last + 1)]

    entry_dates = {r[0] for r in db.query(OEEEntry.entry_date).distinct().all() if r[0]}
    plan_dates = {r[0] for r in db.query(ProductionPlan.plan_date).distinct().all() if r[0]}
    return sorted(entry_dates | plan_dates)


def _manual_entry_to_report_row(
    e: OEEEntry,
    *,
    station_name: str = "",
    machine_name: str = "",
    work_order_no: str = "",
    planned_qty=None,
) -> SimpleNamespace:
    return SimpleNamespace(
        id=e.id,
        source="manual",
        entry_date=_as_date(e.entry_date),
        station_no=e.station_no,
        station_name=station_name,
        machine_id=e.machine_id,
        machine_name=machine_name,
        shift=e.shift,
        work_order_no=work_order_no or "",
        model_variant=e.model_variant or "",
        current_operation=e.current_operation or "",
        next_operation=e.next_operation or "",
        process_time=float(e.process_time or 0),
        loading_unloading=float(e.loading_unloading or 0),
        start_time=e.start_time or "",
        stop_time=e.stop_time or "",
        total_minutes=int(e.total_minutes or 0),
        lunch_break=int(e.lunch_break or 0),
        tea_break=int(e.tea_break or 0),
        tpm_cleaning=int(e.tpm_cleaning or 0),
        other_cleaning=int(e.other_cleaning or 0),
        management_meeting=int(e.management_meeting or 0),
        total_breaks=int(getattr(e, "total_breaks", 0) or 0),
        shift_working_minutes=int(getattr(e, "shift_working_minutes", 0) or 0),
        no_load=int(e.no_load or 0),
        new_model_trial=int(e.new_model_trial or 0),
        power_cut=int(e.power_cut or 0),
        planned_maintenance=int(e.planned_maintenance or 0),
        no_manpower_planned=int(e.no_manpower_planned or 0),
        management_loss_total=int(getattr(e, "management_loss_total", 0) or 0),
        available_shift_time=int(e.available_shift_time or 0),
        setting_time=int(e.setting_time or 0),
        tool_change=int(e.tool_change or 0),
        dimension_correction=int(e.dimension_correction or 0),
        scrap_removal=int(e.scrap_removal or 0),
        break_down=int(e.break_down or 0),
        total_down_time=int(getattr(e, "total_down_time", 0) or 0),
        operating_time=int(e.operating_time or 0),
        possible_qty=int(e.possible_qty or 0),
        actual_qty=int(e.actual_qty or 0),
        production_loss=max(0, int(e.possible_qty or 0) - int(e.actual_qty or 0)),
        accp_qty=int(e.accp_qty or 0),
        defect_qty=int(e.defect_qty or 0),
        ar=float(e.ar or 0),
        pr=float(e.pr or 0),
        qr=float(e.qr or 0),
        oee=float(e.oee or 0),
        ar_raw=float(e.ar_raw) if e.ar_raw is not None else None,
        pr_raw=float(e.pr_raw) if e.pr_raw is not None else None,
        qr_raw=float(e.qr_raw) if e.qr_raw is not None else None,
        oee_raw=float(e.oee_raw) if e.oee_raw is not None else None,
        planned_qty=planned_qty,
    )


def _realtime_to_report_row(rt: dict) -> SimpleNamespace:
    possible = int(rt.get("possible_qty") or 0)
    actual = int(rt.get("actual_qty") or 0)
    return SimpleNamespace(
        id=None,
        source="realtime",
        entry_date=_as_date(rt.get("entry_date")),
        station_no=rt.get("station_no"),
        station_name=rt.get("station_name") or "",
        machine_id=rt.get("machine_id"),
        machine_name=rt.get("machine_name") or "",
        shift=rt.get("shift"),
        work_order_no=rt.get("work_order_no") or "",
        model_variant=rt.get("model_variant") or "",
        current_operation=rt.get("current_operation") or "",
        next_operation=rt.get("next_operation") or "",
        process_time=float(rt.get("process_time") or 0),
        loading_unloading=float(rt.get("loading_unloading") or 0),
        start_time="",
        stop_time="",
        total_minutes=0,
        lunch_break=0,
        tea_break=0,
        tpm_cleaning=0,
        other_cleaning=0,
        management_meeting=0,
        total_breaks=0,
        shift_working_minutes=0,
        no_load=0,
        new_model_trial=0,
        power_cut=0,
        planned_maintenance=0,
        no_manpower_planned=0,
        management_loss_total=0,
        available_shift_time=int(rt.get("available_shift_time") or 0),
        setting_time=0,
        tool_change=0,
        dimension_correction=0,
        scrap_removal=0,
        break_down=0,
        total_down_time=0,
        operating_time=int(rt.get("operating_time") or 0),
        possible_qty=possible,
        actual_qty=actual,
        production_loss=max(0, possible - actual),
        accp_qty=int(rt.get("accp_qty") or actual),
        defect_qty=int(rt.get("defect_qty") or 0),
        ar=float(rt.get("ar") or 0),
        pr=float(rt.get("pr") or 0),
        qr=float(rt.get("qr") or 100),
        oee=float(rt.get("oee") or 0),
        ar_raw=None,
        pr_raw=None,
        qr_raw=None,
        oee_raw=None,
        planned_qty=rt.get("planned_qty"),
    )


def collect_merged_oee_report_rows(
    db: Session,
    *,
    shift: Optional[str] = None,
    entry_date: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    station_no: Optional[int] = None,
    machine_id: Optional[int] = None,
    search: Optional[str] = None,
    prefer_live: bool = True,
) -> List[SimpleNamespace]:
    """Build report rows from live capture + data entry.

    Default ``prefer_live=True``: use realtime when available for a
    machine/shift/date key; otherwise fall back to persisted ``OEEEntry``.
    This matches the requirement that emailed/downloaded reports include
    live Dashboard data even when Data Entry was never submitted.
    """
    cfg = _load_site_config(db)
    target_dates = _resolve_report_dates(
        db,
        entry_date=entry_date,
        date_from=date_from,
        date_to=date_to,
        month=month,
        year=year,
    )
    if not target_dates:
        return []

    station_map = {s.id: (s.display_name or s.name) for s in db.query(Station).all()}
    machine_map = {m.id: m.name for m in db.query(Machine).all()}

    # Manual / data-entry rows
    q = db.query(OEEEntry).filter(OEEEntry.entry_date.in_(target_dates))
    if shift:
        q = q.filter(OEEEntry.shift == shift)
    if station_no:
        q = q.filter(OEEEntry.station_no == station_no)
    if machine_id:
        q = q.filter(OEEEntry.machine_id == machine_id)
    term = (search or "").strip()
    if term:
        like = f"%{term}%"
        q = q.filter(or_(
            OEEEntry.current_operation.like(like),
            OEEEntry.model_variant.like(like),
        ))
    manual_entries = q.all()

    all_plans = db.query(ProductionPlan).filter(ProductionPlan.plan_date.in_(target_dates)).all()
    wo_ids = {p.work_order_id for p in all_plans if p.work_order_id}
    wo_map = (
        {w.id: w.work_order_no for w in db.query(WorkOrder).filter(WorkOrder.id.in_(wo_ids)).all()}
        if wo_ids else {}
    )
    plan_lookup = {}
    for p in all_plans:
        key = (p.machine_id, str(p.plan_date), p.shift, p.current_operation or "")
        wo_no = wo_map.get(p.work_order_id, "")
        existing = plan_lookup.get(key)
        if not existing:
            plan_lookup[key] = {"wo": wo_no, "planned": p.planned_qty or 0}
        else:
            existing["planned"] += (p.planned_qty or 0)
            if wo_no and not existing["wo"]:
                existing["wo"] = wo_no

    # Also aggregate planned qty / WO by machine+shift+date for enrichment
    plan_msd = {}
    for p in all_plans:
        k = (p.machine_id, str(p.plan_date), p.shift)
        wo_no = wo_map.get(p.work_order_id, "")
        existing = plan_msd.get(k)
        if not existing:
            plan_msd[k] = {"wo": wo_no, "planned": p.planned_qty or 0}
        else:
            existing["planned"] += (p.planned_qty or 0)
            if wo_no and wo_no not in (existing["wo"] or ""):
                existing["wo"] = ", ".join(x for x in [existing["wo"], wo_no] if x)

    manual_rows = []
    for e in manual_entries:
        pl = plan_lookup.get(
            (e.machine_id, str(e.entry_date), e.shift, e.current_operation or ""),
            {},
        )
        if not pl:
            pl = plan_msd.get((e.machine_id, str(e.entry_date), e.shift), {})
        manual_rows.append(_manual_entry_to_report_row(
            e,
            station_name=station_map.get(e.station_no, str(e.station_no)),
            machine_name=machine_map.get(e.machine_id, "") if e.machine_id else "",
            work_order_no=pl.get("wo", ""),
            planned_qty=pl.get("planned"),
        ))

    # Live / realtime rows
    live_rows = []
    for d in target_dates:
        try:
            rt_list = _compute_realtime_oee_for_date(
                db,
                d,
                cfg,
                shift=shift,
                station_no=station_no,
                machine_id=machine_id,
                sync_actuals=False,
            )
        except Exception as exc:
            log.warning("Realtime OEE failed for %s: %s", d, exc)
            continue
        for rt in rt_list:
            if term:
                blob = f"{rt.get('current_operation') or ''} {rt.get('model_variant') or ''}".lower()
                if term.lower() not in blob:
                    continue
            live_rows.append(_realtime_to_report_row(rt))

    def _key(row: SimpleNamespace):
        ed = row.entry_date.isoformat() if isinstance(row.entry_date, date) else str(row.entry_date)
        return (row.machine_id, row.shift, ed)

    merged: dict = {}
    if prefer_live:
        # Data entry first as base, live overwrites when present
        for row in manual_rows:
            merged[_key(row)] = row
        for row in live_rows:
            merged[_key(row)] = row
    else:
        for row in live_rows:
            merged[_key(row)] = row
        for row in manual_rows:
            merged[_key(row)] = row

    rows = list(merged.values())
    rows.sort(
        key=lambda r: (
            r.entry_date or date.min,
            r.shift or "",
            r.station_no or 0,
            r.machine_name or "",
        ),
        reverse=True,
    )
    return rows


def _find_oee_entry(db: Session, machine_id: int, entry_date: date, shift: str) -> Optional[OEEEntry]:
    return db.query(OEEEntry).filter(
        OEEEntry.machine_id == machine_id,
        OEEEntry.entry_date == entry_date,
        OEEEntry.shift == shift,
    ).first()


def _materialize_oee_from_realtime(
    db: Session,
    machine_id: int,
    entry_date: date,
    shift: str,
    user_id: int,
    cfg: dict,
) -> OEEEntry:
    """Create a persisted OEE row from the live/realtime snapshot (for QC defect updates)."""
    from .hourly_output import _shift_window

    rt_rows = _compute_realtime_oee_for_date(
        db,
        entry_date,
        cfg,
        shift=shift,
        machine_id=machine_id,
        sync_actuals=False,
    )
    rt = next((r for r in rt_rows if r["machine_id"] == machine_id and r["shift"] == shift), None)
    if not rt:
        raise HTTPException(404, "No live OEE data found for this machine/shift/date")

    shift_def = next((s for s in cfg.get("shifts", []) if s.get("id") == shift), None)
    if not shift_def:
        raise HTTPException(400, f"Unknown shift: {shift}")

    shift_start, shift_end = _shift_window(entry_date, shift_def)
    total_minutes = max(1, int((shift_end - shift_start).total_seconds() / 60))

    entry = OEEEntry(
        entry_date=entry_date,
        station_no=rt["station_no"],
        machine_id=machine_id,
        shift=shift,
        current_operation=rt.get("current_operation") or "",
        next_operation=rt.get("next_operation") or "",
        model_variant=rt.get("model_variant") or None,
        process_time=float(rt.get("process_time") or 0),
        loading_unloading=float(rt.get("loading_unloading") or 0),
        start_time=shift_def.get("start", "06:00"),
        stop_time=shift_def.get("end", "14:00"),
        total_minutes=total_minutes,
        available_shift_time=int(rt.get("available_shift_time") or 0),
        operating_time=int(rt.get("operating_time") or 0),
        possible_qty=int(rt.get("possible_qty") or 0),
        actual_qty=int(rt.get("actual_qty") or 0),
        accp_qty=int(rt.get("accp_qty") or rt.get("actual_qty") or 0),
        defect_qty=int(rt.get("defect_qty") or 0),
        ar=float(rt.get("ar") or 0),
        pr=float(rt.get("pr") or 0),
        qr=float(rt.get("qr") or 100),
        oee=float(rt.get("oee") or 0),
        created_by=user_id,
    )
    db.add(entry)
    db.flush()
    return entry


def _apply_defect_update(entry: OEEEntry, defect_qty: int, note: str, user_id: int, db: Session) -> dict:
    before_defect = entry.defect_qty or 0
    before_accp = entry.accp_qty or 0
    before_qr = float(entry.qr or 0)
    before_oee = float(entry.oee or 0)

    actual = entry.actual_qty or 0
    new_defect = max(0, defect_qty)
    if new_defect > actual:
        raise HTTPException(400, "Defect qty cannot exceed actual qty")
    new_accp = max(0, actual - new_defect)
    qr_raw_new = round(new_accp / actual * 100, 2) if actual > 0 else 0
    ar = float(entry.ar or 0)
    pr = float(entry.pr or 0)
    rates_new = _calculate_display_rates(ar, pr, qr_raw_new)
    new_qr = rates_new["qr"]
    oee_raw_new = round(ar * pr * qr_raw_new / 10000, 2)
    new_oee = rates_new["oee"]

    log = OEEDefectLog(
        oee_entry_id=entry.id,
        updated_at=now_ist(),
        updated_by=user_id,
        before_defect_qty=before_defect,
        before_accp_qty=before_accp,
        before_qr=before_qr,
        before_oee=before_oee,
        after_defect_qty=new_defect,
        after_accp_qty=new_accp,
        after_qr=new_qr,
        after_oee=new_oee,
        note=note or "",
    )
    db.add(log)

    entry.defect_qty = new_defect
    entry.accp_qty = new_accp
    entry.qr = new_qr
    entry.qr_raw = qr_raw_new if qr_raw_new > 100 else None
    entry.oee = new_oee
    entry.oee_raw = oee_raw_new if oee_raw_new > new_oee else None
    db.commit()
    db.refresh(entry)
    return {
        "id": entry.id,
        "defect_qty": new_defect,
        "accp_qty": new_accp,
        "qr": new_qr,
        "oee": new_oee,
        "qr_raw": qr_raw_new if qr_raw_new > 100 else None,
        "oee_raw": oee_raw_new if oee_raw_new > new_oee else None,
        "before": {
            "defect_qty": before_defect,
            "accp_qty": before_accp,
            "qr": before_qr,
            "oee": before_oee,
        },
    }


@router.patch("/defect/by-machine")
def update_defect_by_machine(
    data: DefectUpdateByMachine,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    """Update defect qty for live/realtime dashboard rows (any date/shift).

    Materializes a persisted OEE entry from the realtime snapshot when needed.
    """
    entry = _find_oee_entry(db, data.machine_id, data.entry_date, data.shift)
    if not entry:
        cfg = _load_site_config(db)
        entry = _materialize_oee_from_realtime(
            db,
            data.machine_id,
            data.entry_date,
            data.shift,
            user.id,
            cfg,
        )
    return _apply_defect_update(entry, data.defect_qty, data.note or "", user.id, db)


def _format_defect_log(db: Session, entry_id: int) -> list:
    from ..models import User

    logs = db.query(OEEDefectLog).filter(OEEDefectLog.oee_entry_id == entry_id)\
              .order_by(OEEDefectLog.updated_at.desc()).all()
    user_map = {u.id: u.username for u in db.query(User).all()}
    return [{
        "id": l.id,
        "updated_at": l.updated_at.strftime('%Y-%m-%d %H:%M:%S IST'),
        "updated_by": user_map.get(l.updated_by, str(l.updated_by)),
        "before_defect_qty": l.before_defect_qty, "before_accp_qty": l.before_accp_qty,
        "before_qr": float(l.before_qr or 0), "before_oee": float(l.before_oee or 0),
        "after_defect_qty": l.after_defect_qty, "after_accp_qty": l.after_accp_qty,
        "after_qr": float(l.after_qr or 0), "after_oee": float(l.after_oee or 0),
        "note": l.note or ""
    } for l in logs]


@router.get("/defect-log/by-machine")
def get_defect_log_by_machine(
    machine_id: int,
    entry_date: date,
    shift: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    entry = _find_oee_entry(db, machine_id, entry_date, shift)
    if not entry:
        return []
    return _format_defect_log(db, entry.id)


@router.patch("/{entry_id}/defect")
def update_defect(entry_id: int, data: DefectUpdate,
                  db: Session = Depends(get_db),
                  user=Depends(get_current_user)):
    entry = db.query(OEEEntry).filter(OEEEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Entry not found")
    return _apply_defect_update(entry, data.defect_qty, data.note or "", user.id, db)

@router.get("/{entry_id}/defect-log")
def get_defect_log(entry_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    return _format_defect_log(db, entry_id)

@router.get("/download-xlsx")
def download_xlsx(
    shift: Optional[str] = None,
    entry_date: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    station_no: Optional[int] = None,
    machine_id: Optional[int] = None,
    current_operation: Optional[str] = None,
    model: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import Response
    from ..models import OEEDefectLog, User

    term = (search or model or current_operation or "").strip()
    entries = collect_merged_oee_report_rows(
        db,
        shift=shift,
        entry_date=entry_date,
        date_from=date_from,
        date_to=date_to,
        month=month,
        year=year,
        station_no=station_no,
        machine_id=machine_id,
        search=term or None,
        prefer_live=True,
    )

    user_map = {u.id: u.username for u in db.query(User).all()}

    def fmt_ist(dt_val):
        if not dt_val: return ''
        return dt_val.strftime('%d-%m-%Y %H:%M:%S IST')

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    red_f = Font(bold=True, color="DC2626")
    grn_f = Font(bold=True, color="059669")
    amb_f = Font(bold=True, color="D97706")

    def make_header(ws, hdrs):
        ws.append(hdrs)
        for c in range(1, len(hdrs)+1):
            cell = ws.cell(1, c)
            cell.fill = hdr_fill; cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # ── OEE sheet ──
    ws_oee = wb.active
    ws_oee.title = "OEE Report"
    oee_hdrs = ["Date","Station","Machine","Shift","Work Order","Model / Variant",
                "Current Operation","Next Operation","CT (sec)",
                "Avail (min)","Op Time (min)","Plan Qty","Possible Qty","Actual Qty",
                "Prod Loss","Accepted Qty","Defect Qty",
                "AR%","PR%","QR%","OEE%",
                "AR% (original)","PR% (original)","QR% (original)","OEE% (original)",
                "Source"]
    make_header(ws_oee, oee_hdrs)
    for e in entries:
        ct = (e.process_time or 0) + (e.loading_unloading or 0)
        prod_loss = max(0, (e.possible_qty or 0) - (e.actual_qty or 0))
        ar_val  = float(e.ar  or 0)
        pr_val  = float(e.pr  or 0)
        qr_val  = float(e.qr  or 0)
        oee_val = float(e.oee or 0)
        ar_raw  = float(e.ar_raw  or 0) if e.ar_raw is not None else None
        pr_raw  = float(e.pr_raw  or 0) if e.pr_raw is not None else None
        qr_raw  = float(e.qr_raw  or 0) if e.qr_raw is not None else None
        oee_raw = float(e.oee_raw or 0) if e.oee_raw is not None else None
        plan_qty = e.planned_qty if e.planned_qty is not None else ""
        ws_oee.append([
            str(e.entry_date), e.station_name or str(e.station_no or ""),
            e.machine_name or "",
            e.shift, e.work_order_no or "", e.model_variant or "",
            e.current_operation, e.next_operation,
            ct, e.available_shift_time, e.operating_time,
            plan_qty, e.possible_qty, e.actual_qty, prod_loss,
            e.accp_qty, e.defect_qty,
            ar_val, pr_val, qr_val, oee_val,
            ar_raw  if ar_raw  is not None else "—",
            pr_raw  if pr_raw  is not None else "—",
            qr_raw  if qr_raw  is not None else "—",
            oee_raw if oee_raw is not None else "—",
            "Live" if e.source == "realtime" else "Data Entry",
        ])
        ri = ws_oee.max_row
        ws_oee.cell(ri, 21).font = grn_f if oee_val>=85 else (amb_f if oee_val>=65 else red_f)
        for col, raw in [(18, ar_raw), (19, pr_raw), (20, qr_raw), (21, oee_raw)]:
            if raw is not None:
                ws_oee.cell(ri, col).font = amb_f
    for i, w in enumerate([12,14,12,8,18,16,14,14,10,12,14,10,12,12,10,12,12,8,8,8,8,14,14,14,14,12], 1):
        ws_oee.column_dimensions[ws_oee.cell(1,i).column_letter].width = w

    # ── QC Logs sheet ──
    entry_ids = [e.id for e in entries if e.id]
    qc_logs = []
    if entry_ids:
        qc_logs = db.query(OEEDefectLog)\
                    .filter(OEEDefectLog.oee_entry_id.in_(entry_ids))\
                    .order_by(OEEDefectLog.updated_at.desc()).all()
    entry_map = {e.id: e for e in entries if e.id}

    ws_qc = wb.create_sheet("QC Logs")
    qc_hdrs = [
        "Current Operation","Next Operation","Entry Date","Shift","Station",
        "Updated At (IST)","Updated By",
        "Before Defect","Before Accp","Before QR%","Before OEE%",
        "After Defect", "After Accp", "After QR%", "After OEE%",
        "Note"
    ]
    make_header(ws_qc, qc_hdrs)

    if not qc_logs:
        ws_qc.append(["No QC updates found for the selected filters."])
    else:
        for l in qc_logs:
            e = entry_map.get(l.oee_entry_id)
            before_oee = float(l.before_oee or 0)
            after_oee  = float(l.after_oee  or 0)
            ws_qc.append([
                e.current_operation if e else "", e.next_operation if e else "",
                str(e.entry_date) if e else "", e.shift if e else "",
                (e.station_name if e else "") or (str(e.station_no) if e else ""),
                fmt_ist(l.updated_at),
                user_map.get(l.updated_by, str(l.updated_by) if l.updated_by else ""),
                l.before_defect_qty, l.before_accp_qty,
                f"{float(l.before_qr or 0):.2f}%", f"{before_oee:.2f}%",
                l.after_defect_qty,  l.after_accp_qty,
                f"{float(l.after_qr  or 0):.2f}%", f"{after_oee:.2f}%",
                l.note or "",
            ])
            ri = ws_qc.max_row
            bf = ws_qc.cell(ri, 11)
            bf.font = grn_f if before_oee>=85 else (amb_f if before_oee>=65 else red_f)
            af = ws_qc.cell(ri, 15)
            af.font = grn_f if after_oee>=85  else (amb_f if after_oee>=65  else red_f)

    for i, w in enumerate([18,18,12,8,14,22,14,14,12,12,12,12,12,12,12,35], 1):
        ws_qc.column_dimensions[ws_qc.cell(1,i).column_letter].width = w

    # ── OEE Daywise sheet ──
    from collections import defaultdict
    day_groups = defaultdict(list)
    for e in entries:
        day_groups[(str(e.entry_date), e.station_no)].append(e)

    ws_day = wb.create_sheet("OEE Daywise")
    day_hdrs = ["Date","Station","AR%","PR%","QR%","OEE%","Total Produced","Accepted Qty","Defects"]
    make_header(ws_day, day_hdrs)
    for (dt, pno), grp in sorted(day_groups.items()):
        n = len(grp)
        avg_ar  = round(sum(float(e.ar  or 0) for e in grp) / n, 2)
        avg_pr  = round(sum(float(e.pr  or 0) for e in grp) / n, 2)
        avg_qr  = round(sum(float(e.qr  or 0) for e in grp) / n, 2)
        avg_oee = round(sum(float(e.oee or 0) for e in grp) / n, 2)
        tot_act  = sum(e.actual_qty  or 0 for e in grp)
        tot_accp = sum(e.accp_qty    or 0 for e in grp)
        tot_def  = sum(e.defect_qty  or 0 for e in grp)
        station_label = grp[0].station_name if grp else str(pno)
        ws_day.append([dt, station_label,
                       avg_ar, avg_pr, avg_qr, avg_oee,
                       tot_act, tot_accp, tot_def])
        ri = ws_day.max_row
        ws_day.cell(ri, 6).font = grn_f if avg_oee >= 85 else (amb_f if avg_oee >= 65 else red_f)
    for i, w in enumerate([12,14,8,8,8,8,14,14,10], 1):
        ws_day.column_dimensions[ws_day.cell(1,i).column_letter].width = w

    # ── OEE Shiftwise sheet ──
    shift_groups = defaultdict(list)
    for e in entries:
        shift_groups[(str(e.entry_date), e.station_no, e.shift)].append(e)

    ws_shift = wb.create_sheet("OEE Shiftwise")
    shift_hdrs = ["Date","Station","Shift","Actual","Prod Loss","Accepted","Defects","AR%","PR%","QR%","OEE%"]
    make_header(ws_shift, shift_hdrs)
    for (dt, pno, sh), grp in sorted(shift_groups.items()):
        n = len(grp)
        tot_act  = sum(e.actual_qty  or 0 for e in grp)
        tot_loss = sum(max(0, (e.possible_qty or 0) - (e.actual_qty or 0)) for e in grp)
        tot_accp = sum(e.accp_qty    or 0 for e in grp)
        tot_def  = sum(e.defect_qty  or 0 for e in grp)
        avg_ar   = round(sum(float(e.ar  or 0) for e in grp) / n, 2)
        avg_pr   = round(sum(float(e.pr  or 0) for e in grp) / n, 2)
        avg_qr   = round(sum(float(e.qr  or 0) for e in grp) / n, 2)
        avg_oee  = round(sum(float(e.oee or 0) for e in grp) / n, 2)
        station_label = grp[0].station_name if grp else str(pno)
        ws_shift.append([dt, station_label, sh,
                         tot_act, tot_loss, tot_accp, tot_def,
                         avg_ar, avg_pr, avg_qr, avg_oee])
        ri = ws_shift.max_row
        ws_shift.cell(ri, 11).font = grn_f if avg_oee >= 85 else (amb_f if avg_oee >= 65 else red_f)
    for i, w in enumerate([12,14,8,10,10,12,10,8,8,8,8], 1):
        ws_shift.column_dimensions[ws_shift.cell(1,i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    label = str(entry_date or date_from or month or "report")
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="oee_report_{label}.xlsx"'}
    )


@router.get("/download")
def download_csv(
    shift: Optional[str] = None,
    entry_date: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    station_no: Optional[int] = None,
    machine_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    entries = collect_merged_oee_report_rows(
        db,
        shift=shift,
        entry_date=entry_date,
        month=month,
        year=year,
        station_no=station_no,
        machine_id=machine_id,
        prefer_live=True,
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Date","Station","Machine","Shift","Work Order","Model / Variant",
                     "Current Operation","Next Operation","CT",
                     "Start","Stop","Total Min",
                     "Lunch","Tea","TPM","Other Clean","Mgmt Mtg","No Load","New Model","Power Cut",
                     "Planned Maint","No Manpower","Avail Time","Setting","Tool Change","Dim Corr",
                     "Scrap","Breakdown","Operating Time","Plan Qty","Possible Qty","Actual Qty","Accp Qty",
                     "Defect Qty","AR%","PR%","QR%","OEE%",
                     "AR% (original)","PR% (original)","QR% (original)","OEE% (original)",
                     "Source"])
    for e in entries:
        plan_qty = e.planned_qty if e.planned_qty is not None else ""
        writer.writerow([e.entry_date, e.station_name or str(e.station_no or ""),
                         e.machine_name or "",
                         e.shift, e.work_order_no or "", e.model_variant or "",
                         e.current_operation, e.next_operation,
                         (e.process_time or 0)+(e.loading_unloading or 0), e.start_time, e.stop_time,
                         e.total_minutes, e.lunch_break, e.tea_break, e.tpm_cleaning, e.other_cleaning,
                         e.management_meeting, e.no_load, e.new_model_trial, e.power_cut,
                         e.planned_maintenance, e.no_manpower_planned, e.available_shift_time,
                         e.setting_time, e.tool_change, e.dimension_correction, e.scrap_removal,
                         e.break_down, e.operating_time, plan_qty,
                         e.possible_qty, e.actual_qty, e.accp_qty,
                         e.defect_qty, e.ar, e.pr, e.qr, e.oee,
                         float(e.ar_raw  or 0) if e.ar_raw  is not None else "",
                         float(e.pr_raw  or 0) if e.pr_raw  is not None else "",
                         float(e.qr_raw  or 0) if e.qr_raw  is not None else "",
                         float(e.oee_raw or 0) if e.oee_raw is not None else "",
                         "Live" if e.source == "realtime" else "Data Entry"])
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=oee_report.csv"})
