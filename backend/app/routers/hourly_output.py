"""Machine hourly output aggregation for production dashboard."""
from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session
from datetime import date, datetime, timedelta
from typing import Optional
import json
import io
from ..models import Machine, Station, ProductionPlan, MachineStatusLog, SiteConfig, OEEEntry, ModelChangeRequest, Part, get_db, now_ist
from ..auth import get_current_user
from .config import DEFAULT_CONFIG, merge_config

router = APIRouter(prefix="/api/hourly-output", tags=["hourly-output"])

LD_UNLD_MAX_SEC = 60
MICRO_GAP_SEC = 15
STATE_LABELS = [
    ('running', 'Running'),
    ('expected', 'Exp Hourly Output'),
    ('ar', 'AR %'),
    ('pr', 'PR %'),
    ('qr', 'QR %'),
    ('oee', 'OEE %'),
]

_TIMED_BREAK_DEFAULTS = {
    'lunch_start': '13:00', 'lunch_end': '13:30', 'lunch_break': 30,
    'tea_start': '10:00', 'tea_end': '10:10', 'tea_break': 10,
    'tpm_start': '17:00', 'tpm_end': '17:10', 'tpm_cleaning': 10,
}


def _previous_shift_info(cfg: dict, current_shift_id: str, entry_date: date):
    """Return (prev_shift_id, prev_date) for the shift immediately before the current one."""
    shifts = [s for s in cfg.get('shifts', []) if s.get('enabled', True)]
    ordered = sorted(shifts, key=lambda s: _parse_mins(s['start']))
    ids = [s['id'] for s in ordered]
    if current_shift_id not in ids:
        return None, None
    idx = ids.index(current_shift_id)
    if idx == 0:
        return ids[-1], entry_date - timedelta(days=1)
    return ids[idx - 1], entry_date


def _shift_order_index(cfg: dict, shift_id: str) -> int:
    shifts = [s for s in cfg.get('shifts', []) if s.get('enabled', True)]
    ordered = sorted(shifts, key=lambda s: _parse_mins(s['start']))
    ids = [s['id'] for s in ordered]
    try:
        return ids.index(shift_id)
    except ValueError:
        return -1


def _find_continuous_same_part_prior(
    db: Session, machine_id: int, model_variant: str,
    entry_date: date, shift_id: str, cfg: dict, lookback_days: int = 14,
):
    """Most recent started prior plan on this machine. Same part = continuous.

    Continuity means no intervening different part was run on the machine.
    Day-1 of a new part still requires Start / MCR; subsequent same-part
    shifts/days auto-start. Pending-only plans do not break continuity.
    """
    since = entry_date - timedelta(days=lookback_days)
    candidates = (
        db.query(ProductionPlan)
        .filter(
            ProductionPlan.machine_id == machine_id,
            ProductionPlan.plan_date >= since,
            ProductionPlan.plan_date <= entry_date,
            ProductionPlan.status.in_(['running', 'completed', 'paused', 'incomplete', 'aborted']),
        )
        .all()
    )
    cur_ord = _shift_order_index(cfg, shift_id)

    def _sort_key(p):
        return (p.plan_date, _shift_order_index(cfg, p.shift), p.id or 0)

    prior = [
        p for p in candidates
        if (
            p.plan_date < entry_date
            or (
                p.plan_date == entry_date
                and _shift_order_index(cfg, p.shift) < cur_ord
            )
        )
    ]
    if not prior:
        return None

    latest = max(prior, key=_sort_key)
    if not latest.model_variant or latest.model_variant != model_variant:
        return None
    return latest


def auto_transition_shift_plans(db: Session, entry_date: date, shift_id: str, cfg: dict):
    """Auto-start plans when the same part continues from a prior shift/day.

    Rules:
    - If the most recent prior plan on the machine used the same model_variant
      (running/completed/paused), the current pending plan is auto-started
      (no model-change request needed).
    - Still-running prior plan is auto-completed.
    - Works across days and overnight shift boundaries (C → next-day A).
    - New / changed parts remain pending until Start / MCR.
    """
    current_plans = (
        db.query(ProductionPlan)
        .filter(
            ProductionPlan.plan_date == entry_date,
            ProductionPlan.shift == shift_id,
            ProductionPlan.status == 'pending',
        )
        .all()
    )
    if not current_plans:
        return

    changed = False
    _now = now_ist()
    for plan in current_plans:
        if not plan.machine_id or not plan.model_variant:
            continue

        prev_plan = _find_continuous_same_part_prior(
            db, plan.machine_id, plan.model_variant, entry_date, shift_id, cfg,
        )
        if not prev_plan:
            # Fallback: immediate previous shift only (legacy path)
            prev_shift_id, prev_date = _previous_shift_info(cfg, shift_id, entry_date)
            if prev_shift_id:
                prev_plan = (
                    db.query(ProductionPlan)
                    .filter(
                        ProductionPlan.machine_id == plan.machine_id,
                        ProductionPlan.plan_date == prev_date,
                        ProductionPlan.shift == prev_shift_id,
                        ProductionPlan.model_variant == plan.model_variant,
                        ProductionPlan.status.in_(['running', 'completed', 'paused', 'incomplete', 'aborted']),
                    )
                    .first()
                )
        if not prev_plan:
            continue

        plan.status = 'running'
        plan.updated_at = _now

        if prev_plan.status in ('running', 'paused'):
            prev_actual = int(prev_plan.actual_qty or 0)
            prev_planned = int(prev_plan.planned_qty or 0)
            # Shift/continuity handoff: full qty → completed, shortfall → incomplete
            prev_plan.status = (
                'completed' if prev_planned > 0 and prev_actual >= prev_planned else 'incomplete'
            )
            prev_plan.updated_at = _now

        conflicting = (
            db.query(ProductionPlan)
            .filter(
                ProductionPlan.machine_id == plan.machine_id,
                ProductionPlan.id != plan.id,
                ProductionPlan.plan_date == entry_date,
                ProductionPlan.shift == shift_id,
                ProductionPlan.status == 'running',
            )
            .all()
        )
        for c in conflicting:
            c.status = 'paused'
            c.updated_at = _now

        changed = True

    if changed:
        try:
            db.commit()
        except Exception:
            db.rollback()


def _load_config(db: Session) -> dict:
    row = db.query(SiteConfig).first()
    if not row:
        return dict(DEFAULT_CONFIG)
    try:
        stored = json.loads(row.config_json)
    except (json.JSONDecodeError, TypeError):
        return dict(DEFAULT_CONFIG)
    return merge_config(stored)


def _float_ct(process_time, loading_unloading) -> float:
    """Cycle time in seconds = process_time + loading_unloading."""
    return float(process_time or 0) + float(loading_unloading or 0)


def _plan_ct(plan) -> float:
    """Cycle time from production planning sheet: process_time + L&U."""
    if plan is None:
        return 0.0
    stored = getattr(plan, 'cycle_time', None)
    if stored is not None and float(stored) > 0:
        return float(stored)
    return _float_ct(plan.process_time, plan.loading_unloading)


def _ct_for_plan_variant(plans: list, variant: str) -> float:
    for p in plans:
        if _plan_variant(p) == variant:
            ct = _plan_ct(p)
            if ct > 0:
                return ct
    return 0.0


def _parse_mins(hhmm: str) -> int:
    h, m = map(int, hhmm.split(':'))
    return h * 60 + m


def _shift_slots(shift: dict) -> list:
    start_m = _parse_mins(shift['start'])
    end_m = _parse_mins(shift['end'])
    overnight = end_m <= start_m
    total = (24 * 60 - start_m + end_m) if overnight else max(0, end_m - start_m)
    slots = []
    for i in range(0, total, 60):
        from_m = (start_m + i) % (24 * 60)
        to_m = (from_m + 60) % (24 * 60)
        from_h = from_m // 60
        to_h = to_m // 60
        slots.append({
            'label': f"{from_h:02d}:00-{to_h:02d}:00",
            'from_minutes': from_m,
            'slot_index': i // 60,
        })
    return slots


def _break_windows(break_cfg: dict) -> list:
    """Timed breaks from Configuration — start/end windows deduct from hourly slots."""
    merged = {**_TIMED_BREAK_DEFAULTS, **(break_cfg or {})}
    defs = [
        ('lunch', 'lunch_start', 'lunch_end', 'lunch_break'),
        ('tea', 'tea_start', 'tea_end', 'tea_break'),
        ('tpm', 'tpm_start', 'tpm_end', 'tpm_cleaning'),
    ]
    out = []
    for key, sk, ek, mk in defs:
        start = merged.get(sk) or ''
        end = merged.get(ek) or ''
        mins = merged.get(mk) or 0
        if start and end:
            diff = _parse_mins(end) - _parse_mins(start)
            if diff < 0:
                diff += 24 * 60
            mins = diff or mins
        if start and mins > 0:
            out.append({'key': key, 'start_m': _parse_mins(start), 'end_m': _parse_mins(end), 'minutes': mins})
    return out


def _overlap(a0, a1, b0, b1) -> int:
    def ranges(s, e):
        return [(s, e)] if e > s else [(s, 24 * 60), (0, e)]
    total = 0
    for as_, ae in ranges(a0, a1):
        for bs, be in ranges(b0, b1):
            lo, hi = max(as_, bs), min(ae, be)
            if hi > lo:
                total += hi - lo
    return total


def _avail_in_slot(from_m: int, breaks: list) -> int:
    deducted = sum(_overlap(from_m, from_m + 60, b['start_m'], b['end_m']) for b in breaks)
    return max(0, 60 - deducted)


def _expected_parts(ct: float, mins: float) -> int:
    """Floor parts count — avoids rounding hourly slots above planned shift qty."""
    if not ct or ct <= 0 or mins <= 0:
        return 0
    return int(mins * (60 / ct))


def _slot_bounds(shift_start: datetime, shift_end: datetime, slot_index: int):
    slot_start = shift_start + timedelta(hours=slot_index)
    slot_end = min(shift_start + timedelta(hours=slot_index + 1), shift_end)
    return slot_start, slot_end


def _hourly_avail_minutes(slots: list, shift_start: datetime, shift_end: datetime, breaks: list) -> list:
    """Productive minutes per hourly slot after lunch/tea/TPM windows from config."""
    weights = []
    for s in slots:
        lo, hi = _slot_bounds(shift_start, shift_end, s['slot_index'])
        weights.append(_mins_available(lo, hi, breaks))
    return weights


def _untimed_break_minutes(break_cfg: dict) -> int:
    """Minutes-only breaks from Configuration (no start/end window)."""
    if not break_cfg:
        return 0
    return int(break_cfg.get('other_cleaning') or 0) + int(break_cfg.get('management_meeting') or 0)


def _scale_weights_for_untimed(weights: list, untimed_mins: int) -> list:
    """Reduce slot weights proportionally when untimed breaks are configured."""
    total = sum(weights)
    if total <= 0 or untimed_mins <= 0:
        return list(weights)
    factor = max(0.0, (total - untimed_mins) / total)
    return [w * factor for w in weights]


def _expected_hourly_slots(
    slots: list,
    shift_start: datetime,
    shift_end: datetime,
    breaks: list,
    planned_total: int,
    windows: list,
    break_cfg: Optional[dict] = None,
) -> tuple:
    """
    Expected output per hourly slot.

    When planned_total > 0, distribute planned qty across all slots in proportion
    to productive minutes (shift window minus lunch/tea/TPM and untimed breaks).
    When planned_total is 0, fall back to CT × available minutes capacity.
    """
    avail_mins_total = sum(_hourly_avail_minutes(slots, shift_start, shift_end, breaks))
    slot_weights = []
    ct_capacity = []
    for s in slots:
        lo, hi = _slot_bounds(shift_start, shift_end, s['slot_index'])
        slot_weights.append(_slot_avail_minutes(windows, lo, hi, breaks))
        ct_capacity.append(_slot_expected(windows, lo, hi, breaks))

    if planned_total > 0:
        weights = _scale_weights_for_untimed(slot_weights, _untimed_break_minutes(break_cfg))
        if sum(weights) <= 0:
            weights = _scale_weights_for_untimed(
                _hourly_avail_minutes(slots, shift_start, shift_end, breaks),
                _untimed_break_minutes(break_cfg),
            )
        distributed = _distribute_to_target(weights, planned_total)
        # Do not dump a variant's full plan qty into a short occupancy window
        # (e.g. 15 min before a model change). Cap each slot at CT capacity.
        capped = []
        for i, qty in enumerate(distributed):
            cap = ct_capacity[i] if i < len(ct_capacity) else 0
            capped.append(min(qty, cap) if cap > 0 else qty)
        return capped, sum(capped), avail_mins_total

    return ct_capacity, sum(ct_capacity), avail_mins_total


def _distribute_to_target(raw: list, target: int) -> list:
    """Scale hourly weights so they sum exactly to target (largest remainder)."""
    n = len(raw)
    if n == 0:
        return []
    if target <= 0:
        return [0] * n
    raw_sum = sum(raw)
    if raw_sum <= 0:
        base, rem = divmod(target, n)
        return [base + (1 if i < rem else 0) for i in range(n)]
    scaled, remainders, allocated = [], [], 0
    for i, r in enumerate(raw):
        exact = target * r / raw_sum
        flo = int(exact)
        scaled.append(flo)
        remainders.append((exact - flo, i))
        allocated += flo
    diff = target - allocated
    for _, idx in sorted(remainders, key=lambda x: x[0], reverse=True)[:max(0, diff)]:
        scaled[idx] += 1
    return scaled


def _hhmm_str(value) -> Optional[str]:
    """Normalize DB time (str, time, timedelta) to HH:MM."""
    if value is None:
        return None
    if isinstance(value, str):
        return value[:5] if len(value) >= 5 else value
    if isinstance(value, datetime):
        return value.strftime('%H:%M')
    from datetime import time as dt_time, timedelta
    if isinstance(value, dt_time):
        return value.strftime('%H:%M')
    if isinstance(value, timedelta):
        total = int(value.total_seconds()) % (24 * 3600)
        h, rem = divmod(total, 3600)
        m = rem // 60
        return f'{h:02d}:{m:02d}'
    return str(value)[:5]


def _time_in_shift(entry_date: date, hhmm, shift_start: datetime, shift_end: datetime) -> datetime:
    hhmm = _hhmm_str(hhmm)
    if not hhmm:
        return shift_start
    h, m = map(int, hhmm.split(':'))
    dt = datetime.combine(entry_date, datetime.min.time().replace(hour=h, minute=m))
    if dt < shift_start and shift_end.date() > shift_start.date():
        dt += timedelta(days=1)
    return max(shift_start, min(dt, shift_end))


def _mins_available(lo: datetime, hi: datetime, breaks: list) -> float:
    span = (hi - lo).total_seconds() / 60
    if span <= 0:
        return 0.0
    from_m = lo.hour * 60 + lo.minute
    deducted = sum(_overlap(from_m, from_m + int(span), b['start_m'], b['end_m']) for b in breaks)
    return max(0.0, span - deducted)


def _plan_variant(p) -> str:
    return (getattr(p, 'model_variant', None) or '').strip()


def _norm_variant(value) -> str:
    text = (value or '').strip()
    if not text or text == '—':
        return ''
    return text


def _clip_dt(dt: Optional[datetime], start: datetime, end: datetime) -> Optional[datetime]:
    if dt is None:
        return None
    if dt < start:
        return start
    if dt > end:
        return end
    return dt


def _append_variant_window(windows: list, ws: datetime, we: datetime, variant: str, plans: list):
    variant = _norm_variant(variant)
    if not variant or we <= ws:
        return
    if windows and windows[-1]['variant'] == variant and windows[-1]['end'] == ws:
        windows[-1]['end'] = we
        return
    windows.append({
        'start': ws,
        'end': we,
        'ct': _ct_for_plan_variant(plans, variant),
        'variant': variant,
    })


def _production_windows_from_mcrs(
    shift_start: datetime,
    shift_end: datetime,
    plans: list,
    mcrs: list,
) -> Optional[list]:
    """Build actual occupancy windows from approved/completed model changes.

    Example: PEEN 14:30–14:45, setting change 14:45–14:55, MP101 14:55–shift end.
    Changeover gaps are omitted so produced counts and expected output stay
    on the variant that actually ran in that hour.
    """
    timed = [
        m for m in (mcrs or [])
        if getattr(m, 'status', None) in ('approved', 'in_progress', 'completed')
        and getattr(m, 'start_time', None)
    ]
    if not timed:
        return None
    timed.sort(key=lambda m: (m.start_time or datetime.min, m.id or 0))

    windows = []
    cursor = shift_start
    current = _norm_variant(getattr(timed[0], 'from_model', None))

    for mcr in timed:
        t0 = _clip_dt(mcr.start_time, shift_start, shift_end)
        if t0 is None:
            continue
        from_v = _norm_variant(getattr(mcr, 'from_model', None))
        to_v = _norm_variant(getattr(mcr, 'to_model', None))
        if from_v:
            current = from_v
        if t0 > cursor and current:
            _append_variant_window(windows, cursor, t0, current, plans)
        t1 = _clip_dt(mcr.end_time, shift_start, shift_end) if mcr.end_time else None
        if t1 is None:
            # Setting change still in progress — no production after t0
            return windows or None
        cursor = max(t1, t0)
        current = to_v or current

    if current and cursor < shift_end:
        _append_variant_window(windows, cursor, shift_end, current, plans)
    return windows or None


def _oee_variant(e) -> str:
    return (getattr(e, 'model_variant', None) or '').strip()


def _collect_variants(plans: list, oee_entries: list, mcrs: list) -> list:
    variants, seen = [], set()
    active_plans = [p for p in plans if getattr(p, 'status', None) in ('running', 'completed', 'paused', 'incomplete', 'aborted')]
    for p in sorted(active_plans, key=lambda x: (getattr(x, 'priority', None) or 1, x.id)):
        v = _plan_variant(p)
        if v and v not in seen:
            seen.add(v)
            variants.append(v)
    for e in sorted(oee_entries, key=lambda x: (_hhmm_str(x.start_time) or '', x.id)):
        v = _oee_variant(e)
        if v and v not in seen:
            seen.add(v)
            variants.append(v)
    for m in sorted(mcrs, key=lambda x: (x.start_time or x.created_at or datetime.min, x.id)):
        if m.status not in ('approved', 'in_progress', 'completed'):
            continue
        for v in (m.from_model, m.to_model):
            v = (v or '').strip()
            if v and v not in seen:
                seen.add(v)
                variants.append(v)
    return variants


def _collect_cycle_times(plans: list, oee_entries: list) -> list:
    """Cycle times from active plans; OEE only if no active plan."""
    cts, seen = [], set()
    active = [p for p in plans if getattr(p, 'status', None) in ('running', 'completed', 'paused', 'incomplete', 'aborted')]
    for p in sorted(active, key=lambda x: (getattr(x, 'priority', None) or 1, x.id)):
        ct = round(_plan_ct(p), 2)
        if ct > 0 and ct not in seen:
            seen.add(ct)
            cts.append(ct)
    if not cts:
        for e in oee_entries:
            ct = round(_float_ct(e.process_time, e.loading_unloading), 2)
            if ct > 0 and ct not in seen:
                seen.add(ct)
                cts.append(ct)
    return cts


def _production_windows(
    shift_start: datetime,
    shift_end: datetime,
    plans: list,
    oee_entries: list,
    mcrs: Optional[list] = None,
) -> list:
    """Time windows with CT per variant.

    Priority:
    1. Timed OEE entries (operator-entered start/stop)
    2. Model-change occupancy (actual run + setting-change gaps)
    3. Equal split of active plans (no changeover evidence yet)
    Pending plans are excluded.
    """
    windows = []
    timed = [e for e in oee_entries if _hhmm_str(e.start_time) and _hhmm_str(e.stop_time)]
    if timed:
        for e in sorted(timed, key=lambda x: _hhmm_str(x.start_time) or ''):
            ws = _time_in_shift(shift_start.date(), e.start_time, shift_start, shift_end)
            we = _time_in_shift(shift_start.date(), e.stop_time, shift_start, shift_end)
            if we <= ws:
                continue
            variant = _oee_variant(e)
            ct = _ct_for_plan_variant(plans, variant) or _float_ct(e.process_time, e.loading_unloading)
            windows.append({'start': ws, 'end': we, 'ct': ct, 'variant': variant})
    else:
        mcr_windows = _production_windows_from_mcrs(shift_start, shift_end, plans, mcrs or [])
        if mcr_windows:
            windows = mcr_windows
        else:
            active = [p for p in plans if getattr(p, 'status', None) in ('running', 'completed', 'paused', 'incomplete', 'aborted')]
            if active:
                completed = [p for p in active if getattr(p, 'status', None) in ('completed', 'paused')]
                running = [p for p in active if getattr(p, 'status', None) == 'running']
                ordered = sorted(completed, key=lambda x: (getattr(x, 'priority', None) or 1, x.id)) + \
                          sorted(running, key=lambda x: (getattr(x, 'priority', None) or 1, x.id))
                n = len(ordered)
                span_sec = (shift_end - shift_start).total_seconds()
                for i, p in enumerate(ordered):
                    ws = shift_start + timedelta(seconds=span_sec * i / n)
                    we = shift_end if i == n - 1 else shift_start + timedelta(seconds=span_sec * (i + 1) / n)
                    windows.append({'start': ws, 'end': we, 'ct': _plan_ct(p), 'variant': _plan_variant(p)})
    if not windows:
        windows.append({'start': shift_start, 'end': shift_end, 'ct': 0.0, 'variant': ''})
    return windows


def _effective_ct(windows: list) -> float:
    total_mins = sum((w['end'] - w['start']).total_seconds() / 60 for w in windows)
    if total_mins <= 0:
        return 0.0
    weighted = sum(
        (w['end'] - w['start']).total_seconds() / 60 * w['ct'] for w in windows
    )
    return weighted / total_mins


def _variant_planned_qty(plans: list, variant: str) -> int:
    return sum(
        p.planned_qty or 0 for p in plans
        if _plan_variant(p) == variant and getattr(p, 'status', None) in ('running', 'completed', 'paused', 'incomplete', 'aborted')
    )


def _variant_ct(variant: str, windows: list, plans: list, oee_entries: list) -> float:
    for w in windows:
        if w.get('variant') == variant and w.get('ct'):
            return float(w['ct'])
    ct = _ct_for_plan_variant(plans, variant)
    if ct > 0:
        return ct
    for e in oee_entries:
        if _oee_variant(e) == variant:
            ct = _float_ct(e.process_time, e.loading_unloading)
            if ct > 0:
                return ct
    return 0.0


def _status_windows_for_variant(
    variant: str,
    var_windows: list,
    all_windows: list,
    shift_start: datetime,
    shift_end: datetime,
) -> list:
    """Narrow status attribution when a full-shift variant overlaps shorter variant windows."""
    if not var_windows:
        return []
    shift_span = (shift_end - shift_start).total_seconds() or 1.0
    out = []
    for w in var_windows:
        span = (w['end'] - w['start']).total_seconds()
        if span >= shift_span * 0.85 and len(all_windows) > 1:
            other_ends = [
                ow['end'] for ow in all_windows
                if ow.get('variant') != variant and ow.get('end')
            ]
            clip_start = max(other_ends) if other_ends else w['start']
            if clip_start < w['end']:
                out.append({**w, 'start': max(w['start'], clip_start)})
        else:
            out.append(w)
    return out


def _slot_overlaps_windows(windows: list, slot_start: datetime, slot_end: datetime) -> bool:
    for w in windows:
        lo = max(w['start'], slot_start)
        hi = min(w['end'], slot_end)
        if hi > lo:
            return True
    return False


def _slot_parts_in_windows(
    segments: list,
    var_windows: list,
    slot_start: datetime,
    slot_end: datetime,
    state_filter: set,
    ct: float,
    threshold_ratio: float = 0.3,
) -> int:
    """Count state transitions whose start falls in slot and within variant windows.

    For running segments, applies the same duration threshold as the main
    count_transitions_through to avoid counting micro-runs as parts.
    """
    if not var_windows or not _slot_overlaps_windows(var_windows, slot_start, slot_end):
        return 0
    relevant = []
    for seg in segments:
        if seg['state'] not in state_filter:
            continue
        if seg.get('prior'):
            continue
        if seg['start'] < slot_start or seg['start'] >= slot_end:
            continue
        for w in var_windows:
            if seg['start'] >= w['start'] and seg['start'] < w['end']:
                relevant.append(seg)
                break
    if state_filter == {'running'} and ct and ct > 0:
        return _countable_running_segments(relevant, ct, threshold_ratio)
    return len(relevant)


def _slot_state_minutes(segments, slot_start, slot_end, state_filter):
    """Sum segment durations overlapping with the slot for the given states."""
    total = 0.0
    for seg in segments:
        if seg['state'] not in state_filter:
            continue
        lo = max(seg['start'], slot_start)
        hi = min(seg['end'], slot_end)
        if hi > lo:
            total += (hi - lo).total_seconds() / 60.0
    return total


def _slot_state_mins_windowed(segments, var_windows, slot_start, slot_end, state_filter):
    """Sum durations of matching states within both the hourly slot and the variant windows."""
    if not var_windows:
        return 0.0
    total = 0.0
    for seg in segments:
        if seg['state'] not in state_filter:
            continue
        seg_lo = max(seg['start'], slot_start)
        seg_hi = min(seg['end'], slot_end)
        if seg_hi <= seg_lo:
            continue
        for w in var_windows:
            lo = max(seg_lo, w['start'])
            hi = min(seg_hi, w['end'])
            if hi > lo:
                total += (hi - lo).total_seconds() / 60.0
    return total


def _machine_qr(oee_entries):
    """Average QR from OEE data entry, defaults to 100 if none."""
    if not oee_entries:
        return 100.0
    vals = [float(e.qr) for e in oee_entries if e.qr is not None]
    return sum(vals) / len(vals) if vals else 100.0


def _compute_oee_slots(running_inc, expected, operating_mins, avail_mins, qr_val):
    """Compute per-slot AR, PR, QR, OEE arrays.
    AR = operating_time / available_time (operating = available - downtime).
    PR = actual parts / expected parts.
    """
    ar_arr, pr_arr, qr_arr, oee_arr = [], [], [], []
    for i in range(len(running_inc)):
        avail = avail_mins[i] if i < len(avail_mins) else 60.0
        op = operating_mins[i] if i < len(operating_mins) else 0.0
        actual = running_inc[i]
        exp = expected[i] if i < len(expected) else 0

        ar = round(min(op / avail * 100, 100.0), 1) if avail > 0 else 0.0
        pr = round(actual / exp * 100, 1) if exp > 0 else 0.0
        qr = round(qr_val, 1)
        oee_v = round(ar * pr * qr / 10000, 1)

        ar_arr.append(ar)
        pr_arr.append(pr)
        qr_arr.append(qr)
        oee_arr.append(oee_v)
    return ar_arr, pr_arr, qr_arr, oee_arr


def _oee_shift_totals(avail_mins, operating_mins, total_actual, total_expected, qr_val):
    """Compute shift-level OEE totals."""
    total_avail = sum(avail_mins)
    total_op = sum(operating_mins)
    ar = round(min(total_op / total_avail * 100, 100.0), 1) if total_avail > 0 else 0.0
    pr = round(total_actual / total_expected * 100, 1) if total_expected > 0 else 0.0
    qr = round(qr_val, 1)
    oee = round(ar * pr * qr / 10000, 1)
    return ar, pr, qr, oee


def _build_variant_breakdown(
    variants: list,
    windows: list,
    slots: list,
    shift_start: datetime,
    shift_end: datetime,
    breaks: list,
    segments: list,
    machine_plans: list,
    machine_oee: list,
    qr_val: float = 100.0,
    running_threshold_ratio: float = 0.3,
    break_cfg: Optional[dict] = None,
) -> list:
    if not variants:
        return []

    shift_span = (shift_end - shift_start).total_seconds() or 1.0
    breakdown = []

    for variant in variants:
        planned = _variant_planned_qty(machine_plans, variant)
        var_windows = [w for w in windows if w.get('variant') == variant]
        if not var_windows and len(variants) == 1:
            var_windows = list(windows)

        ct = _variant_ct(variant, windows, machine_plans, machine_oee)

        expected_hourly, exp_shift_total, _ = _expected_hourly_slots(
            slots, shift_start, shift_end, breaks, planned, var_windows, break_cfg,
        )

        running_hourly, ld_hourly, idle_hourly = [], [], []
        status_windows = _status_windows_for_variant(
            variant, var_windows, windows, shift_start, shift_end,
        )
        for s in slots:
            slot_start = shift_start + timedelta(hours=s['slot_index'])
            slot_end = min(shift_start + timedelta(hours=s['slot_index'] + 1), shift_end)
            running_hourly.append(_slot_parts_in_windows(
                segments, status_windows, slot_start, slot_end, {'running'}, ct,
                threshold_ratio=running_threshold_ratio,
            ))
            ld_hourly.append(_slot_parts_in_windows(
                segments, status_windows, slot_start, slot_end, {'ld_unld'}, ct,
            ))
            idle_hourly.append(_slot_parts_in_windows(
                segments, status_windows, slot_start, slot_end, {'idle'}, ct,
            ))

        # OEE metrics per variant per slot — AR uses operating time (running + ld_unld)
        var_op_mins = []
        var_avail_mins = []
        for s in slots:
            s_start = shift_start + timedelta(hours=s['slot_index'])
            s_end = min(shift_start + timedelta(hours=s['slot_index'] + 1), shift_end)
            var_op_mins.append(_slot_state_mins_windowed(
                segments, status_windows, s_start, s_end, {'running', 'ld_unld'},
            ))
            var_avail_mins.append(_mins_available(s_start, s_end, breaks))
        ar_h, pr_h, qr_h, oee_h = _compute_oee_slots(
            running_hourly, expected_hourly, var_op_mins, var_avail_mins, qr_val,
        )
        ar_t, pr_t, qr_t, oee_t = _oee_shift_totals(
            var_avail_mins, var_op_mins, sum(running_hourly), exp_shift_total, qr_val,
        )

        if var_windows:
            w_start = min(w['start'] for w in var_windows)
            w_end = max(w['end'] for w in var_windows)
        else:
            w_start, w_end = shift_start, shift_end

        breakdown.append({
            'variant': variant,
            'planned_qty': planned,
            'is_current': False,
            '_w_start': w_start,
            '_w_end': w_end,
            'cycle_time': ct,
            'cycle_time_display': _fmt_ct_val(ct) if ct else '0',
            'window_start': w_start.strftime('%H:%M'),
            'window_end': w_end.strftime('%H:%M'),
            'timeline_pct_start': round((w_start - shift_start).total_seconds() / shift_span * 100, 2),
            'timeline_pct_end': round((w_end - shift_start).total_seconds() / shift_span * 100, 2),
            'states': {
                'running': running_hourly,
                'ld_unld': ld_hourly,
                'idle': idle_hourly,
                'expected': expected_hourly,
                'ar': ar_h,
                'pr': pr_h,
                'qr': qr_h,
                'oee': oee_h,
            },
            'shift_totals': {
                'running': sum(running_hourly),
                'ld_unld': sum(ld_hourly),
                'idle': sum(idle_hourly),
                'expected': exp_shift_total,
                'ar': ar_t,
                'pr': pr_t,
                'qr': qr_t,
                'oee': oee_t,
            },
        })

    running_plan_variants = {
        (_plan_variant(p))
        for p in machine_plans
        if getattr(p, 'status', None) == 'running'
    }

    _now = datetime.now()
    current_idx = None
    for i, bd in enumerate(breakdown):
        ws, we = bd.pop('_w_start'), bd.pop('_w_end')
        if bd['variant'] in running_plan_variants:
            current_idx = i
    if current_idx is None:
        for i, bd in enumerate(breakdown):
            if bd.get('planned_qty', 0) > 0:
                ws_check = [w for w in windows if w.get('variant') == bd['variant']]
                if ws_check:
                    w_end = max(w['end'] for w in ws_check)
                    if _now <= w_end:
                        current_idx = i
    if current_idx is None and breakdown:
        current_idx = len(breakdown) - 1
    if current_idx is not None:
        breakdown[current_idx]['is_current'] = True

    return breakdown


def _slot_avail_minutes(windows: list, slot_start: datetime, slot_end: datetime, breaks: list) -> float:
    """Productive minutes in slot across production windows (breaks deducted)."""
    total = 0.0
    for w in windows:
        lo = max(w['start'], slot_start)
        hi = min(w['end'], slot_end)
        if hi <= lo:
            continue
        total += _mins_available(lo, hi, breaks)
    return total


def _slot_expected(windows: list, slot_start: datetime, slot_end: datetime, breaks: list) -> int:
    total = 0
    for w in windows:
        lo = max(w['start'], slot_start)
        hi = min(w['end'], slot_end)
        if hi <= lo:
            continue
        avail = _mins_available(lo, hi, breaks)
        total += _expected_parts(w['ct'], avail)
    return total


def _fmt_ct_val(ct) -> str:
    c = float(ct)
    if c == int(c):
        return str(int(c))
    return f"{c:.2f}".rstrip('0').rstrip('.')


def _ct_display_str(cycle_times: list) -> Optional[str]:
    if not cycle_times:
        return None
    if len(cycle_times) == 1:
        return _fmt_ct_val(cycle_times[0])
    return ' · '.join(_fmt_ct_val(c) for c in cycle_times)


def _shift_window(entry_date: date, shift: dict):
    sh, sm = map(int, shift['start'].split(':'))
    eh, em = map(int, shift['end'].split(':'))
    start = datetime.combine(entry_date, datetime.min.time().replace(hour=sh, minute=sm))
    if eh <= sh:
        end = datetime.combine(entry_date + timedelta(days=1), datetime.min.time().replace(hour=eh, minute=em))
    else:
        end = datetime.combine(entry_date, datetime.min.time().replace(hour=eh, minute=em))
    return start, end


def _effective_shift_end(shift_start: datetime, shift_end: datetime) -> datetime:
    """Cap an in-progress shift at current IST time; full window for completed shifts."""
    now = now_ist()
    if now <= shift_start:
        return shift_start
    if now >= shift_end:
        return shift_end
    return now


def _get_cycle_profile(db: Session, variant: str) -> Optional[dict]:
    """Return cycle_profile for a part only if it has interruptions > 0 configured.
    Returns None for all normal parts (no profile or interruptions=0).
    """
    if not variant:
        return None
    part = db.query(Part).filter(
        (Part.part_no == variant) | (Part.model_variant == variant),
        Part.active == 1,
    ).first()
    if not part or not getattr(part, 'cycle_profile_json', None):
        return None
    try:
        profile = json.loads(part.cycle_profile_json)
        # Only activate stitching when explicitly configured with interruptions > 0
        if isinstance(profile, dict) and int(profile.get('interruptions') or 0) > 0:
            return profile
    except Exception:
        pass
    return None


def _build_status_segments(
    db: Session,
    machine_id: int,
    shift_start: datetime,
    effective_end: datetime,
    cycle_profile: Optional[dict] = None,
    ld_unld_max_sec: int = LD_UNLD_MAX_SEC,
    micro_gap_sec: int = MICRO_GAP_SEC,
) -> list:
    """Status segments from shift_start through effective_end.
    cycle_profile is only passed for parts that explicitly have interruptions > 0.
    For all other parts (cycle_profile=None) the raw segments are returned unchanged.
    """
    if effective_end <= shift_start:
        return []

    prior = (
        db.query(MachineStatusLog)
        .filter(
            MachineStatusLog.machine_id == machine_id,
            MachineStatusLog.changed_at < shift_start,
        )
        .order_by(MachineStatusLog.changed_at.desc())
        .first()
    )
    logs = (
        db.query(MachineStatusLog)
        .filter(
            MachineStatusLog.machine_id == machine_id,
            MachineStatusLog.changed_at >= shift_start,
            MachineStatusLog.changed_at <= effective_end,
        )
        .order_by(MachineStatusLog.changed_at.asc())
        .all()
    )

    if not prior and not logs:
        return []

    timeline: list = []
    if prior:
        timeline.append((shift_start, prior.status))
    elif logs:
        # No prior carry-over: project first in-shift status back to shift open.
        # Do not also append logs[0] below — timestamps differ so dedup would miss it.
        timeline.append((shift_start, logs[0].status))

    # When we seeded from logs[0] at shift_start, skip that log to avoid a duplicate
    # segment boundary at logs[0].changed_at with the same status.
    for log in (logs[1:] if (not prior and logs) else logs):
        if timeline and timeline[-1][0] == log.changed_at and timeline[-1][1] == log.status:
            continue
        timeline.append((log.changed_at, log.status))

    raw_segments = []
    for i, (t_start, status) in enumerate(timeline):
        t_end = timeline[i + 1][0] if i + 1 < len(timeline) else effective_end
        seg_start = max(t_start, shift_start)
        seg_end = min(t_end, effective_end)
        if seg_end <= seg_start:
            continue
        dur = int((seg_end - seg_start).total_seconds())
        # Mark segments whose status was carried over from before the shift start
        # (prior log entry). These represent machine state at shift open, not a
        # completed cycle within this shift, so they must not be counted as parts.
        is_prior_carryover = (i == 0 and prior is not None and t_start == shift_start)
        raw_segments.append({
            'state': _classify(status, dur, ld_unld_max_sec),
            'start': seg_start,
            'end': seg_end,
            'seconds': dur,
            'prior': is_prior_carryover,
        })

    raw_segments = _merge_micro_gaps(raw_segments, micro_gap_sec)

    if not cycle_profile or not raw_segments:
        return raw_segments

    return _stitch_segments(raw_segments, cycle_profile)


def _classify(status: str, duration_sec: int, ld_unld_max: int = LD_UNLD_MAX_SEC) -> str:
    if status == 'idle' and duration_sec < ld_unld_max:
        return 'ld_unld'
    if status in ('running', 'idle'):
        return status
    return 'idle'


def _merge_micro_gaps(segments: list, gap_sec: int = MICRO_GAP_SEC) -> list:
    """Auto-merge running segments separated by very short ld_unld gaps.

    Handles brief corrections where the operator stops for a few seconds
    without removing the part, then restarts. Without this, each stop/start
    would be counted as a separate part.
    """
    if len(segments) < 3 or gap_sec <= 0:
        return segments

    out = []
    i = 0
    n = len(segments)
    while i < n:
        if (i + 2 < n
                and not segments[i].get('prior')
                and segments[i]['state'] == 'running'
                and segments[i + 1]['state'] == 'ld_unld'
                and segments[i + 1]['seconds'] < gap_sec
                and segments[i + 2]['state'] == 'running'):
            merged = {
                'state': 'running',
                'start': segments[i]['start'],
                'end': segments[i + 2]['end'],
                'seconds': segments[i]['seconds'] + segments[i + 1]['seconds'] + segments[i + 2]['seconds'],
                'prior': False,
            }
            i += 3
            while (i + 1 < n
                   and segments[i]['state'] == 'ld_unld'
                   and segments[i]['seconds'] < gap_sec
                   and segments[i + 1]['state'] == 'running'):
                merged['end'] = segments[i + 1]['end']
                merged['seconds'] += segments[i]['seconds'] + segments[i + 1]['seconds']
                i += 2
            out.append(merged)
        else:
            out.append(segments[i])
            i += 1
    return out


def _running_part_threshold_ratio(cfg: Optional[dict]) -> float:
    """Read the running-part threshold from config as a percentage (default 30%)."""
    if not cfg:
        return 0.3
    hourly_cfg = cfg.get('hourly_output') or {}
    raw = hourly_cfg.get('running_part_threshold_pct', cfg.get('running_part_threshold_pct', 30))
    try:
        pct = float(raw)
    except (TypeError, ValueError):
        return 0.3
    if pct <= 0:
        return 0.0
    return min(max(pct / 100.0, 0.0), 1.0)


def _cfg_ld_unld_max_sec(cfg: Optional[dict]) -> int:
    """Idle duration below this (in seconds) is classified as Ld/UnLd. Default 60."""
    if not cfg:
        return LD_UNLD_MAX_SEC
    hourly_cfg = cfg.get('hourly_output') or {}
    try:
        return max(1, int(hourly_cfg.get('ld_unld_max_sec', LD_UNLD_MAX_SEC)))
    except (TypeError, ValueError):
        return LD_UNLD_MAX_SEC


def _cfg_micro_gap_sec(cfg: Optional[dict]) -> int:
    """Ld/UnLd gaps shorter than this (in seconds) between two running segments
    are auto-merged into a single running segment. Default 15. Set 0 to disable."""
    if not cfg:
        return MICRO_GAP_SEC
    hourly_cfg = cfg.get('hourly_output') or {}
    try:
        return max(0, int(hourly_cfg.get('micro_gap_sec', MICRO_GAP_SEC)))
    except (TypeError, ValueError):
        return MICRO_GAP_SEC


def _countable_running_segments(segments: list, process_time_sec: Optional[float], ratio: float = 0.3) -> int:
    """Count running segments only when they cross the configured threshold.

    Segments tagged 'prior=True' are carry-over state from before the shift
    and are never counted as completed parts.
    """
    if not process_time_sec or process_time_sec <= 0:
        return sum(1 for seg in segments
                   if seg.get('state') == 'running' and not seg.get('prior'))

    threshold_sec = max(1, int(float(process_time_sec) * ratio))
    return sum(1 for seg in segments
               if seg.get('state') == 'running'
               and not seg.get('prior')
               and seg.get('seconds', 0) >= threshold_sec)


def sync_plan_actuals_from_status_logs(
    db: Session,
    *,
    entry_date: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    machine_id: Optional[int] = None,
    shift: Optional[str] = None,
    commit: bool = True,
) -> int:
    """Write dashboard-style countable running parts into production_plans.actual_qty.

    Dashboard OEE computes Actual from status logs but does not persist it.
    Work orders and planning screens read plan.actual_qty, so keep them in sync.

    Auto-sync uses max(existing, computed) so a higher manual edit is preserved
    until machine count catches up. Returns number of plans updated.
    """
    cfg = _load_config(db)
    shifts = [s for s in cfg.get('shifts', []) if s.get('enabled', True)]
    if not shifts:
        return 0
    shift_map = {s['id']: s for s in shifts}
    if shift and shift not in shift_map:
        return 0

    q = db.query(ProductionPlan).filter(
        ProductionPlan.machine_id.isnot(None),
        ProductionPlan.status.in_(['running', 'completed', 'paused', 'incomplete', 'aborted']),
    )
    if entry_date:
        q = q.filter(ProductionPlan.plan_date == entry_date)
    if date_from:
        q = q.filter(ProductionPlan.plan_date >= date_from)
    if date_to:
        q = q.filter(ProductionPlan.plan_date <= date_to)
    if machine_id:
        q = q.filter(ProductionPlan.machine_id == machine_id)
    if shift:
        q = q.filter(ProductionPlan.shift == shift)

    plans = q.all()
    if not plans:
        return 0

    # Cap sync window to recent days for unbounded list calls
    if not entry_date and not date_from and not date_to:
        today = now_ist().date()
        plans = [p for p in plans if p.plan_date and p.plan_date >= today - timedelta(days=7)]
        if not plans:
            return 0

    threshold_ratio = _running_part_threshold_ratio(cfg)
    ld_max = _cfg_ld_unld_max_sec(cfg)
    micro_gap = _cfg_micro_gap_sec(cfg)

    groups = {}
    for p in plans:
        key = (p.machine_id, p.plan_date, p.shift)
        groups.setdefault(key, []).append(p)

    status_rank = {'running': 0, 'paused': 1, 'completed': 2}
    updated = 0
    touched_wo_ids = set()
    _now = now_ist()

    for (mid, pdate, sh_id), group in groups.items():
        sh_def = shift_map.get(sh_id)
        if not sh_def or not pdate:
            continue

        shift_start, shift_end = _shift_window(pdate, sh_def)
        effective_end = _effective_shift_end(shift_start, shift_end)
        if effective_end <= shift_start:
            continue

        group.sort(key=lambda p: (status_rank.get(p.status, 9), p.priority or 9999, p.id))
        target = group[0]

        mcrs = db.query(ModelChangeRequest).filter(
            ModelChangeRequest.machine_id == mid,
            ModelChangeRequest.entry_date == pdate,
            ModelChangeRequest.shift == sh_id,
            ModelChangeRequest.status.in_(['approved', 'in_progress', 'completed']),
        ).all()
        windows = _production_windows(shift_start, shift_end, group, [], mcrs)

        profile = _get_cycle_profile(db, getattr(target, 'model_variant', None) or '')
        segments = _build_status_segments(
            db, mid, shift_start, effective_end,
            cycle_profile=profile,
            ld_unld_max_sec=ld_max,
            micro_gap_sec=micro_gap,
        )

        for plan in group:
            variant = _plan_variant(plan)
            var_windows = [w for w in windows if w.get('variant') == variant] if variant else []
            if not var_windows and len(group) == 1:
                var_windows = list(windows)
            ct = _plan_ct(plan)
            running_segs = []
            for seg in segments:
                if seg.get('state') != 'running' or seg.get('prior'):
                    continue
                if seg['start'] < shift_start or seg['start'] >= effective_end:
                    continue
                if not var_windows:
                    running_segs.append(seg)
                    continue
                for w in var_windows:
                    if seg['start'] >= w['start'] and seg['start'] < w['end']:
                        running_segs.append(seg)
                        break
            computed = _countable_running_segments(
                running_segs, ct if ct > 0 else None, threshold_ratio,
            )
            new_actual = max(int(plan.actual_qty or 0), int(computed or 0))
            if new_actual != int(plan.actual_qty or 0):
                plan.actual_qty = new_actual
                plan.updated_at = _now
                updated += 1
                if plan.work_order_id:
                    touched_wo_ids.add(plan.work_order_id)
                if (
                    plan.status == 'running'
                    and plan.planned_qty
                    and new_actual >= plan.planned_qty
                ):
                    plan.status = 'completed'

    if touched_wo_ids:
        from .work_orders import sync_work_order_after_plan_change
        for wo_id in touched_wo_ids:
            sync_work_order_after_plan_change(db, wo_id)

    if updated and commit:
        try:
            db.commit()
        except Exception:
            db.rollback()
            return 0

    return updated


def _stitch_segments(segments: list, profile: Optional[dict]) -> list:
    """Merge multi-segment cycles. Returns segments unchanged when profile is
    None or interruptions <= 0 — so normal parts are never affected.

    Full pattern matched (oldest-first):
        Running  →  (Ld/UnLd → Running) × N

    The leading Running may optionally be a micro-run (<=threshold_sec).
    For VMC-style parts the leading Running is a full operation block —
    threshold_sec=0 means any Running qualifies as the cycle opener.
    """
    if not profile or not segments:
        return segments
    interruptions = int(profile.get('interruptions') or 0)
    threshold_sec = int(profile.get('micro_run_threshold_sec') or 0)
    if interruptions <= 0:
        return segments

    out = []
    i = 0
    n = len(segments)
    while i < n:
        pos = i
        group = []

        # Leading Running: required as cycle opener.
        # If threshold_sec > 0, only accept it when duration <= threshold (micro-run).
        # If threshold_sec == 0, accept any Running as the opener.
        if pos < n and segments[pos]['state'] == 'running':
            if threshold_sec == 0 or segments[pos]['seconds'] <= threshold_sec:
                group.append(segments[pos])
                pos += 1

        # Require all N (ld_unld -> running) pairs
        matched = len(group) > 0
        for _ in range(interruptions):
            if pos + 1 >= n:
                matched = False
                break
            if segments[pos]['state'] != 'ld_unld':
                matched = False
                break
            if segments[pos + 1]['state'] != 'running':
                matched = False
                break
            group.append(segments[pos])
            group.append(segments[pos + 1])
            pos += 2

        if matched and len(group) >= 2:
            merged = {
                'state': 'running',
                'start': group[0]['start'],
                'end': group[-1]['end'],
                'seconds': sum(g['seconds'] for g in group),
            }
            out.append(merged)
            i = pos
        else:
            out.append(segments[i])
            i += 1
    return out


def _line_meta(cfg: dict, line_id: str):
    for factory in (cfg.get('factory') or {}).get('factories') or []:
        for dept in factory.get('departments') or []:
            for line in dept.get('lines') or []:
                if line.get('enabled') is False:
                    continue
                if line.get('id') == line_id:
                    label = ' / '.join(filter(None, [factory.get('name'), dept.get('name'), line.get('name')]))
                    return {
                        'label': label,
                        'station_ids': set(line.get('stationIds') or []),
                        'factory_name': factory.get('name') or '',
                    }
    return None


def _factory_station_ids(cfg: dict, factory_id: str) -> set:
    ids = set()
    for factory in (cfg.get('factory') or {}).get('factories') or []:
        if factory.get('id') != factory_id:
            continue
        for dept in factory.get('departments') or []:
            for line in dept.get('lines') or []:
                ids.update(line.get('stationIds') or [])
    return ids


def _scope_label(scope: str, cfg: dict, station_id, line_id, factory_id, db: Session) -> str:
    if scope == 'station' and station_id:
        st = db.query(Station).filter(Station.id == station_id).first()
        return f"Station: {st.display_name if st else station_id}"
    if scope == 'line' and line_id:
        meta = _line_meta(cfg, line_id)
        return f"Line: {meta['label'] if meta else line_id}"
    if scope == 'factory' and factory_id:
        for f in (cfg.get('factory') or {}).get('factories') or []:
            if f.get('id') == factory_id:
                return f"Factory: {f.get('name') or factory_id}"
    return 'All machines (entire fleet)'


def _filter_machines(machines: list, cfg: dict, scope: str, station_id, line_id, factory_id):
    if scope == 'station' and station_id:
        return [m for m in machines if m.station_id == station_id]
    if scope == 'line' and line_id:
        meta = _line_meta(cfg, line_id)
        if not meta:
            return machines
        return [
            m for m in machines
            if m.station_id in meta['station_ids'] or (m.location and m.location == meta['label'])
        ]
    if scope == 'factory' and factory_id:
        station_ids = _factory_station_ids(cfg, factory_id)
        if not station_ids:
            return machines
        return [m for m in machines if m.station_id in station_ids]
    return machines


def build_hourly_output(
    db: Session,
    entry_date: date,
    shift: str,
    scope: str = 'all',
    station_id: Optional[int] = None,
    line_id: Optional[str] = None,
    factory_id: Optional[str] = None,
):
    cfg = _load_config(db)
    shifts = [s for s in cfg.get('shifts', []) if s.get('enabled', True)]
    shift_def = next((s for s in shifts if s['id'] == shift), None)

    if shift_def:
        s_start, s_end = _shift_window(entry_date, shift_def)
        _now = now_ist()
        if s_start <= _now < s_end:
            auto_transition_shift_plans(db, entry_date, shift, cfg)
    if not shift_def:
        return {
            'slots': [], 'machines': [], 'scope': scope,
            'scope_label': _scope_label(scope, cfg, station_id, line_id, factory_id, db),
        }

    break_cfg = (cfg.get('breaks') or {}).get(shift, {})
    breaks = _break_windows(break_cfg)
    running_part_threshold_ratio = _running_part_threshold_ratio(cfg)
    ld_unld_max = _cfg_ld_unld_max_sec(cfg)
    micro_gap = _cfg_micro_gap_sec(cfg)
    shift_start, shift_end = _shift_window(entry_date, shift_def)
    effective_end = _effective_shift_end(shift_start, shift_end)
    is_live = shift_start < effective_end < shift_end
    slots = _shift_slots(shift_def)
    shift_avail_mins = sum(_hourly_avail_minutes(slots, shift_start, shift_end, breaks))
    shift_total_mins = int((shift_end - shift_start).total_seconds() / 60)

    all_machines = db.query(Machine).order_by(Machine.station_id, Machine.id).all()
    all_machines = [
        m for m in all_machines
        if int(getattr(m, 'is_enabled', 1) or 0) != 0
    ]
    machines = _filter_machines(all_machines, cfg, scope, station_id, line_id, factory_id)

    plans = db.query(ProductionPlan).filter(
        ProductionPlan.plan_date == entry_date,
        ProductionPlan.shift == shift,
    ).all()
    plans_by_machine: dict = {}
    for p in plans:
        if p.machine_id:
            plans_by_machine.setdefault(p.machine_id, []).append(p)

    oee_all = db.query(OEEEntry).filter(
        OEEEntry.entry_date == entry_date,
        OEEEntry.shift == shift,
    ).all()
    oee_by_machine: dict = {}
    for e in oee_all:
        if e.machine_id:
            oee_by_machine.setdefault(e.machine_id, []).append(e)

    mcrs_all = db.query(ModelChangeRequest).filter(
        ModelChangeRequest.entry_date == entry_date,
        ModelChangeRequest.shift == shift,
        ModelChangeRequest.status.in_(['approved', 'in_progress', 'completed']),
    ).all()
    mcr_by_machine: dict = {}
    for mcr in mcrs_all:
        mcr_by_machine.setdefault(mcr.machine_id, []).append(mcr)

    from ..operator_presence import get_live_operator_map, operator_fields_for_machine
    _hourly_op_map = get_live_operator_map(db, entry_date=entry_date, shift_id=shift)

    result_machines = []
    for m in machines:
        station = db.query(Station).filter(Station.id == m.station_id).first()
        machine_plans = plans_by_machine.get(m.id, [])
        machine_oee = oee_by_machine.get(m.id, [])
        machine_mcrs = mcr_by_machine.get(m.id, [])

        variants = _collect_variants(machine_plans, machine_oee, machine_mcrs)
        cycle_times = _collect_cycle_times(machine_plans, machine_oee)
        planned_total = sum(p.planned_qty or 0 for p in machine_plans)

        windows = _production_windows(shift_start, shift_end, machine_plans, machine_oee, machine_mcrs)
        eff_ct = _effective_ct(windows) if windows else 0.0
        primary_ct = _plan_ct(machine_plans[0]) if len(machine_plans) == 1 else eff_ct
        if not primary_ct and cycle_times:
            primary_ct = cycle_times[0]

        expected_hourly, exp_shift_total, _ = _expected_hourly_slots(
            slots, shift_start, shift_end, breaks, planned_total, windows, break_cfg,
        )

        cycle_profile = None
        active_variants = [
            _plan_variant(p) for p in machine_plans
            if getattr(p, 'status', None) in ('running', 'completed', 'paused', 'incomplete', 'aborted')
        ]
        if active_variants:
            profiles = [_get_cycle_profile(db, v) for v in active_variants]
            # All variants must have the same non-None profile for stitching to apply
            non_null = [p for p in profiles if p is not None]
            if non_null and len(non_null) == len(profiles):
                # Check all profiles are identical (same interruptions + threshold)
                first = non_null[0]
                all_same = all(
                    p.get('interruptions') == first.get('interruptions') and
                    p.get('micro_run_threshold_sec') == first.get('micro_run_threshold_sec')
                    for p in non_null
                )
                if all_same:
                    cycle_profile = first

        segments = _build_status_segments(db, m.id, shift_start, effective_end, cycle_profile, ld_unld_max, micro_gap)

        def count_transitions_through(slot_end_dt: datetime, state_filter):
            """Count transitions (occurrences) from shift_start up to slot_end_dt."""
            slot_end_dt = min(slot_end_dt, effective_end)
            relevant = [
                seg for seg in segments
                if seg['state'] in state_filter
                and not seg.get('prior')
                and seg['start'] >= shift_start
                and seg['start'] < slot_end_dt
            ]
            if state_filter == {'running'}:
                threshold_reference_sec = primary_ct if primary_ct else None
                return _countable_running_segments(relevant, threshold_reference_sec, running_part_threshold_ratio)
            return len(relevant)

        running_cum, ld_cum, idle_cum = [], [], []
        for s in slots:
            slot_end_dt = min(shift_start + timedelta(hours=s['slot_index'] + 1), shift_end, effective_end)
            running_cum.append(count_transitions_through(slot_end_dt, {'running'}))
            ld_cum.append(count_transitions_through(slot_end_dt, {'ld_unld'}))
            idle_cum.append(count_transitions_through(slot_end_dt, {'idle'}))

        ct_display = _ct_display_str(cycle_times)

        # Per-slot OEE metrics
        # AR = Operating Time / Available Time
        # Operating Time = time in running + ld_unld (productive states)
        # Downtime = idle + setting_change + breakdown + alarm + offline
        qr_val = _machine_qr(machine_oee)
        avail_mins = _hourly_avail_minutes(slots, shift_start, shift_end, breaks)
        operating_mins_per_slot = []
        for s in slots:
            s_start = shift_start + timedelta(hours=s['slot_index'])
            s_end = min(shift_start + timedelta(hours=s['slot_index'] + 1), shift_end, effective_end)
            operating_mins_per_slot.append(
                _slot_state_minutes(segments, s_start, s_end, {'running', 'ld_unld'})
            )
        running_inc = [running_cum[0]] + [max(0, running_cum[i] - running_cum[i - 1]) for i in range(1, len(running_cum))]
        ar_hourly, pr_hourly, qr_hourly, oee_hourly = _compute_oee_slots(
            running_inc, expected_hourly, operating_mins_per_slot, avail_mins, qr_val,
        )
        ar_tot, pr_tot, qr_tot, oee_tot = _oee_shift_totals(
            avail_mins, operating_mins_per_slot,
            running_cum[-1] if running_cum else 0, exp_shift_total, qr_val,
        )

        variant_breakdown = _build_variant_breakdown(
            variants, windows, slots, shift_start, shift_end, breaks,
            segments, machine_plans, machine_oee, qr_val,
            running_threshold_ratio=running_part_threshold_ratio,
            break_cfg=break_cfg,
        )

        from ..operator_presence import operator_fields_for_machine
        _op = operator_fields_for_machine(_hourly_op_map, m.id)

        result_machines.append({
            'machine_id': m.id,
            'machine_name': m.name,
            'machine_status': m.status or 'idle',
            'station_id': m.station_id,
            'station_name': station.display_name if station else str(m.station_id),
            'location': m.location or '',
            'model_variants': variants,
            'model_variant': ' · '.join(variants) if variants else '',
            'variant_breakdown': variant_breakdown,
            'planned_qty': planned_total,
            'cycle_time': primary_ct if primary_ct else 0,
            'cycle_times': cycle_times,
            'cycle_time_display': ct_display,
            'process_time': machine_plans[0].process_time if machine_plans else 0,
            'loading_unloading': machine_plans[0].loading_unloading if machine_plans else 0,
            'cycle_time_source': 'production_plan' if machine_plans else ('oee' if machine_oee else 'none'),
            **_op,
            'states': {
                'running': running_cum,
                'ld_unld': ld_cum,
                'idle': idle_cum,
                'expected': expected_hourly,
                'ar': ar_hourly,
                'pr': pr_hourly,
                'qr': qr_hourly,
                'oee': oee_hourly,
            },
            'shift_totals': {
                'running': running_cum[-1] if running_cum else 0,
                'ld_unld': ld_cum[-1] if ld_cum else 0,
                'idle': idle_cum[-1] if idle_cum else 0,
                'expected': exp_shift_total,
                'ar': ar_tot,
                'pr': pr_tot,
                'qr': qr_tot,
                'oee': oee_tot,
            },
        })

    return {
        'entry_date': str(entry_date),
        'shift': shift,
        'shift_name': shift_def.get('name', shift),
        'shift_start': shift_def['start'],
        'shift_end': shift_def['end'],
        'as_of': effective_end.isoformat(sep=' '),
        'is_live': is_live,
        'shift_total_minutes': shift_total_mins,
        'shift_available_minutes': round(shift_avail_mins, 1),
        'break_windows': [
            {'key': b['key'], 'start_m': b['start_m'], 'end_m': b['end_m'], 'minutes': b['minutes']}
            for b in breaks
        ],
        'scope': scope,
        'scope_label': _scope_label(scope, cfg, station_id, line_id, factory_id, db),
        'machine_count': len(result_machines),
        'total_fleet': len(all_machines),
        'slots': [s['label'] for s in slots],
        'machines': result_machines,
    }


def _build_xlsx(payload: dict) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hourly Output'

    hdr_fill = PatternFill('solid', fgColor='1E3A5F')
    hdr_font = Font(bold=True, color='FFFFFF')
    total_fill = PatternFill('solid', fgColor='FEF08A')

    slots = payload.get('slots') or []
    headers = ['Station', 'Machine', 'Model/Variant', 'Planned Qty', 'Set Cycle time', 'State'] + slots + ['Shift Total']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.freeze_panes = 'A2'

    for m in payload.get('machines') or []:
        start_row = ws.max_row + 1
        for key, label in STATE_LABELS:
            row = [
                m.get('station_name', ''),
                m.get('machine_name', ''),
                m.get('model_variant', '') or '—',
                m.get('planned_qty', 0) or '—',
                m.get('cycle_time_display') or m.get('cycle_time', '') or '—',
                label,
            ]
            row.extend((m.get('states') or {}).get(key, []))
            row.append((m.get('shift_totals') or {}).get(key, 0))
            ws.append(row)

        end_row = ws.max_row
        if end_row > start_row:
            for col in range(1, 6):
                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                ws.cell(start_row, col).alignment = Alignment(vertical='center')

        total_col = len(headers)
        for r in range(start_row, end_row + 1):
            ws.cell(r, total_col).fill = total_fill
            ws.cell(r, total_col).font = Font(bold=True)

    for i, w in enumerate([14, 14, 16, 12, 14, 18] + [12] * len(slots) + [12], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    meta = wb.create_sheet('Report Info')
    meta.append(['Date', payload.get('entry_date', '')])
    meta.append(['Shift', f"{payload.get('shift_name', '')} ({payload.get('shift_start', '')}–{payload.get('shift_end', '')})"])
    meta.append(['Scope', payload.get('scope_label', '')])
    meta.append(['Machines shown', payload.get('machine_count', 0)])
    meta.append(['Total fleet', payload.get('total_fleet', 0)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/machine-ct")
def get_machine_ct(
    machine_id: int = Query(...),
    entry_date: date = Query(...),
    shift: str = Query(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Lightweight endpoint: returns the active cycle time for a machine on a date/shift."""
    cfg = _load_config(db)
    hourly_cfg = cfg.get('hourly_output') or {}
    plans = db.query(ProductionPlan).filter(
        ProductionPlan.machine_id == machine_id,
        ProductionPlan.plan_date == entry_date,
        ProductionPlan.shift == shift,
        ProductionPlan.status.in_(['running', 'completed', 'pending']),
    ).order_by(ProductionPlan.status.desc(), ProductionPlan.priority).all()
    ct = _plan_ct(plans[0]) if plans else 0.0
    pt = float(plans[0].process_time or 0) if plans else 0.0
    return {
        "cycle_time_sec": ct,
        "process_time_sec": pt,
        "running_part_threshold_pct": hourly_cfg.get('running_part_threshold_pct', 30),
        "ld_unld_max_sec": hourly_cfg.get('ld_unld_max_sec', 60),
        "micro_gap_sec": hourly_cfg.get('micro_gap_sec', 15),
    }


@router.get("/")
def get_hourly_output(
    entry_date: date = Query(...),
    shift: str = Query(...),
    scope: str = Query('all'),
    station_id: Optional[int] = None,
    line_id: Optional[str] = None,
    factory_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    try:
        return build_hourly_output(db, entry_date, shift, scope, station_id, line_id, factory_id)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@router.get("/download-xlsx")
def download_hourly_output_xlsx(
    entry_date: date = Query(...),
    shift: str = Query(...),
    scope: str = Query('all'),
    station_id: Optional[int] = None,
    line_id: Optional[str] = None,
    factory_id: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    payload = build_hourly_output(db, entry_date, shift, scope, station_id, line_id, factory_id)
    content = _build_xlsx(payload)
    fname = f"hourly_output_{entry_date}_{shift}.xlsx"
    return Response(
        content=content,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )
