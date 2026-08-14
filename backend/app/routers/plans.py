from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import extract, func
from datetime import date, datetime, timedelta
from typing import Optional, List
from pydantic import BaseModel
import io, smtplib, os
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from ..models import ProductionPlan, Machine, ModelChangeRequest, EmailLog, get_db, now_ist, WorkOrder
from ..auth import get_current_user, require_role
from ..ws_manager import manager
from .work_orders import sync_work_order_after_plan_change

router = APIRouter(prefix="/api/plans", tags=["plans"])

class PlanCreate(BaseModel):
    plan_date: date
    end_date: Optional[date] = None   # if set, creates plans for each day in range
    shift: str
    shifts: Optional[List[str]] = None  # if set, creates one plan per day per shift
    station_no: int
    machine_id: Optional[int] = None
    work_order_id: Optional[int] = None
    current_operation: str
    next_operation: str = ""
    model_variant: Optional[str] = None
    process_time: float
    loading_unloading: float = 10
    planned_qty: int
    priority: int = 1
    plan_type: str = "scheduled"
    notes: Optional[str] = None
    tool_shortage_ack: bool = False  # planner ack when forecast requires it

class PlanUpdate(BaseModel):
    planned_qty: Optional[int] = None
    priority: Optional[int] = None
    plan_type: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None

class ActualQtyUpdate(BaseModel):
    actual_qty: int
    source: str = "manual"  # manual | modbus | opcua | mqtt

class RescheduleRequest(BaseModel):
    new_date: Optional[date] = None
    new_shift: Optional[str] = None
    new_machine_id: Optional[int] = None
    new_station_no: Optional[int] = None
    mode: str = "custom"  # custom | next_week | split_remaining
    days_offset: Optional[int] = None
    split_remaining: bool = False

class BulkRescheduleRequest(BaseModel):
    plan_ids: List[int]
    mode: str = "next_week"  # next_week | custom
    days_offset: int = 7
    new_start_date: Optional[date] = None
    split_remaining: bool = True

def _resolve_shifts(data: PlanCreate) -> List[str]:
    if data.shifts:
        return list(dict.fromkeys(data.shifts))
    return [data.shift]

def _plan_copy_fields(plan: ProductionPlan) -> dict:
    return {
        "work_order_id": plan.work_order_id,
        "station_no": plan.station_no,
        "machine_id": plan.machine_id,
        "current_operation": plan.current_operation,
        "next_operation": plan.next_operation,
        "model_variant": plan.model_variant,
        "process_time": plan.process_time,
        "loading_unloading": plan.loading_unloading,
        "priority": plan.priority,
        "plan_type": plan.plan_type,
        "created_by": plan.created_by,
    }

def _compute_new_date(plan: ProductionPlan, data: RescheduleRequest) -> date:
    if data.mode == "next_week":
        return plan.plan_date + timedelta(days=7)
    if data.new_date:
        return data.new_date
    if data.days_offset is not None:
        return plan.plan_date + timedelta(days=data.days_offset)
    raise HTTPException(400, "Provide new_date, days_offset, or use mode=next_week")

def _validate_move_date(new_date: date) -> None:
    today = now_ist().date()
    if new_date < today:
        raise HTTPException(
            400,
            f"Cannot move plan to a past date ({new_date}). Earliest allowed date is {today}.",
        )

def _validate_may_complete_plan(plan: ProductionPlan) -> None:
    today = now_ist().date()
    if plan.plan_date > today:
        raise HTTPException(
            400,
            f"Cannot complete plan scheduled for {plan.plan_date}. "
            f"Completion is only allowed on or after the plan date (today: {today}).",
        )

def _validate_may_start_plan(plan: ProductionPlan) -> None:
    today = now_ist().date()
    if plan.plan_date > today:
        raise HTTPException(
            400,
            f"Cannot start plan scheduled for {plan.plan_date}. "
            f"Start is only allowed on or after the plan date (today: {today}).",
        )


def _finalize_plan_at_shift_end(plan: ProductionPlan):
    """Close a running/paused plan when its shift window has ended.

    Full target met → completed; otherwise → incomplete (production incomplete).
    """
    actual = int(plan.actual_qty or 0)
    planned = int(plan.planned_qty or 0)
    if planned > 0 and actual >= planned:
        plan.status = "completed"
    else:
        plan.status = "incomplete"
    plan.updated_at = now_ist()


def _heal_shift_ended_plans(db: Session) -> int:
    """Mark running/paused plans whose shift has ended as completed or incomplete."""
    today = now_ist().date()
    now = now_ist()
    candidates = (
        db.query(ProductionPlan)
        .filter(
            ProductionPlan.status.in_(("running", "paused")),
            ProductionPlan.plan_date <= today,
        )
        .all()
    )
    if not candidates:
        return 0

    try:
        from .hourly_output import _load_config, _shift_window
        cfg = _load_config(db)
        shifts = (cfg or {}).get("shifts") or {}
    except Exception:
        shifts = {}

    changed = 0
    touched_wo = set()
    for plan in candidates:
        sh_def = shifts.get(plan.shift)
        if sh_def:
            try:
                _start, shift_end = _shift_window(plan.plan_date, sh_def)
                ended = now >= shift_end
            except Exception:
                ended = plan.plan_date < today
        else:
            ended = plan.plan_date < today
        if not ended:
            continue
        _finalize_plan_at_shift_end(plan)
        if plan.work_order_id:
            touched_wo.add(plan.work_order_id)
        changed += 1

    if changed:
        for wo_id in touched_wo:
            sync_work_order_after_plan_change(db, wo_id)
        try:
            db.commit()
        except Exception:
            db.rollback()
            return 0
    return changed


TERMINAL_PLAN_STATUSES = frozenset({"completed", "cancelled", "aborted", "incomplete"})


def _find_running_conflicts(db: Session, plan: ProductionPlan) -> list:
    """Other running plans on the same machine (trial plans may run concurrently)."""
    if not plan.machine_id or plan.plan_type == "trial":
        return []
    return (
        db.query(ProductionPlan)
        .filter(
            ProductionPlan.machine_id == plan.machine_id,
            ProductionPlan.id != plan.id,
            ProductionPlan.status == "running",
        )
        .all()
    )


def _raise_if_part_already_running(db: Session, plan: ProductionPlan) -> None:
    """Block starting a new plan while another part is already running on the machine."""
    conflicts = _find_running_conflicts(db, plan)
    if not conflicts:
        return
    names = []
    for c in conflicts:
        label = (c.model_variant or c.current_operation or f"Plan #{c.id}").strip()
        names.append(label)
    parts = ", ".join(names)
    raise HTTPException(
        status_code=409,
        detail={
            "code": "part_already_running",
            "message": (
                "A part is already running on this machine. "
                "Please pause or complete the planned production before starting the new one."
                + (f" (Running: {parts})" if parts else "")
            ),
            "conflicting_plan_ids": [c.id for c in conflicts],
            "conflicting_parts": names,
        },
    )

def _should_split(plan: ProductionPlan, data: RescheduleRequest) -> bool:
    if data.mode == "split_remaining" or data.split_remaining:
        return plan.actual_qty > 0 and plan.actual_qty < plan.planned_qty
    return plan.status == "paused" and plan.actual_qty > 0 and plan.actual_qty < plan.planned_qty

def _apply_machine_relocation(plan: ProductionPlan, data: RescheduleRequest, db: Session) -> Optional[str]:
    if data.new_machine_id is None:
        return None
    if data.new_machine_id == plan.machine_id:
        return None
    machine = db.query(Machine).filter(Machine.id == data.new_machine_id).first()
    if not machine:
        raise HTTPException(404, "Target machine not found")
    if machine.status == "breakdown":
        raise HTTPException(400, f"Cannot move plan to {machine.name} — machine is in breakdown")
    old_m = db.query(Machine).filter(Machine.id == plan.machine_id).first() if plan.machine_id else None
    old_name = old_m.name if old_m else "unassigned"
    plan.machine_id = data.new_machine_id
    plan.station_no = data.new_station_no or machine.station_id
    return f" [Machine relocated: {old_name} → {machine.name}]"

@router.post("/")
async def create_plan(data: PlanCreate, db: Session = Depends(get_db),
                      user=Depends(get_current_user)):
    start = data.plan_date
    end   = data.end_date if data.end_date and data.end_date >= start else start
    delta = (end - start).days + 1
    shifts = _resolve_shifts(data)

    if data.work_order_id:
        wo = db.query(WorkOrder).filter(WorkOrder.id == data.work_order_id).first()
        if not wo:
            raise HTTPException(404, "Work order not found")
        if wo.status == "cancelled":
            raise HTTPException(400, "Cannot plan against a cancelled work order")
        existing_planned = db.query(func.coalesce(func.sum(ProductionPlan.planned_qty), 0)).filter(
            ProductionPlan.work_order_id == wo.id,
            ProductionPlan.status != "cancelled",
        ).scalar() or 0
        total_new = data.planned_qty * delta * len(shifts)
        if existing_planned + total_new > wo.target_qty:
            remaining = max(wo.target_qty - existing_planned, 0)
            slot_count = delta * len(shifts)
            max_per_slot = remaining // slot_count if slot_count else remaining
            raise HTTPException(
                400,
                f"Total planned qty ({total_new} pcs = {data.planned_qty} per shift × {slot_count} slots) "
                f"exceeds work order capacity ({remaining} pcs left to plan). "
                f"Max per shift for this date/shift range: {max_per_slot} pcs.",
            )

        # Tool stock / life forecast — notify planner; require ack if short / near EOL
        from ..tool_service import build_forecast, log_event, find_tool
        forecast = build_forecast(db, work_order_id=data.work_order_id, planned_qty=total_new)
        if forecast.get("blocks_plan"):
            raise HTTPException(status_code=409, detail={
                "code": "tool_eol_blocked",
                "message": forecast.get("message"),
                "forecast": forecast,
            })
        if forecast.get("requires_ack") and not data.tool_shortage_ack:
            raise HTTPException(status_code=409, detail={
                "code": "tool_forecast_ack_required",
                "message": forecast.get("message"),
                "forecast": forecast,
            })
        if data.tool_shortage_ack and forecast.get("requires_ack"):
            from ..models import ToolStock
            for trow in forecast.get("tools") or []:
                if not trow.get("requires_ack"):
                    continue
                tool = None
                if trow.get("tool_id"):
                    tool = db.query(ToolStock).filter(ToolStock.id == trow["tool_id"]).first()
                if not tool:
                    tool = find_tool(db, tool_code=trow.get("tool_code"), tool_name=trow.get("tool_name"))
                if tool:
                    log_event(
                        db, tool, "forecast_ack",
                        user_id=user.id,
                        work_order_id=data.work_order_id,
                        notes=f"Planner acknowledged forecast: {trow.get('message')}",
                        acknowledged_by=user.id,
                    )

    created = []
    base = data.model_dump(exclude={"end_date", "shifts", "tool_shortage_ack"})
    for i in range(delta):
        for shift in shifts:
            day_data = {**base, "plan_date": start + timedelta(days=i), "shift": shift}
            plan = ProductionPlan(**day_data, created_by=user.id, created_at=now_ist())
            db.add(plan)
            db.flush()
            created.append(plan)
    if data.work_order_id:
        sync_work_order_after_plan_change(db, data.work_order_id)
    db.commit()
    for plan in created:
        db.refresh(plan)
        await manager.broadcast({"type": "plan_created", "plan_id": plan.id,
                                  "station_no": plan.station_no, "shift": plan.shift,
                                  "current_operation": plan.current_operation, "next_operation": plan.next_operation,
                                  "planned_qty": plan.planned_qty})
    return created if delta > 1 else created[0]

@router.get("/")
def get_plans(
    plan_date: Optional[date] = None,
    shift: Optional[str] = None,
    station_no: Optional[int] = None,
    week: Optional[int] = None,
    week_start: Optional[date] = None,
    week_end: Optional[date] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    q = db.query(ProductionPlan)
    if plan_date:  q = q.filter(ProductionPlan.plan_date == plan_date)
    if shift:      q = q.filter(ProductionPlan.shift == shift)
    if station_no:    q = q.filter(ProductionPlan.station_no == station_no)
    if month:      q = q.filter(extract("month", ProductionPlan.plan_date) == month)
    if year:       q = q.filter(extract("year",  ProductionPlan.plan_date) == year)
    if week:       q = q.filter(extract("week",  ProductionPlan.plan_date) == week)
    if week_start: q = q.filter(ProductionPlan.plan_date >= week_start)
    if week_end:   q = q.filter(ProductionPlan.plan_date <= week_end)
    if date_from:  q = q.filter(ProductionPlan.plan_date >= date_from)
    if date_to:    q = q.filter(ProductionPlan.plan_date <= date_to)
    if status:     q = q.filter(ProductionPlan.status == status)

    # Close shift-ended running/paused plans (completed or production incomplete)
    try:
        _heal_shift_ended_plans(db)
    except Exception as exc:
        print(f"[Plans] shift-end heal skipped: {exc}")

    plans = q.order_by(ProductionPlan.plan_date, ProductionPlan.shift, ProductionPlan.priority).all()

    # Keep plan.actual_qty aligned with dashboard running-part count
    try:
        from .hourly_output import sync_plan_actuals_from_status_logs
        sync_kwargs = {"commit": True}
        if plan_date:
            sync_kwargs["entry_date"] = plan_date
        else:
            if date_from or week_start:
                sync_kwargs["date_from"] = date_from or week_start
            if date_to or week_end:
                sync_kwargs["date_to"] = date_to or week_end
        if shift:
            sync_kwargs["shift"] = shift
        sync_plan_actuals_from_status_logs(db, **sync_kwargs)
        # Re-query so response reflects synced actuals
        plans = q.order_by(ProductionPlan.plan_date, ProductionPlan.shift, ProductionPlan.priority).all()
    except Exception as exc:
        print(f"[Plans] actual sync skipped: {exc}")

    return [_plan_dict(p, db) for p in plans]

@router.get("/summary")
def get_summary(
    plan_date: Optional[date] = None,
    shift: Optional[str] = None,
    station_no: Optional[int] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    week_start: Optional[date] = None,
    week_end: Optional[date] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    q = db.query(ProductionPlan)
    if plan_date:  q = q.filter(ProductionPlan.plan_date == plan_date)
    if shift:      q = q.filter(ProductionPlan.shift == shift)
    if station_no:    q = q.filter(ProductionPlan.station_no == station_no)
    if month:      q = q.filter(extract("month", ProductionPlan.plan_date) == month)
    if year:       q = q.filter(extract("year",  ProductionPlan.plan_date) == year)
    if week_start: q = q.filter(ProductionPlan.plan_date >= week_start)
    if week_end:   q = q.filter(ProductionPlan.plan_date <= week_end)
    if date_from:  q = q.filter(ProductionPlan.plan_date >= date_from)
    if date_to:    q = q.filter(ProductionPlan.plan_date <= date_to)

    try:
        _heal_shift_ended_plans(db)
    except Exception as exc:
        print(f"[Plans] summary shift-end heal skipped: {exc}")

    # Sync running-part counts into plan.actual_qty so tiles match dashboard
    try:
        from .hourly_output import sync_plan_actuals_from_status_logs
        sync_kwargs = {"commit": True}
        if plan_date:
            sync_kwargs["entry_date"] = plan_date
        else:
            if date_from or week_start:
                sync_kwargs["date_from"] = date_from or week_start
            if date_to or week_end:
                sync_kwargs["date_to"] = date_to or week_end
        if shift:
            sync_kwargs["shift"] = shift
        sync_plan_actuals_from_status_logs(db, **sync_kwargs)
    except Exception as exc:
        print(f"[Plans] summary actual sync skipped: {exc}")

    plans = q.all()
    total_planned = sum(p.planned_qty for p in plans)
    total_actual  = sum(p.actual_qty  for p in plans)
    return {
        "total_plans": len(plans),
        "total_planned": total_planned,
        "total_actual": total_actual,
        "achievement_pct": round(total_actual / total_planned * 100, 1) if total_planned else 0,
        "by_status": {s: sum(1 for p in plans if p.status == s)
                      for s in ["pending","running","completed","paused","cancelled","aborted","incomplete"]},
        "by_shift": {sh: {"planned": sum(p.planned_qty for p in plans if p.shift == sh),
                           "actual":  sum(p.actual_qty  for p in plans if p.shift == sh)}
                     for sh in ["A","B"]}
    }

def _plan_dict(plan: ProductionPlan, db: Session) -> dict:
    """Serialize plan with linked model-change interlock state."""
    d = {c.name: getattr(plan, c.name) for c in plan.__table__.columns}
    for key in ("plan_date", "created_at", "updated_at"):
        val = d.get(key)
        if val is not None and hasattr(val, "isoformat"):
            d[key] = val.isoformat()
    for key in ("process_time", "loading_unloading"):
        if d.get(key) is not None:
            d[key] = float(d[key])
    pending = (
        db.query(ModelChangeRequest)
        .filter(
            ModelChangeRequest.plan_id == plan.id,
            ModelChangeRequest.status == "pending",
        )
        .order_by(ModelChangeRequest.id.desc())
        .first()
    )
    active = (
        db.query(ModelChangeRequest)
        .filter(
            ModelChangeRequest.plan_id == plan.id,
            ModelChangeRequest.status.in_(["approved", "in_progress"]),
        )
        .order_by(ModelChangeRequest.id.desc())
        .first()
    )
    mcr = pending or active
    d["model_change_request_id"] = mcr.id if mcr else None
    d["model_change_status"] = mcr.status if mcr else None
    d["awaiting_model_change"] = bool(pending)
    return d


def _same_part_in_previous_shift(
    db: Session,
    machine_id: int,
    model_variant: str,
    plan_date: date,
    shift_id: str,
) -> bool:
    """Check if the same model_variant was running/completed on this machine
    in the immediately preceding shift (including cross-day C→A)."""
    from .hourly_output import _load_config, _previous_shift_info

    cfg = _load_config(db)
    prev_shift_id, prev_date = _previous_shift_info(cfg, shift_id, plan_date)
    if not prev_shift_id or not prev_date:
        return False

    prev_plan = (
        db.query(ProductionPlan)
        .filter(
            ProductionPlan.machine_id == machine_id,
            ProductionPlan.plan_date == prev_date,
            ProductionPlan.shift == prev_shift_id,
            ProductionPlan.model_variant == model_variant,
            ProductionPlan.status.in_(["running", "completed", "paused"]),
        )
        .first()
    )
    return prev_plan is not None


def _previous_model_on_machine(db: Session, machine_id: int, exclude_plan_id: int) -> str:
    prev = (
        db.query(ProductionPlan)
        .filter(
            ProductionPlan.machine_id == machine_id,
            ProductionPlan.id != exclude_plan_id,
            ProductionPlan.status.in_(["running", "completed", "paused"]),
        )
        .order_by(ProductionPlan.updated_at.desc(), ProductionPlan.id.desc())
        .first()
    )
    if prev and prev.model_variant:
        return prev.model_variant
    return "—"


def _norm_model(value: Optional[str]) -> str:
    text = (value or "").strip()
    if not text or text == "—":
        return ""
    return text.casefold()


def _last_active_plan_on_machine(db: Session, machine_id: int):
    """Most recent running plan, else the latest paused/completed/running plan on the machine."""
    running = (
        db.query(ProductionPlan)
        .filter(
            ProductionPlan.machine_id == machine_id,
            ProductionPlan.status == "running",
        )
        .order_by(ProductionPlan.updated_at.desc(), ProductionPlan.id.desc())
        .first()
    )
    if running:
        return running
    return (
        db.query(ProductionPlan)
        .filter(
            ProductionPlan.machine_id == machine_id,
            ProductionPlan.status.in_(["running", "paused", "completed"]),
        )
        .order_by(ProductionPlan.updated_at.desc(), ProductionPlan.id.desc())
        .first()
    )


def _is_same_model_as_last_on_machine(db: Session, plan: ProductionPlan) -> bool:
    """True when starting/resuming the same part that last ran on this machine.

    Pause/resume of the same plan is not a model change. Switching to a different
    part/variant always requires a model-change request, even if that part ran earlier.
    """
    if not plan.machine_id:
        return False
    last = _last_active_plan_on_machine(db, plan.machine_id)
    if not last:
        return False
    if last.id == plan.id:
        return True
    this_model = _norm_model(plan.model_variant)
    return bool(this_model) and this_model == _norm_model(last.model_variant)


def _ensure_machine_on_plan(plan: ProductionPlan, db: Session) -> None:
    if plan.machine_id:
        return
    machine = (
        db.query(Machine)
        .filter(Machine.station_id == plan.station_no)
        .order_by(Machine.id)
        .first()
    )
    if not machine:
        raise HTTPException(
            400,
            "No machine found for this station — assign a machine on the plan before starting",
        )
    plan.machine_id = machine.id


def _plan_started_event(plan: ProductionPlan, plan_id: int) -> dict:
    return {
        "type": "plan_started",
        "plan_id": plan_id,
        "machine_id": plan.machine_id,
        "station_no": plan.station_no,
        "current_operation": plan.current_operation,
        "next_operation": plan.next_operation,
        "model_variant": plan.model_variant,
        "process_time": float(plan.process_time) if plan.process_time is not None else None,
        "loading_unloading": float(plan.loading_unloading) if plan.loading_unloading is not None else None,
        "shift": plan.shift,
        "plan_date": str(plan.plan_date),
    }


async def _issue_model_change_for_start(plan: ProductionPlan, db: Session, user, plan_id: int):
    """Raise or reuse a model-change request. Plan status is left unchanged until approval."""
    existing_pending = (
        db.query(ModelChangeRequest)
        .filter(
            ModelChangeRequest.plan_id == plan.id,
            ModelChangeRequest.status == "pending",
        )
        .first()
    )
    if existing_pending:
        return {
            **_plan_dict(plan, db),
            "model_change_pending": True,
            "model_change_request_id": existing_pending.id,
            "message": (
                f"Model change request #{existing_pending.id} is awaiting approval "
                "on the Model Change page"
            ),
        }

    existing_active = (
        db.query(ModelChangeRequest)
        .filter(
            ModelChangeRequest.plan_id == plan.id,
            ModelChangeRequest.status.in_(["approved", "in_progress"]),
        )
        .first()
    )
    if existing_active:
        _raise_if_part_already_running(db, plan)
        plan.status = "running"
        plan.updated_at = now_ist()
        db.commit()
        return {
            **_plan_dict(plan, db),
            "message": "Plan is running (model change already approved)",
        }

    from_model = _previous_model_on_machine(db, plan.machine_id, plan.id)
    to_model = (plan.model_variant or plan.current_operation or "—").strip() or "—"
    mcr = ModelChangeRequest(
        machine_id=plan.machine_id,
        plan_id=plan.id,
        requested_by=user.id,
        from_model=from_model,
        to_model=to_model,
        ideal_minutes=60,
        shift=plan.shift or "A",
        entry_date=plan.plan_date or now_ist().date(),
        reason="setting_change",
        status="pending",
        created_at=now_ist(),
    )
    db.add(mcr)
    plan.updated_at = now_ist()
    db.commit()
    db.refresh(mcr)
    await manager.broadcast({
        "type": "model_change_request",
        "id": mcr.id,
        "machine_id": plan.machine_id,
        "plan_id": plan.id,
        "from_model": from_model,
        "to_model": to_model,
        "status": "pending",
    })
    await manager.broadcast({"type": "plan_updated", "plan_id": plan_id, "status": plan.status})
    return {
        **_plan_dict(plan, db),
        "message": (
            f"Model change request #{mcr.id} raised ({from_model} → {to_model}). "
            "Approve on Model Change page to start the plan and apply the part on WI."
        ),
        "model_change_pending": True,
        "model_change_request_id": mcr.id,
    }


@router.patch("/{plan_id}/status")
async def update_status(plan_id: int, data: PlanUpdate, db: Session = Depends(get_db),
                        user=Depends(get_current_user)):
    plan = db.query(ProductionPlan).filter(ProductionPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(404, "Plan not found")
    if data.status == "completed":
        _validate_may_complete_plan(plan)
    if data.status == "running" and plan.status != "running":
        _validate_may_start_plan(plan)
    if data.status == "aborted":
        if plan.status not in ("paused", "running"):
            raise HTTPException(400, "Only running or paused plans can be aborted")
        plan.status = "aborted"
        plan.updated_at = now_ist()
        machine = db.query(Machine).filter(Machine.id == plan.machine_id).first() if plan.machine_id else None
        if machine and machine.status == "running":
            # Only idle if no other plan is still running on this machine
            other_running = (
                db.query(ProductionPlan)
                .filter(
                    ProductionPlan.machine_id == plan.machine_id,
                    ProductionPlan.id != plan.id,
                    ProductionPlan.status == "running",
                )
                .first()
            )
            if not other_running:
                machine.status = "idle"
        sync_work_order_after_plan_change(db, plan.work_order_id)
        db.commit()
        await manager.broadcast({"type": "plan_updated", "plan_id": plan_id, "status": "aborted"})
        return _plan_dict(plan, db)

    # ── Interlock: start/resume requires model-change when the part/variant changes ──
    if data.status == "running" and plan.status in ("pending", "paused"):
        _ensure_machine_on_plan(plan, db)
        # Block start while another part is already running (trial may run concurrently)
        _raise_if_part_already_running(db, plan)

        # Same part already on this machine (including pause/resume or previous-shift
        # continuity) does not need a setting change. Any other part/variant does —
        # including switching back to a previously paused model.
        last_active = _last_active_plan_on_machine(db, plan.machine_id)
        if _is_same_model_as_last_on_machine(db, plan):
            plan.status = "running"
            plan.updated_at = now_ist()
            machine = db.query(Machine).filter(Machine.id == plan.machine_id).first()
            if machine and machine.status not in ("breakdown", "offline", "setting_change", "alarm"):
                machine.status = "running"
            db.commit()
            await manager.broadcast(_plan_started_event(plan, plan_id))
            resumed = last_active is not None and last_active.id == plan.id
            return {
                **_plan_dict(plan, db),
                "message": (
                    "Same part resumed without model change"
                    if resumed else
                    "Same part continues — started without model change"
                ),
            }

        return await _issue_model_change_for_start(plan, db, user, plan_id)

    if data.status:
        plan.status = data.status
    if data.planned_qty is not None:
        plan.planned_qty = data.planned_qty
    if data.priority is not None:
        plan.priority = data.priority
    if data.plan_type:
        plan.plan_type = data.plan_type
    if data.notes is not None:
        plan.notes = data.notes
    plan.updated_at = now_ist()

    # Resume from paused is gated above; this path is a same-status refresh
    if data.status == "running":
        # Block if another part is already running (trial plans may run concurrently)
        _raise_if_part_already_running(db, plan)

        machine = db.query(Machine).filter(Machine.id == plan.machine_id).first()
        if machine and machine.status not in ("breakdown", "offline", "setting_change", "alarm"):
            machine.status = "running"
        db.commit()
        await manager.broadcast({
            "type": "plan_started", "plan_id": plan_id,
            "machine_id": plan.machine_id, "station_no": plan.station_no,
            "current_operation": plan.current_operation, "next_operation": plan.next_operation,
            "model_variant": plan.model_variant,
            "process_time": float(plan.process_time) if plan.process_time is not None else None,
            "loading_unloading": float(plan.loading_unloading) if plan.loading_unloading is not None else None,
            "shift": plan.shift, "plan_date": str(plan.plan_date),
        })
    elif data.status == "completed":
        machine = db.query(Machine).filter(Machine.id == plan.machine_id).first()
        if machine:
            machine.status = "idle"
        sync_work_order_after_plan_change(db, plan.work_order_id)
        db.commit()
        await manager.broadcast({"type": "plan_completed", "plan_id": plan_id,
                                  "machine_id": plan.machine_id, "station_no": plan.station_no})
    elif data.status == "incomplete":
        machine = db.query(Machine).filter(Machine.id == plan.machine_id).first()
        if machine and machine.status == "running":
            other_running = (
                db.query(ProductionPlan)
                .filter(
                    ProductionPlan.machine_id == plan.machine_id,
                    ProductionPlan.id != plan.id,
                    ProductionPlan.status == "running",
                )
                .first()
            )
            if not other_running:
                machine.status = "idle"
        sync_work_order_after_plan_change(db, plan.work_order_id)
        db.commit()
        await manager.broadcast({"type": "plan_updated", "plan_id": plan_id, "status": "incomplete"})
    else:
        db.commit()
        await manager.broadcast({"type": "plan_updated", "plan_id": plan_id, "status": data.status})
    return _plan_dict(plan, db)

@router.patch("/{plan_id}/actual")
async def update_actual_qty(plan_id: int, data: ActualQtyUpdate, db: Session = Depends(get_db),
                             user=Depends(get_current_user)):
    """Called by manual entry, MQTT bridge, Modbus bridge, or OPC-UA bridge"""
    plan = db.query(ProductionPlan).filter(ProductionPlan.id == plan_id).first()
    if not plan: raise HTTPException(404, "Plan not found")
    prev_actual = int(plan.actual_qty or 0)
    plan.actual_qty = data.actual_qty
    plan.updated_at = now_ist()
    if data.actual_qty >= plan.planned_qty and plan.status == "running":
        _validate_may_complete_plan(plan)
        plan.status = "completed"
        machine = db.query(Machine).filter(Machine.id == plan.machine_id).first()
        if machine: machine.status = "idle"
    try:
        from ..tool_service import apply_consumption_safe
        apply_consumption_safe(db, plan, prev_actual, data.actual_qty, user_id=getattr(user, "id", None))
    except Exception as exc:
        print(f"[WARN] tool consumption skipped: {exc}")
    sync_work_order_after_plan_change(db, plan.work_order_id)
    db.commit()
    await manager.broadcast({
        "type": "actual_qty_updated", "plan_id": plan_id,
        "actual_qty": data.actual_qty, "planned_qty": plan.planned_qty,
        "source": data.source, "station_no": plan.station_no
    })
    return plan

@router.delete("/{plan_id}")
async def delete_plan(plan_id: int, db: Session = Depends(get_db),
                      user=Depends(require_role("supervisor", "admin"))):
    plan = db.query(ProductionPlan).filter(ProductionPlan.id == plan_id).first()
    if not plan: raise HTTPException(404, "Plan not found")
    db.delete(plan)
    db.commit()
    await manager.broadcast({"type": "plan_deleted", "plan_id": plan_id})
    return {"ok": True}


@router.post("/{plan_id}/reschedule")
async def reschedule_plan(
    plan_id: int,
    data: RescheduleRequest,
    db: Session = Depends(get_db),
    user=Depends(require_role("supervisor", "admin")),
):
    plan = db.query(ProductionPlan).filter(ProductionPlan.id == plan_id).first()
    if not plan:
        raise HTTPException(404, "Plan not found")
    if plan.status in TERMINAL_PLAN_STATUSES:
        raise HTTPException(400, f"Cannot reschedule a {plan.status} plan")
    if plan.status == "running":
        raise HTTPException(400, "Pause the plan before moving it")

    new_date = _compute_new_date(plan, data)
    _validate_move_date(new_date)
    new_shift = data.new_shift or plan.shift
    old_date = plan.plan_date
    created = None

    if _should_split(plan, data):
        remaining = plan.planned_qty - plan.actual_qty
        _validate_may_complete_plan(plan)
        plan.planned_qty = plan.actual_qty
        plan.status = "completed"
        suffix = f" [Partial: {remaining} pcs moved to {new_date} shift {new_shift}]"
        plan.notes = ((plan.notes or "") + suffix).strip()
        plan.updated_at = now_ist()

        created = ProductionPlan(
            **_plan_copy_fields(plan),
            plan_date=new_date,
            shift=new_shift,
            planned_qty=remaining,
            actual_qty=0,
            status="pending",
            notes=f"Moved from plan #{plan_id} ({old_date} shift {plan.shift})",
            created_at=now_ist(),
            updated_at=now_ist(),
        )
        machine_note = _apply_machine_relocation(created, data, db)
        if machine_note:
            created.notes = (created.notes + machine_note).strip()
        db.add(created)
    else:
        old_shift = plan.shift
        plan.plan_date = new_date
        plan.shift = new_shift
        if plan.status == "paused":
            plan.status = "pending"
        suffix = f" [Rescheduled from {old_date} shift {old_shift}]"
        plan.notes = ((plan.notes or "") + suffix).strip()
        machine_note = _apply_machine_relocation(plan, data, db)
        if machine_note:
            plan.notes = (plan.notes + machine_note).strip()
        plan.updated_at = now_ist()

    if plan.work_order_id:
        sync_work_order_after_plan_change(db, plan.work_order_id)
    db.commit()
    db.refresh(plan)
    if created:
        db.refresh(created)

    await manager.broadcast({
        "type": "plan_rescheduled",
        "plan_id": plan_id,
        "new_plan_id": created.id if created else None,
        "new_date": str(new_date),
    })
    return {
        "original": plan,
        "new_plan": created,
        "split": created is not None,
    }


@router.post("/bulk-reschedule")
async def bulk_reschedule(
    data: BulkRescheduleRequest,
    db: Session = Depends(get_db),
    user=Depends(require_role("supervisor", "admin")),
):
    if not data.plan_ids:
        raise HTTPException(400, "No plans selected")

    plans = db.query(ProductionPlan).filter(ProductionPlan.id.in_(data.plan_ids)).all()
    if len(plans) != len(data.plan_ids):
        raise HTTPException(404, "One or more plans not found")

    movable = [p for p in plans if p.status in ("pending", "paused")]
    if not movable:
        raise HTTPException(400, "No pending or paused plans to move")

    if data.new_start_date:
        anchor = min(p.plan_date for p in movable)
        offset_days = (data.new_start_date - anchor).days
    elif data.mode == "next_week":
        offset_days = data.days_offset if data.days_offset else 7
    else:
        offset_days = data.days_offset or 0

    today = now_ist().date()
    if data.new_start_date and data.new_start_date < today:
        raise HTTPException(
            400,
            f"Start date cannot be in the past ({data.new_start_date}). Earliest allowed date is {today}.",
        )

    invalid_moves = []
    for plan in movable:
        candidate = plan.plan_date + timedelta(days=offset_days)
        if candidate < today:
            invalid_moves.append(plan.id)
    if invalid_moves:
        raise HTTPException(
            400,
            f"Cannot move plan(s) {invalid_moves} to a past date. Earliest allowed date is {today}.",
        )

    results = []
    work_order_ids = set()

    for plan in movable:
        if plan.status == "running":
            continue
        req = RescheduleRequest(
            mode="custom",
            days_offset=offset_days,
            split_remaining=data.split_remaining,
        )
        new_date = plan.plan_date + timedelta(days=offset_days)
        new_shift = plan.shift
        old_date = plan.plan_date
        created = None

        if _should_split(plan, req):
            remaining = plan.planned_qty - plan.actual_qty
            _validate_may_complete_plan(plan)
            plan.planned_qty = plan.actual_qty
            plan.status = "completed"
            plan.notes = ((plan.notes or "") + f" [Bulk move: {remaining} pcs → {new_date}]").strip()
            plan.updated_at = now_ist()
            created = ProductionPlan(
                **_plan_copy_fields(plan),
                plan_date=new_date,
                shift=new_shift,
                planned_qty=remaining,
                actual_qty=0,
                status="pending",
                notes=f"Bulk moved from plan #{plan.id} ({old_date})",
                created_at=now_ist(),
                updated_at=now_ist(),
            )
            db.add(created)
        else:
            plan.plan_date = new_date
            if plan.status == "paused":
                plan.status = "pending"
            plan.notes = ((plan.notes or "") + f" [Bulk moved from {old_date}]").strip()
            plan.updated_at = now_ist()

        if plan.work_order_id:
            work_order_ids.add(plan.work_order_id)
        results.append({"plan_id": plan.id, "new_date": str(new_date), "split": created is not None})

    for wo_id in work_order_ids:
        sync_work_order_after_plan_change(db, wo_id)
    db.commit()

    await manager.broadcast({"type": "plans_bulk_rescheduled", "count": len(results)})
    return {"moved": len(results), "results": results, "days_offset": offset_days}

@router.get("/pipeline/{station_no}")
def get_pipeline(station_no: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Returns ordered queue for a machine — next part to load is first"""
    plans = db.query(ProductionPlan).filter(
        ProductionPlan.station_no == station_no,
        ProductionPlan.status.in_(["pending", "running"])
    ).order_by(ProductionPlan.plan_date, ProductionPlan.shift, ProductionPlan.priority).all()
    return [_plan_dict(p, db) for p in plans]


def _style_header_row(ws, ncols):
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

def _build_excel(plans, work_orders: Optional[dict] = None) -> io.BytesIO:
    work_orders = work_orders or {}
    wb = openpyxl.Workbook()
    now = datetime.now()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    total_planned = sum(p.planned_qty for p in plans)
    total_actual  = sum(p.actual_qty  for p in plans)
    achievement   = round(total_actual / total_planned * 100, 1) if total_planned else 0
    by_status = {s: sum(1 for p in plans if p.status == s)
                 for s in ["pending", "running", "completed", "paused", "cancelled", "aborted", "incomplete"]}
    summary_rows = [
        ["PRODUCTION PLANNING REPORT"],
        [""],
        ["Report Date:", now.strftime("%d-%m-%Y")],
        ["Period:", f"{plans[0].plan_date} to {plans[-1].plan_date}" if len(plans) > 1 else (str(plans[0].plan_date) if plans else "")],
        [""],
        ["SUMMARY METRICS"],
        ["Total Plans:",    len(plans)],
        ["Planned Quantity:", total_planned],
        ["Actual Quantity:",  total_actual],
        ["Achievement %:",    str(achievement) + "%"],
        [""],
        ["STATUS BREAKDOWN"],
        ["Pending:",   by_status["pending"]],
        ["Running:",   by_status["running"]],
        ["Completed:", by_status["completed"]],
        ["Paused:",    by_status["paused"]],
        ["Aborted:",   by_status["aborted"]],
        ["Production Incomplete:", by_status["incomplete"]],
        ["Cancelled:", by_status["cancelled"]],
    ]
    for row in summary_rows:
        ws1.append(row)
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 18

    # ── Sheet 2: Plans Details ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Plans Details")
    headers2 = ["Date","Shift","Station","Work Order No","WO Target Qty","Current Operation","Next Operation","Process Time (s)",
                "Loading/Unloading (s)","Cycle Time (s)","Type","Priority",
                "Planned Qty","Actual Qty","Achievement %","Status","Notes"]
    ws2.append(headers2)
    _style_header_row(ws2, len(headers2))
    for p in plans:
        pct = round(p.actual_qty / p.planned_qty * 100, 1) if p.planned_qty else 0
        wo = work_orders.get(p.work_order_id) if p.work_order_id else None
        ws2.append([str(p.plan_date), p.shift, getattr(p, '_station_label', p.station_no),
                    wo.work_order_no if wo else "", wo.target_qty if wo else "",
                    p.current_operation, p.next_operation,
                    p.process_time, p.loading_unloading, p.process_time + p.loading_unloading,
                    p.plan_type, p.priority, p.planned_qty, p.actual_qty,
                    str(pct) + "%", p.status, p.notes or ""])
    col_widths2 = [12,8,6,14,12,14,14,16,20,14,12,10,12,12,14,12,20]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    # ── Sheet 3: By Pair ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Station")
    headers3 = ["Station No","Total Plans","Planned Qty","Actual Qty",
                "Achievement %","Completed","Running","Pending"]
    ws3.append(headers3)
    _style_header_row(ws3, len(headers3))
    station_stats = {}
    for p in plans:
        key = getattr(p, '_station_label', str(p.station_no))
        ps = station_stats.setdefault(key, {"planned": 0, "actual": 0, "total": 0,
                                         "completed": 0, "running": 0, "pending": 0})
        ps["total"] += 1; ps["planned"] += p.planned_qty; ps["actual"] += p.actual_qty
        if p.status in ps: ps[p.status] += 1
    for pn, ps in sorted(station_stats.items(), key=lambda x: str(x[0])):
        pct = round(ps["actual"] / ps["planned"] * 100, 1) if ps["planned"] else 0
        ws3.append([pn, ps["total"], ps["planned"], ps["actual"],
                    str(pct) + "%", ps["completed"], ps["running"], ps["pending"]])
    for i, w in enumerate([10,12,12,12,14,12,12,12], 1):
        ws3.column_dimensions[ws3.cell(row=1, column=i).column_letter].width = w

    # ── Sheet 4: By Shift ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("By Shift")
    headers4 = ["Shift","Total Plans","Planned Qty","Actual Qty",
                "Achievement %","Completed","Running"]
    ws4.append(headers4)
    _style_header_row(ws4, len(headers4))
    shift_stats = {}
    for p in plans:
        ss = shift_stats.setdefault(p.shift, {"planned": 0, "actual": 0, "total": 0,
                                              "completed": 0, "running": 0})
        ss["total"] += 1; ss["planned"] += p.planned_qty; ss["actual"] += p.actual_qty
        if p.status in ss: ss[p.status] += 1
    for sh, ss in sorted(shift_stats.items()):
        pct = round(ss["actual"] / ss["planned"] * 100, 1) if ss["planned"] else 0
        ws4.append([sh, ss["total"], ss["planned"], ss["actual"],
                    str(pct) + "%", ss["completed"], ss["running"]])
    for i, w in enumerate([10,12,12,12,14,12,12], 1):
        ws4.column_dimensions[ws4.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ExportParams(BaseModel):
    plan_date: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    month: Optional[int] = None
    year: Optional[int] = None
    shift: Optional[str] = None
    station_no: Optional[int] = None


class EmailReportRequest(ExportParams):
    recipients: List[str]
    subject: Optional[str] = "Production Planning Report"


def _filter_plans(q, params: ExportParams):
    if params.plan_date: q = q.filter(ProductionPlan.plan_date == params.plan_date)
    if params.date_from: q = q.filter(ProductionPlan.plan_date >= params.date_from)
    if params.date_to:   q = q.filter(ProductionPlan.plan_date <= params.date_to)
    if params.shift:     q = q.filter(ProductionPlan.shift == params.shift)
    if params.station_no:   q = q.filter(ProductionPlan.station_no == params.station_no)
    if params.month:     q = q.filter(extract("month", ProductionPlan.plan_date) == params.month)
    if params.year:      q = q.filter(extract("year",  ProductionPlan.plan_date) == params.year)
    return q


def _work_orders_for_plans(db: Session, plans) -> dict:
    wo_ids = {p.work_order_id for p in plans if p.work_order_id}
    if not wo_ids:
        return {}
    rows = db.query(WorkOrder).filter(WorkOrder.id.in_(wo_ids)).all()
    return {w.id: w for w in rows}


@router.post("/export")
def export_excel(params: ExportParams, db: Session = Depends(get_db),
                 _=Depends(get_current_user)):
    q = _filter_plans(db.query(ProductionPlan), params)
    plans = q.order_by(ProductionPlan.plan_date, ProductionPlan.shift, ProductionPlan.priority).all()
    buf = _build_excel(plans, _work_orders_for_plans(db, plans))
    filename = f"production_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/email-logs")
def get_email_logs(db: Session = Depends(get_db), _=Depends(get_current_user)):
    logs = db.query(EmailLog).order_by(EmailLog.sent_at.desc()).limit(100).all()
    return [{"id": l.id, "sent_at": l.sent_at, "recipients": l.recipients,
             "subject": l.subject, "report_type": l.report_type,
             "status": l.status, "error_msg": l.error_msg} for l in logs]


@router.post("/email-report")
def email_report(req: EmailReportRequest, db: Session = Depends(get_db),
                 user=Depends(get_current_user)):
    q = _filter_plans(db.query(ProductionPlan), req)
    plans = q.order_by(ProductionPlan.plan_date, ProductionPlan.shift, ProductionPlan.priority).all()
    buf = _build_excel(plans, _work_orders_for_plans(db, plans))
    filename = f"production_plan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    from ..models import EmailSmtpConfig
    cfg = db.query(EmailSmtpConfig).first()
    smtp_host = cfg.smtp_server if cfg else os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = cfg.smtp_port   if cfg else int(os.getenv("SMTP_PORT", 587))
    smtp_user = cfg.email_address  if cfg else os.getenv("SMTP_USER", "")
    smtp_pass = cfg.email_password if cfg else os.getenv("SMTP_PASS", "")

    if not smtp_user or not smtp_pass:
        raise HTTPException(400, "SMTP credentials not configured")

    msg = MIMEMultipart()
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(req.recipients)
    msg["Subject"] = req.subject
    msg.attach(MIMEText(f"Please find attached the production planning report generated on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}.", "plain"))
    part = MIMEBase("application", "octet-stream")
    part.set_payload(buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={filename}")
    msg.attach(part)

    log = EmailLog(sent_at=now_ist(), recipients=", ".join(req.recipients),
                   subject=req.subject, report_type="planning",
                   sent_by=user.id if user else None)
    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, req.recipients, msg.as_string())
        log.status = "sent"
    except Exception as e:
        log.status = "failed"
        log.error_msg = str(e)
        db.add(log)
        db.commit()
        raise HTTPException(500, f"Email send failed: {str(e)}")
    db.add(log)
    db.commit()
    return {"ok": True, "sent_to": req.recipients}
