"""Work order management — master orders, plan linkage, track record, machine suggestions."""
import json
import io
from datetime import date, timedelta, datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import extract, func, or_, and_
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ..auth import get_current_user, require_capability
from ..models import WorkOrder, ProductionPlan, Machine, Part, ToolEvent, GsapSync, get_db, now_ist
from ..ws_manager import manager

router = APIRouter(prefix="/api/work-orders", tags=["work-orders"])


class SpareToolItem(BaseModel):
    name: str
    qty: Optional[float] = None
    unit: Optional[str] = "pcs"
    notes: Optional[str] = None
    tool_no: Optional[str] = None
    stock_available: Optional[float] = None
    remaining_qty: Optional[float] = None
    source: Optional[str] = None  # "part" | "manual"


class WorkOrderCreate(BaseModel):
    work_order_no: str
    part_id: Optional[int] = None
    part_source: Optional[str] = "part_master"  # part_master | gsap
    gsap_sync_id: Optional[int] = None
    model_variant: Optional[str] = None
    description: Optional[str] = None
    target_qty: int = Field(gt=0)
    start_date: date
    end_date: date
    spares_tools: Optional[List[SpareToolItem]] = None
    # Closed WOs whose remaining qty is clubbed / consumed into this new order
    source_wo_ids: Optional[List[int]] = None


class WorkOrderUpdate(BaseModel):
    work_order_no: Optional[str] = None
    part_id: Optional[int] = None
    part_source: Optional[str] = None
    gsap_sync_id: Optional[int] = None
    model_variant: Optional[str] = None
    description: Optional[str] = None
    target_qty: Optional[int] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    status: Optional[str] = None
    spares_tools: Optional[List[SpareToolItem]] = None


class OutstandingDiscardBody(BaseModel):
    wo_ids: List[int] = Field(min_length=1)


class WorkOrderExportParams(BaseModel):
    date_from: Optional[date] = None
    date_to: Optional[date] = None
    status: Optional[str] = None
    search: Optional[str] = None
    historic: bool = False


def _parse_spares(raw: Optional[str]) -> list:
    if not raw:
        return []
    try:
        data = json.loads(raw)
        return data if isinstance(data, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def _resolve_wo_part_fields(
    db: Session,
    *,
    part_source: Optional[str],
    part_id: Optional[int],
    gsap_sync_id: Optional[int],
    model_variant: Optional[str],
    description: Optional[str],
) -> dict:
    src = (part_source or "part_master").strip().lower()
    if src not in ("part_master", "gsap"):
        raise HTTPException(400, "part_source must be part_master or gsap")
    if src == "part_master":
        if not part_id:
            raise HTTPException(400, "Select a part from Part Master")
        part = db.query(Part).filter(Part.id == part_id).first()
        if not part:
            raise HTTPException(404, "Part not found")
        return {
            "part_source": "part_master",
            "part_id": part.id,
            "gsap_sync_id": None,
            "model_variant": model_variant or part.model_variant,
            "description": description,
        }
    if not gsap_sync_id:
        raise HTTPException(400, "Select a GSAP material / operation row")
    gsap = db.query(GsapSync).filter(GsapSync.id == gsap_sync_id).first()
    if not gsap:
        raise HTTPException(404, "GSAP row not found")
    desc = description or gsap.op_short_text
    return {
        "part_source": "gsap",
        "part_id": None,
        "gsap_sync_id": gsap.id,
        "model_variant": model_variant or gsap.material,
        "description": desc,
    }


def _wo_stats(db: Session, wo: WorkOrder) -> dict:
    plans = db.query(ProductionPlan).filter(
        ProductionPlan.work_order_id == wo.id,
        ProductionPlan.status != "cancelled",
    ).all()
    planned_qty = sum(p.planned_qty for p in plans)
    completed_qty = sum(p.actual_qty for p in plans)
    remaining_qty = max(wo.target_qty - completed_qty, 0)
    unplanned_qty = max(wo.target_qty - planned_qty, 0)
    pct = round(completed_qty / wo.target_qty * 100, 1) if wo.target_qty else 0
    return {
        "planned_qty": planned_qty,
        "completed_qty": completed_qty,
        "remaining_qty": remaining_qty,
        "unplanned_qty": unplanned_qty,
        "complete_pct": pct,
        "plan_count": len(plans),
    }


def _effective_end_date(db: Session, wo: WorkOrder) -> Optional[date]:
    """WO end_date, else latest linked plan date (so overdue still closes without end_date)."""
    if wo.end_date:
        return wo.end_date
    row = (
        db.query(func.max(ProductionPlan.plan_date))
        .filter(
            ProductionPlan.work_order_id == wo.id,
            ProductionPlan.status != "cancelled",
        )
        .scalar()
    )
    return row


def _mark_closed_outstanding(wo: WorkOrder, remaining: int):
    wo.status = "closed"
    wo.outstanding_qty = max(int(remaining), 0)
    # Do not revive discarded/consumed leftovers when syncing again
    if (wo.outstanding_status or "none") in ("none", "available", None):
        wo.outstanding_status = "available" if wo.outstanding_qty > 0 else "none"


def _sync_wo_status(db: Session, wo_id: int):
    wo = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not wo or wo.status == "cancelled":
        return
    stats = _wo_stats(db, wo)
    today = date.today()
    end = _effective_end_date(db, wo)
    remaining = stats["remaining_qty"]

    if stats["completed_qty"] >= wo.target_qty:
        wo.status = "completed"
        wo.outstanding_qty = 0
        if (wo.outstanding_status or "none") == "available":
            wo.outstanding_status = "none"
    elif end and end < today and remaining > 0:
        # Planned window passed with unfinished qty → closed + outstanding
        if (wo.outstanding_status or "none") in ("consumed", "discarded"):
            wo.status = "closed"
            wo.outstanding_qty = remaining
        else:
            _mark_closed_outstanding(wo, remaining)
    elif wo.status == "closed" and (wo.outstanding_status or "none") in ("consumed", "discarded"):
        # Keep closed once disposition decided
        wo.outstanding_qty = remaining
    elif stats["plan_count"] > 0 or stats["completed_qty"] > 0:
        wo.status = "in_progress"
        if (wo.outstanding_status or "none") == "available":
            wo.outstanding_qty = 0
            wo.outstanding_status = "none"
    elif wo.status == "completed":
        wo.status = "in_progress"
    wo.updated_at = now_ist()


def _close_overdue_for_part(db: Session, part_id: Optional[int], except_wo_id: Optional[int] = None):
    """When a new WO starts for a part, close any overdue incomplete WOs for that part."""
    if not part_id:
        return
    today = date.today()
    candidates = (
        db.query(WorkOrder)
        .filter(
            WorkOrder.part_id == part_id,
            WorkOrder.status.in_(("draft", "in_progress")),
        )
        .all()
    )
    for wo in candidates:
        if except_wo_id and wo.id == except_wo_id:
            continue
        end = _effective_end_date(db, wo)
        if not end or end >= today:
            continue
        stats = _wo_stats(db, wo)
        if stats["remaining_qty"] > 0:
            _mark_closed_outstanding(wo, stats["remaining_qty"])
            wo.updated_at = now_ist()


def _serialize_wo(db: Session, wo: WorkOrder, include_plans: bool = False) -> dict:
    part = db.query(Part).filter(Part.id == wo.part_id).first() if wo.part_id else None
    gsap = db.query(GsapSync).filter(GsapSync.id == wo.gsap_sync_id).first() if wo.gsap_sync_id else None
    stats = _wo_stats(db, wo)
    out_status = wo.outstanding_status or "none"
    out_qty = int(wo.outstanding_qty or 0)
    if out_status == "available" and out_qty <= 0:
        out_qty = stats["remaining_qty"]
    status_label = (
        f"Closed with leftover qty ({out_qty})"
        if wo.status == "closed" and out_qty > 0
        else wo.status.replace("_", " ").title()
    )
    out = {
        "id": wo.id,
        "work_order_no": wo.work_order_no,
        "part_id": wo.part_id,
        "part_source": wo.part_source or "part_master",
        "gsap_sync_id": wo.gsap_sync_id,
        "part_no": part.part_no if part else (gsap.material if gsap else None),
        "gsap_material": gsap.material if gsap else None,
        "gsap_operation": gsap.operation if gsap else None,
        "gsap_work_centre": gsap.work_centre if gsap else None,
        "gsap_op_short_text": gsap.op_short_text if gsap else None,
        "gsap_setup_time": gsap.setup_time if gsap else None,
        "gsap_machine_time": gsap.machine_time if gsap else None,
        "model_variant": wo.model_variant or (part.model_variant if part else (gsap.material if gsap else None)),
        "description": wo.description or (gsap.op_short_text if gsap else None),
        "target_qty": wo.target_qty,
        "start_date": str(wo.start_date) if wo.start_date else None,
        "end_date": str(wo.end_date) if wo.end_date else None,
        "status": wo.status,
        "status_label": status_label,
        "spares_tools": _parse_spares(wo.spares_tools_json),
        "outstanding_qty": out_qty,
        "outstanding_status": out_status,
        "consumed_by_wo_id": wo.consumed_by_wo_id,
        "created_at": wo.created_at,
        "updated_at": wo.updated_at,
        **stats,
    }
    if include_plans:
        plans = db.query(ProductionPlan).filter(
            ProductionPlan.work_order_id == wo.id
        ).order_by(ProductionPlan.plan_date, ProductionPlan.shift).all()
        machines = {m.id: m for m in db.query(Machine).all()}
        out["plans"] = [_serialize_plan(p, machines) for p in plans]
    return out


def _serialize_plan(p: ProductionPlan, machines: dict) -> dict:
    m = machines.get(p.machine_id)
    return {
        "id": p.id,
        "work_order_id": p.work_order_id,
        "plan_date": str(p.plan_date),
        "shift": p.shift,
        "station_no": p.station_no,
        "machine_id": p.machine_id,
        "machine_name": m.name if m else None,
        "current_operation": p.current_operation,
        "next_operation": p.next_operation,
        "model_variant": p.model_variant,
        "planned_qty": p.planned_qty,
        "actual_qty": p.actual_qty,
        "status": p.status,
        "plan_type": p.plan_type,
        "notes": p.notes,
    }


def _apply_wo_date_range(q, date_from: Optional[date], date_to: Optional[date]):
    """Include work orders whose schedule overlaps the requested date window."""
    if date_from:
        q = q.filter(or_(WorkOrder.end_date.is_(None), WorkOrder.end_date >= date_from))
    if date_to:
        q = q.filter(or_(WorkOrder.start_date.is_(None), WorkOrder.start_date <= date_to))
    return q


def _month_end(d: date) -> date:
    if d.month == 12:
        return date(d.year, 12, 31)
    return date(d.year, d.month + 1, 1) - timedelta(days=1)


def _apply_active_in_range(q, db: Session, date_from: date, date_to: date):
    """Active WOs with production plans in range, or whose WO schedule overlaps the range."""
    plan_wo_ids = [
        row[0] for row in db.query(ProductionPlan.work_order_id).filter(
            ProductionPlan.work_order_id.isnot(None),
            ProductionPlan.plan_date >= date_from,
            ProductionPlan.plan_date <= date_to,
            ProductionPlan.status != "cancelled",
        ).distinct().all()
    ]
    overlap = and_(
        or_(WorkOrder.end_date.is_(None), WorkOrder.end_date >= date_from),
        or_(WorkOrder.start_date.is_(None), WorkOrder.start_date <= date_to),
    )
    if plan_wo_ids:
        return q.filter(or_(WorkOrder.id.in_(plan_wo_ids), overlap))
    return q.filter(overlap)


def _consume_source_wos(db: Session, source_ids: List[int], new_wo_id: int, part_id: Optional[int]) -> int:
    """Mark selected closed/available outstanding WOs as consumed. Returns clubbed qty."""
    if not source_ids:
        return 0
    rows = (
        db.query(WorkOrder)
        .filter(WorkOrder.id.in_(source_ids))
        .all()
    )
    total = 0
    for src in rows:
        if part_id and src.part_id and src.part_id != part_id:
            raise HTTPException(
                400,
                f"Source work order {src.work_order_no} belongs to a different part",
            )
        # Ensure overdue incomplete rows are closed first
        _sync_wo_status(db, src.id)
        db.refresh(src)
        status = src.outstanding_status or "none"
        qty = int(src.outstanding_qty or 0)
        if status != "available" or qty <= 0:
            # Allow consuming freshly closed remaining even if status sync lagged
            stats = _wo_stats(db, src)
            if src.status == "closed" and stats["remaining_qty"] > 0 and status in ("none", "available"):
                qty = stats["remaining_qty"]
            else:
                raise HTTPException(
                    400,
                    f"Work order {src.work_order_no} has no available outstanding qty",
                )
        src.status = "closed"
        src.outstanding_qty = qty
        src.outstanding_status = "consumed"
        src.consumed_by_wo_id = new_wo_id
        src.updated_at = now_ist()
        total += qty
    return total


@router.post("/")
async def create_work_order(
    data: WorkOrderCreate,
    db: Session = Depends(get_db),
    user=Depends(require_capability("capability.edit_work_orders", "supervisor", "admin")),
):
    if db.query(WorkOrder).filter(WorkOrder.work_order_no == data.work_order_no).first():
        raise HTTPException(400, "Work order number already exists")
    part_fields = _resolve_wo_part_fields(
        db,
        part_source=data.part_source,
        part_id=data.part_id,
        gsap_sync_id=data.gsap_sync_id,
        model_variant=data.model_variant,
        description=data.description,
    )
    # New work orders require both dates and may only be scheduled from today onward.
    today = now_ist().date()
    if not data.start_date:
        raise HTTPException(400, "Period Start is required")
    if not data.end_date:
        raise HTTPException(400, "Period End is required")
    if data.start_date < today:
        raise HTTPException(400, "Period Start cannot be before today")
    if data.end_date < today:
        raise HTTPException(400, "Period End cannot be before today")
    if data.end_date < data.start_date:
        raise HTTPException(400, "Period End cannot be before Period Start")
    spares_json = json.dumps([s.model_dump() for s in data.spares_tools]) if data.spares_tools else None
    wo = WorkOrder(
        work_order_no=data.work_order_no.strip(),
        part_id=part_fields["part_id"],
        part_source=part_fields["part_source"],
        gsap_sync_id=part_fields["gsap_sync_id"],
        model_variant=part_fields["model_variant"],
        description=part_fields["description"],
        target_qty=data.target_qty,
        start_date=data.start_date,
        end_date=data.end_date,
        spares_tools_json=spares_json,
        status="draft",
        outstanding_qty=0,
        outstanding_status="none",
        created_by=user.id,
        created_at=now_ist(),
        updated_at=now_ist(),
    )
    db.add(wo)
    db.flush()

    clubbed = 0
    if data.source_wo_ids:
        clubbed = _consume_source_wos(db, data.source_wo_ids, wo.id, part_fields["part_id"])
        if clubbed > 0:
            note = f"Clubbed outstanding qty {clubbed}."
            combined = f"{wo.description} {note}".strip() if wo.description else note
            wo.description = combined[:255]

    # New order for this part → close any other overdue incomplete WOs
    _close_overdue_for_part(db, part_fields["part_id"], except_wo_id=wo.id)

    db.commit()
    db.refresh(wo)
    await manager.broadcast({
        "type": "work_order_created",
        "work_order_id": wo.id,
        "clubbed_outstanding_qty": clubbed,
    })
    out = _serialize_wo(db, wo)
    out["clubbed_outstanding_qty"] = clubbed
    return out


@router.get("/outstanding")
def list_outstanding_work_orders(
    part_id: Optional[int] = None,
    model_variant: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Available outstanding (closed leftover) work orders, optionally filtered by part."""
    # Heal overdue rows first so UI sees fresh outstanding
    q_heal = db.query(WorkOrder).filter(WorkOrder.status.in_(("draft", "in_progress", "closed")))
    if part_id:
        q_heal = q_heal.filter(WorkOrder.part_id == part_id)
    for wo in q_heal.all():
        _sync_wo_status(db, wo.id)
    db.commit()

    q = db.query(WorkOrder).filter(
        WorkOrder.status == "closed",
        WorkOrder.outstanding_status == "available",
        WorkOrder.outstanding_qty > 0,
    )
    if part_id:
        q = q.filter(WorkOrder.part_id == part_id)
    if model_variant:
        q = q.filter(WorkOrder.model_variant.ilike(f"%{model_variant}%"))
    rows = q.order_by(WorkOrder.end_date.asc(), WorkOrder.id.asc()).all()
    items = [_serialize_wo(db, wo) for wo in rows]
    return {
        "items": items,
        "count": len(items),
        "total_outstanding_qty": sum(i.get("outstanding_qty") or 0 for i in items),
    }


@router.post("/outstanding/discard")
def discard_outstanding(
    data: OutstandingDiscardBody,
    db: Session = Depends(get_db),
    _=Depends(require_capability("capability.edit_work_orders", "admin", "superadmin", "supervisor")),
):
    """Discard leftover qty so it is no longer offered for reuse/clubbing."""
    discarded = []
    for wo_id in data.wo_ids:
        wo = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
        if not wo:
            raise HTTPException(404, f"Work order {wo_id} not found")
        _sync_wo_status(db, wo.id)
        db.refresh(wo)
        if wo.status != "closed" and (wo.outstanding_status or "none") != "available":
            # Still allow discard of overdue incomplete by closing first
            stats = _wo_stats(db, wo)
            end = _effective_end_date(db, wo)
            if end and end < date.today() and stats["remaining_qty"] > 0:
                _mark_closed_outstanding(wo, stats["remaining_qty"])
            else:
                raise HTTPException(400, f"{wo.work_order_no} has no outstanding qty to discard")
        wo.status = "closed"
        wo.outstanding_status = "discarded"
        wo.updated_at = now_ist()
        discarded.append({
            "id": wo.id,
            "work_order_no": wo.work_order_no,
            "outstanding_qty": wo.outstanding_qty or 0,
        })
    db.commit()
    return {"ok": True, "discarded": discarded, "count": len(discarded)}


@router.get("/")
def list_work_orders(
    search: Optional[str] = None,
    status: Optional[str] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    active_only: bool = False,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(WorkOrder)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(
            WorkOrder.work_order_no.ilike(like),
            WorkOrder.model_variant.ilike(like),
            WorkOrder.description.ilike(like),
        ))
    if status:
        q = q.filter(WorkOrder.status == status)
    if month:
        q = q.filter(extract("month", WorkOrder.start_date) == month)
    if year:
        q = q.filter(extract("year", WorkOrder.start_date) == year)
    if active_only and date_from and date_to:
        q = _apply_active_in_range(q, db, date_from, date_to)
    else:
        q = _apply_wo_date_range(q, date_from, date_to)
    orders = q.order_by(WorkOrder.start_date.desc(), WorkOrder.id.desc()).all()

    # Sync plan actuals from machine running count before rolling up WO completed qty
    try:
        from .hourly_output import sync_plan_actuals_from_status_logs
        sync_plan_actuals_from_status_logs(
            db,
            date_from=date_from,
            date_to=date_to,
            commit=True,
        )
    except Exception as exc:
        print(f"[WorkOrders] actual sync skipped: {exc}")

    for wo in orders:
        _sync_wo_status(db, wo.id)
    db.commit()
    if active_only:
        orders = [wo for wo in orders if wo.status in ("draft", "in_progress")]
    return [_serialize_wo(db, wo) for wo in orders]


@router.get("/overview")
def work_order_overview(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    search: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Overview list with gantt bar data per work order."""
    today = date.today()
    if not date_from:
        date_from = today - timedelta(days=7)
    if not date_to:
        date_to = today + timedelta(days=14)

    q = db.query(WorkOrder)
    if search:
        like = f"%{search}%"
        q = q.filter(or_(
            WorkOrder.work_order_no.ilike(like),
            WorkOrder.model_variant.ilike(like),
        ))
    if status:
        q = q.filter(WorkOrder.status == status)
    q = q.filter(or_(
        WorkOrder.start_date.is_(None),
        WorkOrder.start_date <= date_to,
    )).filter(or_(
        WorkOrder.end_date.is_(None),
        WorkOrder.end_date >= date_from,
    ))
    orders = q.order_by(WorkOrder.start_date, WorkOrder.id).all()
    machines = {m.id: m for m in db.query(Machine).all()}

    items = []
    for wo in orders:
        _sync_wo_status(db, wo.id)
        stats = _wo_stats(db, wo)
        plans = db.query(ProductionPlan).filter(
            ProductionPlan.work_order_id == wo.id,
            ProductionPlan.status != "cancelled",
        ).all()
        plan_dates = [p.plan_date for p in plans]
        bar_start = min(plan_dates) if plan_dates else wo.start_date
        bar_end = max(plan_dates) if plan_dates else wo.end_date
        if not bar_start:
            bar_start = date_from
        if not bar_end:
            bar_end = date_to

        out_qty = int(wo.outstanding_qty or 0)
        if (wo.outstanding_status or "none") == "available" and out_qty <= 0:
            out_qty = stats["remaining_qty"]
        delay = (
            wo.status == "closed"
            or (wo.end_date and wo.end_date < today and stats["completed_qty"] < wo.target_qty)
        )
        gantt_status = "completed" if wo.status == "completed" else (
            "delay" if delay else ("running" if stats["completed_qty"] > 0 else "schedule")
        )

        segments = []
        for p in plans:
            seg_end = p.plan_date
            pct = round(p.actual_qty / p.planned_qty * 100) if p.planned_qty else 0
            seg_status = p.status
            if p.plan_date < today and p.status not in ("completed", "cancelled"):
                seg_status = "delay"
            segments.append({
                "plan_id": p.id,
                "start": str(p.plan_date),
                "end": str(seg_end),
                "planned_qty": p.planned_qty,
                "actual_qty": p.actual_qty,
                "machine_name": machines.get(p.machine_id).name if p.machine_id and machines.get(p.machine_id) else None,
                "status": seg_status,
                "complete_pct": pct,
            })

        part = db.query(Part).filter(Part.id == wo.part_id).first() if wo.part_id else None
        items.append({
            "id": wo.id,
            "work_order_no": wo.work_order_no,
            "part_label": wo.model_variant or (part.part_no if part else "—"),
            "target_qty": wo.target_qty,
            "completed_qty": stats["completed_qty"],
            "remaining_qty": stats["remaining_qty"],
            "outstanding_qty": out_qty,
            "complete_pct": stats["complete_pct"],
            "status": wo.status,
            "status_label": (
                f"Closed with leftover qty ({out_qty})"
                if wo.status == "closed" and out_qty > 0
                else wo.status.replace("_", " ").title()
            ),
            "gantt_status": gantt_status,
            "bar_start": str(bar_start),
            "bar_end": str(bar_end),
            "segments": segments,
            "plans": [_serialize_plan(p, machines) for p in plans],
        })

    db.commit()
    return {
        "date_from": str(date_from),
        "date_to": str(date_to),
        "today": str(today),
        "items": items,
    }


@router.get("/planned")
def list_planned_work_orders(
    search: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Work orders with upcoming production plans or a future WO schedule in the date window."""
    today = now_ist().date()
    plan_from = date_from or today.replace(day=1)
    plan_to = date_to or _month_end(today)

    plans_q = db.query(ProductionPlan).filter(
        ProductionPlan.work_order_id.isnot(None),
        ProductionPlan.plan_date >= plan_from,
        ProductionPlan.plan_date <= plan_to,
        ProductionPlan.status.in_(["pending", "paused"]),
    )
    if search:
        like = f"%{search}%"
        plans_q = plans_q.join(WorkOrder, WorkOrder.id == ProductionPlan.work_order_id).filter(or_(
            WorkOrder.work_order_no.ilike(like),
            WorkOrder.model_variant.ilike(like),
            WorkOrder.description.ilike(like),
        ))

    plans = plans_q.order_by(ProductionPlan.plan_date, ProductionPlan.shift).all()
    machines = {m.id: m for m in db.query(Machine).all()}

    by_wo: dict = {}
    for p in plans:
        by_wo.setdefault(p.work_order_id, []).append(p)

    items = []
    seen_ids = set()

    for wo_id, wo_plans in by_wo.items():
        wo = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
        if not wo or wo.status in ("completed", "cancelled"):
            continue
        _sync_wo_status(db, wo.id)
        if wo.status in ("completed", "cancelled", "closed"):
            continue
        stats = _wo_stats(db, wo)
        if stats["completed_qty"] >= wo.target_qty:
            continue
        future_plans = [_serialize_plan(p, machines) for p in wo_plans]
        future_qty = sum(p.planned_qty - p.actual_qty for p in wo_plans)
        items.append({
            **_serialize_wo(db, wo),
            "future_plans": future_plans,
            "future_plan_count": len(future_plans),
            "future_planned_qty": future_qty,
            "next_plan_date": future_plans[0]["plan_date"] if future_plans else None,
            "unplanned_qty": stats["unplanned_qty"],
            "schedule_only": False,
        })
        seen_ids.add(wo_id)

    # Future-scheduled work orders with no production plans yet (WO period in range, start in future)
    wo_q = db.query(WorkOrder).filter(
        WorkOrder.status.notin_(["completed", "cancelled"]),
        WorkOrder.start_date.isnot(None),
        WorkOrder.start_date > today,
        WorkOrder.start_date <= plan_to,
        or_(WorkOrder.end_date.is_(None), WorkOrder.end_date >= plan_from),
    )
    if search:
        like = f"%{search}%"
        wo_q = wo_q.filter(or_(
            WorkOrder.work_order_no.ilike(like),
            WorkOrder.model_variant.ilike(like),
            WorkOrder.description.ilike(like),
        ))

    for wo in wo_q.order_by(WorkOrder.start_date, WorkOrder.id).all():
        if wo.id in seen_ids:
            continue
        _sync_wo_status(db, wo.id)
        if wo.status in ("completed", "cancelled", "closed"):
            continue
        stats = _wo_stats(db, wo)
        if stats["completed_qty"] >= wo.target_qty:
            continue
        items.append({
            **_serialize_wo(db, wo),
            "future_plans": [],
            "future_plan_count": 0,
            "future_planned_qty": 0,
            "next_plan_date": str(wo.start_date),
            "unplanned_qty": stats["unplanned_qty"],
            "schedule_only": True,
        })
        seen_ids.add(wo.id)

    db.commit()
    items.sort(key=lambda x: (x["next_plan_date"] or "", x["work_order_no"]))
    return {
        "today": str(today),
        "date_from": str(plan_from),
        "date_to": str(plan_to),
        "items": items,
    }


@router.get("/suggest-machines")
def suggest_machines(
    part_id: Optional[int] = None,
    model_variant: Optional[str] = None,
    smart: bool = Query(True, description="Only suggest machines near completion or idle"),
    limit: int = Query(5, ge=1, le=20),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Suggest machines based on historical runs for the same part."""
    if not part_id and not model_variant:
        return {"suggestions": [], "smart_mode": smart}

    part = db.query(Part).filter(Part.id == part_id).first() if part_id else None
    variant = model_variant or (part.model_variant if part else None)

    hist_q = db.query(ProductionPlan).filter(
        ProductionPlan.machine_id.isnot(None),
        ProductionPlan.status.in_(["completed", "running", "paused"]),
    )
    if variant:
        hist_q = hist_q.filter(ProductionPlan.model_variant == variant)
    elif part:
        hist_q = hist_q.filter(ProductionPlan.model_variant == part.part_no)

    historical = hist_q.order_by(ProductionPlan.plan_date.desc()).all()
    machine_ids = []
    seen = set()
    for p in historical:
        if p.machine_id not in seen:
            seen.add(p.machine_id)
            machine_ids.append(p.machine_id)

    if not smart:
        machines = {m.id: m for m in db.query(Machine).filter(Machine.id.in_(machine_ids)).all()}
        suggestions = []
        for mid in machine_ids[:limit]:
            m = machines.get(mid)
            if not m:
                continue
            last = next((p for p in historical if p.machine_id == mid), None)
            suggestions.append({
                "machine_id": mid,
                "machine_name": m.name,
                "station_id": m.station_id,
                "machine_status": m.status,
                "reason": "Previously used for this part",
                "last_run_date": str(last.plan_date) if last else None,
                "total_qty_produced": sum(p.actual_qty for p in historical if p.machine_id == mid),
            })
        return {"suggestions": suggestions, "smart_mode": False}

    machines = {m.id: m for m in db.query(Machine).all()}
    suggestions = []
    completion_threshold = 0.15  # within 15% of plan completion

    for mid in machine_ids:
        m = machines.get(mid)
        if not m:
            continue
        current = db.query(ProductionPlan).filter(
            ProductionPlan.machine_id == mid,
            ProductionPlan.status.in_(["pending", "running", "paused"]),
        ).order_by(ProductionPlan.priority, ProductionPlan.plan_date).first()

        reason = "Previously used for this part"
        include = False

        if m.status in ("idle", "offline") and not current:
            include = True
            reason = "Machine idle — available for this part"
        elif current:
            remaining = max(current.planned_qty - current.actual_qty, 0)
            ratio = remaining / current.planned_qty if current.planned_qty else 1
            if ratio <= completion_threshold or current.status == "completed":
                include = True
                reason = f"Current job ~{100 - int(ratio * 100)}% complete"
            elif m.status == "idle":
                include = True
                reason = "Machine idle after recent run"
        elif m.status == "running":
            continue

        if not include:
            continue

        last = next((p for p in historical if p.machine_id == mid), None)
        suggestions.append({
            "machine_id": mid,
            "machine_name": m.name,
            "station_id": m.station_id,
            "machine_status": m.status,
            "reason": reason,
            "last_run_date": str(last.plan_date) if last else None,
            "total_qty_produced": sum(p.actual_qty for p in historical if p.machine_id == mid),
            "current_plan_remaining": (
                max(current.planned_qty - current.actual_qty, 0) if current else 0
            ),
        })

    suggestions.sort(key=lambda s: (
        0 if s["machine_status"] == "idle" else 1,
        -(s.get("total_qty_produced") or 0),
    ))
    return {"suggestions": suggestions[:limit], "smart_mode": True}


def _filter_work_orders(q, params: WorkOrderExportParams):
    if params.search:
        like = f"%{params.search}%"
        q = q.filter(or_(
            WorkOrder.work_order_no.ilike(like),
            WorkOrder.model_variant.ilike(like),
            WorkOrder.description.ilike(like),
        ))
    if params.status:
        q = q.filter(WorkOrder.status == params.status)
    elif params.historic:
        q = q.filter(WorkOrder.status.in_(["completed", "cancelled"]))
    else:
        q = q.filter(WorkOrder.status.in_(["draft", "in_progress"]))
    q = _apply_wo_date_range(q, params.date_from, params.date_to)
    return q


def _style_header_row(ws, ncols):
    hdr_fill = PatternFill("solid", fgColor="1E3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")


def _build_work_order_excel(db: Session, orders: list, report_label: str) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["WORK ORDER REPORT", report_label])
    ws1.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws1.append([])
    ws1.append(["Total Work Orders", len(orders)])
    ws1.append(["Total Target Qty", sum(o.target_qty for o in orders)])
    ws1.append(["Total Completed Qty", sum(_wo_stats(db, o)["completed_qty"] for o in orders)])

    ws2 = wb.create_sheet("Work Orders")
    wo_headers = [
        "Work Order No", "Part No", "Model/Variant", "Description", "Target Qty",
        "Planned Qty", "Completed Qty", "Remaining Qty", "Complete %", "Status",
        "Period Start", "Period End", "Spares/Tools", "Created At",
    ]
    ws2.append(wo_headers)
    _style_header_row(ws2, len(wo_headers))
    for wo in orders:
        stats = _wo_stats(db, wo)
        part = db.query(Part).filter(Part.id == wo.part_id).first() if wo.part_id else None
        spares = _parse_spares(wo.spares_tools_json)
        spares_txt = "; ".join(
            f"{s.get('name', '')}{' x' + str(s['qty']) if s.get('qty') else ''}" for s in spares
        )
        ws2.append([
            wo.work_order_no,
            part.part_no if part else "",
            wo.model_variant or (part.model_variant if part else ""),
            wo.description or "",
            wo.target_qty,
            stats["planned_qty"],
            stats["completed_qty"],
            stats["remaining_qty"],
            stats["complete_pct"],
            wo.status,
            str(wo.start_date) if wo.start_date else "",
            str(wo.end_date) if wo.end_date else "",
            spares_txt,
            str(wo.created_at) if wo.created_at else "",
        ])

    ws3 = wb.create_sheet("Production Runs")
    run_headers = [
        "Work Order No", "Plan Date", "Shift", "Station", "Machine", "Operation",
        "Planned Qty", "Actual Qty", "Complete %", "Status", "Notes",
    ]
    ws3.append(run_headers)
    _style_header_row(ws3, len(run_headers))
    machines = {m.id: m for m in db.query(Machine).all()}
    for wo in orders:
        plans = db.query(ProductionPlan).filter(
            ProductionPlan.work_order_id == wo.id,
        ).order_by(ProductionPlan.plan_date, ProductionPlan.shift).all()
        for p in plans:
            m = machines.get(p.machine_id)
            pct = round(p.actual_qty / p.planned_qty * 100, 1) if p.planned_qty else 0
            ws3.append([
                wo.work_order_no,
                str(p.plan_date),
                p.shift,
                p.station_no,
                m.name if m else "",
                p.current_operation,
                p.planned_qty,
                p.actual_qty,
                pct,
                p.status,
                p.notes or "",
            ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.post("/export")
def export_work_orders(
    params: WorkOrderExportParams,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = _filter_work_orders(db.query(WorkOrder), params)
    orders = q.order_by(WorkOrder.start_date.desc(), WorkOrder.id.desc()).all()
    label = "Historic" if params.historic else "Current"
    if params.date_from and params.date_to:
        label += f" ({params.date_from} to {params.date_to})"
    buf = _build_work_order_excel(db, orders, label)
    filename = f"work_orders_{label.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/{wo_id}")
def get_work_order(wo_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    wo = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not wo:
        raise HTTPException(404, "Work order not found")
    try:
        from .hourly_output import sync_plan_actuals_from_status_logs
        sync_plan_actuals_from_status_logs(
            db,
            date_from=wo.start_date,
            date_to=wo.end_date or now_ist().date(),
            commit=True,
        )
    except Exception as exc:
        print(f"[WorkOrders] detail actual sync skipped: {exc}")
    return _serialize_wo(db, wo, include_plans=True)


@router.get("/{wo_id}/track-record")
def track_record(wo_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    wo = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not wo:
        raise HTTPException(404, "Work order not found")
    try:
        from .hourly_output import sync_plan_actuals_from_status_logs
        sync_plan_actuals_from_status_logs(
            db,
            date_from=wo.start_date,
            date_to=wo.end_date or now_ist().date(),
            commit=True,
        )
    except Exception as exc:
        print(f"[WorkOrders] track-record actual sync skipped: {exc}")
    plans = db.query(ProductionPlan).filter(
        ProductionPlan.work_order_id == wo_id,
    ).order_by(ProductionPlan.plan_date.desc(), ProductionPlan.shift).all()
    machines = {m.id: m for m in db.query(Machine).all()}
    stats = _wo_stats(db, wo)

    records = []
    for p in plans:
        m = machines.get(p.machine_id)
        records.append({
            "plan_id": p.id,
            "run_date": str(p.plan_date),
            "shift": p.shift,
            "station_no": p.station_no,
            "machine_id": p.machine_id,
            "machine_name": m.name if m else "—",
            "machine_status": m.status if m else None,
            "planned_qty": p.planned_qty,
            "actual_qty": p.actual_qty,
            "complete_pct": round(p.actual_qty / p.planned_qty * 100, 1) if p.planned_qty else 0,
            "status": p.status,
            "current_operation": p.current_operation,
            "next_operation": p.next_operation,
            "model_variant": p.model_variant,
            "updated_at": p.updated_at,
        })

    return {
        "work_order": _serialize_wo(db, wo),
        "summary": stats,
        "records": records,
    }


@router.patch("/{wo_id}")
async def update_work_order(
    wo_id: int,
    data: WorkOrderUpdate,
    db: Session = Depends(get_db),
    user=Depends(require_capability("capability.edit_work_orders", "supervisor", "admin")),
):
    wo = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not wo:
        raise HTTPException(404, "Work order not found")

    if data.work_order_no is not None:
        new_no = data.work_order_no.strip()
        if not new_no:
            raise HTTPException(400, "Work order number is required")
        clash = (
            db.query(WorkOrder)
            .filter(WorkOrder.work_order_no == new_no, WorkOrder.id != wo_id)
            .first()
        )
        if clash:
            raise HTTPException(400, "Work order number already exists")
        wo.work_order_no = new_no

    part_source_in = data.part_source if data.part_source is not None else wo.part_source
    part_id_in = data.part_id if data.part_id is not None else wo.part_id
    gsap_in = data.gsap_sync_id if data.gsap_sync_id is not None else wo.gsap_sync_id
    if (
        data.part_source is not None
        or data.part_id is not None
        or data.gsap_sync_id is not None
    ):
        part_fields = _resolve_wo_part_fields(
            db,
            part_source=part_source_in,
            part_id=part_id_in,
            gsap_sync_id=gsap_in,
            model_variant=data.model_variant if data.model_variant is not None else wo.model_variant,
            description=data.description if data.description is not None else wo.description,
        )
        wo.part_source = part_fields["part_source"]
        wo.part_id = part_fields["part_id"]
        wo.gsap_sync_id = part_fields["gsap_sync_id"]
        wo.model_variant = part_fields["model_variant"]
        if data.description is None and part_fields["description"]:
            wo.description = part_fields["description"]
    else:
        if data.model_variant is not None:
            wo.model_variant = data.model_variant
        if data.description is not None:
            wo.description = data.description
    if data.target_qty is not None:
        if data.target_qty <= 0:
            raise HTTPException(400, "Target qty must be greater than 0")
        wo.target_qty = data.target_qty

    new_start = data.start_date if data.start_date is not None else wo.start_date
    new_end = data.end_date if data.end_date is not None else wo.end_date
    if new_start is None or new_end is None:
        raise HTTPException(400, "Period Start and Period End are both required")
    if new_end < new_start:
        raise HTTPException(400, "Period End cannot be before Period Start")
    today = now_ist().date()
    # Do not allow moving dates into the past relative to today (existing past dates may be kept).
    if data.start_date is not None and data.start_date < today and data.start_date != wo.start_date:
        raise HTTPException(400, "Period Start cannot be before today")
    if data.end_date is not None and data.end_date < today and data.end_date != wo.end_date:
        raise HTTPException(400, "Period End cannot be before today")
    if data.start_date is not None:
        wo.start_date = data.start_date
    if data.end_date is not None:
        wo.end_date = data.end_date

    if data.status is not None:
        wo.status = data.status
    if data.spares_tools is not None:
        wo.spares_tools_json = json.dumps([s.model_dump() for s in data.spares_tools])
    wo.updated_at = now_ist()
    _sync_wo_status(db, wo_id)
    db.commit()
    db.refresh(wo)
    await manager.broadcast({"type": "work_order_updated", "work_order_id": wo_id})
    return _serialize_wo(db, wo, include_plans=True)


@router.delete("/{wo_id}")
async def delete_work_order(
    wo_id: int,
    db: Session = Depends(get_db),
    user=Depends(require_capability("capability.edit_work_orders", "supervisor", "admin")),
):
    """Delete a work order. Linked plans/tool events are unlinked (not deleted)."""
    wo = db.query(WorkOrder).filter(WorkOrder.id == wo_id).first()
    if not wo:
        raise HTTPException(404, "Work order not found")

    wo_no = wo.work_order_no

    # Restore outstanding WOs that were clubbed into this one
    sources = db.query(WorkOrder).filter(WorkOrder.consumed_by_wo_id == wo_id).all()
    for src in sources:
        src.consumed_by_wo_id = None
        if (src.outstanding_status or "") == "consumed":
            src.outstanding_status = "available"
            src.status = "closed"
        src.updated_at = now_ist()

    # Unlink production plans (keep history; remove WO association)
    db.query(ProductionPlan).filter(ProductionPlan.work_order_id == wo_id).update(
        {ProductionPlan.work_order_id: None},
        synchronize_session=False,
    )
    # Unlink tool events
    db.query(ToolEvent).filter(ToolEvent.work_order_id == wo_id).update(
        {ToolEvent.work_order_id: None},
        synchronize_session=False,
    )

    db.delete(wo)
    db.commit()
    await manager.broadcast({"type": "work_order_deleted", "work_order_id": wo_id})
    return {"ok": True, "deleted_id": wo_id, "work_order_no": wo_no}


def sync_work_order_after_plan_change(db: Session, work_order_id: Optional[int]):
    if work_order_id:
        _sync_wo_status(db, work_order_id)
